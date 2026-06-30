# -*- coding: utf-8 -*-
"""
SV pipeline - Process, parse, priorization and annotation of SV data to
generate Excel reports.

Last modification: April 2026
"""

import os
import logging
import gzip
import pandas as pd
import numpy as np
import pickle
import csv
import sys
import glob
from natsort import natsorted
from intervaltree import IntervalTree
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


# Obtener la raíz del proyecto 'VUSCAN' para permitir imports relativos entre carpetas
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

# Importar la utilidad que genera la hoja 'Execution_summary'
from general_scripts.execution_summary_excel import add_summary_sheet


# -------------------------------------------------------------------------
# CONFIGURACIÓN DE CONSTANTES
# -------------------------------------------------------------------------

# Rutas a archivos
BASE_PATH = '/mnt/c/Users/bioinfor/Desktop/VUSCAN_PIPELINES/DATA'
COSMIC_FILE = f'{BASE_PATH}/cancer_genes/tsg_og.tsv'
HEREDITARY_CANCER_FILE = f'{BASE_PATH}/cancer_genes/lista_CH_genes.txt'
GNOMAD_FILE = f'{BASE_PATH}/gnomad/gnomad.v4.1.sv.non_neuro_controls_final.tsv'
UNIPROT_FILE = f'{BASE_PATH}/pathway_annotation_db/uniprot.txt'
KEGG_FILE = f'{BASE_PATH}/pathway_annotation_db/kegg_pathways.tsv'
REACTOME_FILE = f'{BASE_PATH}/pathway_annotation_db/Gene2Reactome.txt'
SAMPLE_PHENO_FILE = f'{BASE_PATH}/muestras_fenotipo.txt'
GLOSSARY_FILE = f'{BASE_PATH}/glosarios/glosario_sv.xlsx'

# Parametros de filtrado
OVERLAP_THRESHOLD = 0.7   # 70% reciprocal overlap
QUAL_MIN = 30             # Minimum quality QUAL
INTERNAL_FREQ_MAX = 10    # Maximum internal frequency count

# Colores para formatos condicionales de Excel
EXCEL_COLORS = {
    'red': PatternFill(bgColor="FFC7CE"),
    'orange': PatternFill(bgColor="FFEB9C"),
    'yellow': PatternFill(bgColor="FFFFCC")
}

# Anotación de ACMG para la clasificación de AnnotSV
ACMG_CLASSIFICATION = {1: 'B', 2: 'LB', 3: 'VUS', 4: 'LP', 5: 'P'}


# -------------------------------------------------------------------------
# LOGGING SETUP
# -------------------------------------------------------------------------

logging.basicConfig(
    level=logging.DEBUG,  # Nivel mínimo que se mostrará
    format='%(asctime)s - [ %(levelname)s ] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


# -------------------------------------------------------------------------
# CARGA Y GUARDADO DE ARCHIVOS
# -------------------------------------------------------------------------

def save_pickle(data, file_path):
    """Guarda un diccionario en formato pickle."""
    with open(file_path, 'wb') as f:
        pickle.dump(data, f)


def load_pickle(file_path):
    """Carga un archivo en formato pickle."""
    with open(file_path, 'rb') as f:
        return pickle.load(f)


def set_csv_field_size():
    """Configura el límite de tamaño de campo CSV de forma segura."""
    max_size = sys.maxsize
    while True:
        try:
            csv.field_size_limit(max_size)
            break
        except OverflowError:
            max_size = int(max_size / 10)


def build_sv_id(row):
    """Genera un ID de SV con formato: chrom_start_end_type."""
    return f"{row['SV_chrom']}_{row['SV_start']}_{row['SV_end']}_{row['SV_type']}"


# -------------------------------------------------------------------------
# PROCESAMIENTO DE DATOS DE LAS MUESTRAS Y MAPEO DE IDS
# -------------------------------------------------------------------------

def split_analysis_folder(folder_path):
    """Extrae nodo y ID de la familia a partir de la ruta de analisis."""
    parts = [p for p in os.path.normpath(folder_path).split(os.sep) if p]
    if len(parts) < 3:
        raise ValueError(f'Ruta de analisis invalida: {folder_path}')

    if parts[-1].upper() != 'SV':
        raise ValueError(f"La ruta debe terminar en 'SV': {folder_path}")

    node = parts[-3]
    family_id = parts[-2]
    return node, family_id


def normalize_chromosome(chrom):
    """Normaliza cromosomas para obtener en formato base (sin 'chr')."""
    chrom_str = str(chrom)
    base = chrom_str[3:] if chrom_str.startswith('chr') else chrom_str
    return base


def get_variant_genes(variant):
    """Devuelve los genes de una variante sin duplicados y preservando el orden."""
    raw_genes = variant.get('Gene_name') or variant.get('Genes') or ''
    genes = []
    seen = set()

    for gene in raw_genes.split(';'):
        gene = gene.strip()
        if not gene or gene in seen:
            continue
        seen.add(gene)
        genes.append(gene)

    return genes


def load_phenotype_data(file_path, family_id):
    """Carga fenotipos de la familia en un diccionario por ID IMPaCT."""
    try:
        df = pd.read_csv(file_path, sep='\t', encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, sep='\t', encoding='ISO-8859-1')

    pattern = f'-{family_id}-'
    df = df[df['Muestra ID'].astype(str).str.contains(pattern)]

    if df.empty:
        raise ValueError('No se encontraron muestras para esta familia')
    
    # Convertir a diccionario (clave: ID IMPaCT)
    pheno_dict = df.set_index('Muestra ID').T.to_dict()
    return pheno_dict


def parse_ped_file(ped_file):
    """
    Lee el archivo PED de la familia y devuelve sexo/estatus por muestra.

    Args:
        ped_file (str): Ruta al archivo PED de la familia.

    Returns:
        dict: Diccionario con la información del PED.
    """
    # Leer archivo PED filtrando filas con 'impact' en el ID
    try:
        ped_data = pd.read_csv(ped_file, sep='\t', encoding='utf-8', dtype=str)
        ped_data = ped_data[ped_data['id'].str.contains('impact', na=False)]
        
        # Crear diccionario con información de sexo y estatus
        ped_dict = {}
        for _, row in ped_data.iterrows():
            try:
                sample_id = row['id'].replace('\u200b', '')  # Eliminar caracteres invisibles
                ped_dict[sample_id] = {
                    'sex': int(row['sex']),
                    'status': int(row['aff'])
                }
            except ValueError as ve:
                logging.warning(f'Error al procesar el PED para {sample_id}: {ve}')

        return ped_dict

    except FileNotFoundError:
        logging.error(f'El archivo {ped_file} no se encuentra')
        raise
    except Exception as e:
        logging.error(f'Error al leer el archivo PED: {e}')
        raise


def build_cnag_id_maps(pheno_dict):
    """
    Construye los mapas de conversión entre IDs de muestras de CNAG e IMPaCT.

    Args:
        pheno_dict (dict): Diccionario de fenotipos de las muestras.

    Returns:
        dict (imp_to_cnag): De IMPaCT a CNAG (key: IMPaCT ID, value: CNAG ID)
        dict (cnag_to_imp): De CNAG a IMPaCT (key: CNAG ID, value: IMPaCT ID)
    """
    imp_to_cnag, cnag_to_imp = {}, {}

    for imp_id, data in pheno_dict.items():
        cnag_id = data.get('ID_CNAG')
        # Construir diccionarios con el mapa de IDs
        if cnag_id:
            imp_to_cnag[imp_id] = cnag_id
            cnag_to_imp[cnag_id] = imp_id

    return imp_to_cnag, cnag_to_imp


def get_full_id(short_id, pheno_dict):
    """Obtiene el ID completo de una muestra germinal a partir del ID corto."""
    matches = [
        key for key, val in pheno_dict.items()
        if key.startswith(short_id) and 'ADN germinal' in val.get('Muestra', '')
    ]
    return matches[0] if matches else None


def get_proband_id(family_id, pheno_dict):
    """Devuelve el ID IMPaCT del probando con el diccionario de fenotipos."""
    proband_prefix = f'{family_id}-{family_id}-4impact'
    return get_full_id(proband_prefix, pheno_dict)


# -------------------------------------------------------
# Funciones de solapamiento y comparación entre variantes
# -------------------------------------------------------

def assert_full_split_id_match(full_dict, split_dict, sample_id):
    """Comprueba si las SVs 'full' y 'split' son las mismas. Detiene la
    ejecución si no coinciden, excepto si son SVs sin genes asociados."""
    full_ids = set(full_dict.keys())
    split_ids = set(split_dict.keys())

    only_full = full_ids - split_ids
    only_split = split_ids - full_ids

    # Permitir SVs solo en 'full' cuando no tienen genes asociados.
    allowed_only_full = set()
    for sv_id in only_full:
        gene_count = full_dict.get(sv_id, {}).get('Gene_count', None)
        try:
            if float(gene_count) == 0:
                allowed_only_full.add(sv_id)
        except (TypeError, ValueError):
            pass

    unexpected_only_full = only_full - allowed_only_full

    if unexpected_only_full or only_split:
        only_full_preview = ', '.join(sorted(unexpected_only_full)) if unexpected_only_full else '-'
        only_split_preview = ', '.join(sorted(only_split)) if only_split else '-'
        logging.error(
            f'Muestra {sample_id}: SVs full/split inconsistentes. '
            f'Solo en "full" ({len(unexpected_only_full)}): {only_full_preview} | '
            f'Solo en "split" ({len(only_split)}): {only_split_preview}'
        )
        raise SystemExit(1)

    logging.info(
        f'{sample_id}: {len(full_dict)} SVs "full", {len(split_dict)} SVs "split". '
        'SVs "full" y "split" consistentes.'
    )


### Función para los archivos de SVs de CNAG que tienen discrepancias entre full y split (VUSCAN_03_04 y VUSCAN_06_07)
# def assert_full_split_id_match(full_dict, split_dict, sample_id, allow_cleanup=False):
#     """
#     Comprueba si las SVs 'full' y 'split' son las mismas.

#     Si 'allow_cleanup' es True, elimina SVs inconsistentes en ambos
#     diccionarios y continúa la ejecución. En caso contrario, si no
#     coinciden, detiene la ejecución (excepto SVs 'full' sin genes asociados).
#     """
#     full_ids = set(full_dict.keys())
#     split_ids = set(split_dict.keys())

#     only_full = full_ids - split_ids
#     only_split = split_ids - full_ids

#     # Permitir SVs solo en 'full' cuando no tienen genes asociados.
#     allowed_only_full = set()
#     for sv_id in only_full:
#         gene_count = full_dict.get(sv_id, {}).get('Gene_count', None)
#         try:
#             if float(gene_count) == 0:
#                 allowed_only_full.add(sv_id)
#         except (TypeError, ValueError):
#             pass

#     unexpected_only_full = only_full - allowed_only_full

#     if unexpected_only_full or only_split:
#         only_full_preview = ', '.join(sorted(unexpected_only_full)) if unexpected_only_full else '-'
#         only_split_preview = ', '.join(sorted(only_split)) if only_split else '-'

#         if allow_cleanup:
#             mismatched_ids = unexpected_only_full | only_split
#             for sv_id in mismatched_ids:
#                 full_dict.pop(sv_id, None)
#                 split_dict.pop(sv_id, None)

#             logging.warning(
#                 f'Muestra {sample_id}: SVs full/split inconsistentes. '
#                 f'Solo en "full" ({len(unexpected_only_full)}): {only_full_preview} | '
#                 f'Solo en "split" ({len(only_split)}): {only_split_preview}. '
#                 f'Se eliminaron {len(mismatched_ids)} SVs inconsistentes y se continúa la ejecución.'
#             )
#         else:
#             logging.error(
#                 f'Muestra {sample_id}: SVs full/split inconsistentes. '
#                 f'Solo en "full" ({len(unexpected_only_full)}): {only_full_preview} | '
#                 f'Solo en "split" ({len(only_split)}): {only_split_preview}'
#             )
#             raise SystemExit(1)

#     logging.info(
#         f'{sample_id}: {len(full_dict)} SVs "full", {len(split_dict)} SVs "split". '
#         'SVs "full" y "split" consistentes.'
#     )


def check_reciprocal_overlap(a, b, threshold=OVERLAP_THRESHOLD):  ### EHI: revisar calculo longitud
    """
    Comprueba si hay solapamiento recíproco ≥ umbral entre dos variantes.
    
    Args:
        a, b (dict): Diccionarios con 'SV_start', 'SV_end', 'SV_length'.
        threshold (float): Umbral mínimo de superposición recíproca.
    
    Returns:
        bool: True hay superposición recíproca ≥ threshold, False si no.
    """
    # Obtener coordenadas de las variantes
    start1, end1 = int(a['SV_start']), int(a['SV_end'])
    start2, end2 = int(b['SV_start']), int(b['SV_end'])
    
    # Calcular la longitud de la superposición
    overlap = max(0, min(end1, end2) - max(start1, start2) + 1)

    # Obtener longitud de las variantes
    try:
        len_a = abs(int(float(a.get('SV_length'))))
    except (TypeError, ValueError):
        len_a = end1 - start1 + 1

    try:
        len_b = abs(int(float(b.get('SV_length'))))
    except (TypeError, ValueError):
        len_b = end2 - start2 + 1

    # Ajustar la longitud a 1 si es 0
    len_a = max(len_a, 1)
    len_b = max(len_b, 1)

    # Verificar si la superposición recíproca es al menos el umbral
    return (overlap / len_a >= threshold) and (overlap / len_b >= threshold)


def build_chrom_interval_tree(variants_dict, include_score=False):
    """
    Construye un índice IntervalTree de CNVs o SVs por cromosoma.
    Se añade la puntuación de AnnotSV si se solicita.
    
    Args:
        variants_dict (dict): Diccionario de CNVs o SVs.
        include_score (bool): Si True, incluye puntuación de AnnotSV al índice.

    Returns:
        dict: {cromosoma: IntervalTree con las variantes}
    """
    chrom_index = {}

    for var_id, var_data in variants_dict.items():
        try:
            chrom = normalize_chromosome(var_data['SV_chrom'])
            start = int(var_data['SV_start'])
            end = int(var_data['SV_end'])
            length = var_data['SV_length']
            sv_type = var_data['SV_type']

            # Corregir inserciones
            if start == end:
                end += 1
            # Validar coordenadas
            elif start > end:
                raise ValueError(f'Intervalo inválido en {var_id}: start ({start}) > end ({end})')

            # Crear el índice para el cromosoma si no existe
            if chrom not in chrom_index:
                chrom_index[chrom] = IntervalTree()

            # Añadir variante al índice del cromosoma e incluir puntuación si se indica
            if include_score:
                score = var_data.get('AnnotSV_ranking_score', 0)
                chrom_index[chrom][start:end] = (var_id, start, end, sv_type, length, score)
            else:
                chrom_index[chrom][start:end] = (var_id, start, end, sv_type, length)
        
        except Exception as e:
            logging.error(f'Error procesando la variante {var_id}: {e}')
    
    return chrom_index


def should_replace_best_sv(best_id, best_data, candidate_id, candidate_data):
    """Devuelve True si la variante candidata debe sustituir a la mejor actual."""
    best_score = float(best_data['AnnotSV_ranking_score']) if best_data['AnnotSV_ranking_score'] not in [None, ''] else 0.0
    candidate_score = float(candidate_data['AnnotSV_ranking_score']) if candidate_data['AnnotSV_ranking_score'] not in [None, ''] else 0.0
    
    # Normalización de puntuaciones negativas
    if best_score < 0 and candidate_score < 0:
        best_score = abs(best_score)
        candidate_score = abs(candidate_score)
    # Detectar si tienen signos diferentes
    elif (best_score < 0 and candidate_score > 0) or (best_score > 0 and candidate_score < 0):
        logging.info(f'Diferente signo de AnnotSV entre {best_id} ({best_score}) y {candidate_id} ({candidate_score})')

    best_length = best_data['SV_length']
    candidate_length = candidate_data['SV_length']

    # Comprobar si la candidata tiene mejor puntuación y, si hay empate, es más larga que la mejor actual
    return (
        candidate_score > best_score
        or (candidate_score == best_score and candidate_length > best_length)
    )


def get_best_overlapping(sv_id, sv_data, sv_type, overlaps):
    """
    Entre una SV y un IntervalTree de SVs, devuelve la mejor (mayor
    score AnnotSV / mayor longitud en empate) y las SVs solapantes.

    Returns:
        tuple: (best_sv_id, merged_variants_set)
    """
    best_variant = (sv_id, sv_data)  # Inicializar mejor SV como la dada
    merged_variants = set()          # Conjunto para las SV solapantes

    # Buscar SVs solapantes
    for overlap in overlaps:
        candidate_id, candidate_start, candidate_end, candidate_type, candidate_length, candidate_score = overlap.data

        # No comparar SV consigo misma
        if candidate_id == sv_id:
            continue

        # Comparar solo SVs del mismo tipo
        if sv_type != candidate_type:
            continue

        # Crear diccionario para la SV candidata
        candidate_data = {
            'SV_start': candidate_start,
            'SV_end': candidate_end,
            'SV_type': candidate_type,
            'AnnotSV_ranking_score': candidate_score,
            'SV_length': candidate_length
        }

        # Comprobar si hay solapamiento recíproco
        if check_reciprocal_overlap(sv_data, candidate_data):
            best_id, best_data = best_variant
            # Comprobar si la SV candidata reemplaza a la mejor actual
            # (mayor puntuación AnnotSV y, si hay empate, más larga)
            if should_replace_best_sv(best_id, best_data, candidate_id, candidate_data):
                merged_variants.add(best_id)
                best_variant = (candidate_id, candidate_data)
            else:
                merged_variants.add(candidate_id)

    best_sv_id, _ = best_variant
    return best_sv_id, merged_variants


# -------------------------------------------------
# PASO 1: Parseo y procesamiento de archivos
# -------------------------------------------------

def process_sv_cnag(file_path):
    """
    Procesa archivos TSV de SVs de CNAG, aplica filtros (Qual, 'PASS' y frecuencia
    interna).

    Args:
        file_path (str): Ruta al archivo TSV.

    Returns:
        tuple: full_dict (SVs 'full'), split_dict (SVs 'split')
    """
    # Inicializar diccionarios
    full_dict = {}   # SVs 'full'
    split_dict = {}  # SVs 'split'

    try:
        with gzip.open(file_path, 'rt') as f:
            reader = csv.DictReader(f, delimiter='\t')
            for row in reader:
                # Obtener ID de la SV
                sv_id = build_sv_id(row)
                
                # Filtrar por calidad
                if float(row['QUAL']) < QUAL_MIN:
                    continue

                # Filtrar 'PASS'
                if row['FILTER'] != 'PASS':
                    continue

                # Filtrar por frecuencia interna
                try:
                    row['Illumina.exact.counts'] = 0 if row['Illumina.exact.counts'] == '#' or \
                    row['Illumina.exact.counts'] == '' else float(row['Illumina.exact.counts'])
                    if float(row['Illumina.exact.counts']) >= INTERNAL_FREQ_MAX:
                        continue
                except ValueError:
                    logging.error(f'SV {sv_id}: su frecuencia interna no es un valor numérico ({row["Illumina.exact.counts"]})')
                    continue

                # Convertir '#' o '' a 0 en la frecuencia interna similar
                if row['Illumina.similar.counts'] == '#' or row['Illumina.similar.counts'] == '':
                    row['Illumina.similar.counts'] = 0
                
                # Guardar según tipo de anotación ('full'/'split')
                if row['Annotation_mode'] == 'full':
                    if sv_id not in full_dict:
                        full_dict[sv_id] = {k: v for k, v in row.items() if k != 'sv_id'}
                elif row['Annotation_mode'] == 'split':
                    if sv_id in full_dict:
                        split_dict.setdefault(sv_id, []).append(row)

    except csv.Error as e:
        logging.error(f'Error de CSV en el archivo {file_path}: {e}')
    except Exception as e:
        logging.error(f'Error al leer el archivo {file_path}: {e}')
    
    return full_dict, split_dict


def process_sv_fpgmx(full_file, split_file):
    """
    Procesa archivos .tab.gz con CNVs y SVs de FPGMX. Selecciona SVs, aplica
    filtros (calidad y frecuencia interna).
    
    Args:
        full_file (str): Ruta al archivo de SVs 'full' (*.full.tab.gz).
        split_file (str): Ruta al archivo de SVs 'split' (*.split.tab.gz).

    Returns:
        tuple: full_dict (SVs 'full'), split_dict (SVs 'split')
    """
    # Inicializar diccionarios
    full_dict = {}   # SVs 'full'
    split_dict = {}  # SVs 'split'

    # Procesar SVs 'full' y 'split'
    for sv_file, is_full in [(full_file, True), (split_file, False)]:
        try:
            with gzip.open(sv_file, 'rt') as f:
                reader = csv.DictReader(f, delimiter='\t')
                
                # Para cada SV del archivo
                for row in reader:
                    try:
                        # Seleccionar SVs
                        if row['SV_type_original'] in {'GAIN', 'LOSS'}:
                            continue

                        # Filtrar por calidad
                        if int(row['Qual']) < QUAL_MIN:
                            continue
                        
                        # Filtrar por frecuencia interna
                        if int(row['Illumina_DRAGEN.exact.counts']) >= INTERNAL_FREQ_MAX:
                            continue

                        sv_id = build_sv_id(row)
                        
                        # Guardar según si es 'full' o 'split'
                        if is_full:
                            full_dict.setdefault(sv_id, {k: v for k, v in row.items() if k != 'sv_id'})
                        else:
                            split_dict.setdefault(sv_id, []).append(row)
                        
                    except ValueError as ve:
                        logging.error(f'Error al convertir Qual en {row}: {ve}')

        except csv.Error as e:
            logging.error(f'Error de CSV en el archivo {sv_file}: {e}')
        except Exception as e:
            logging.error(f'Error al leer el archivo {sv_file}: {e}')
        
    return full_dict, split_dict


def process_sv_nasertic(file_path):
    """
    Procesa un archivo TAB o TSV de SVs de NASERTIC, aplica filtros (Qual,
    'PASS' y frecuencia interna).

    Args:
        file_path (str): Ruta al archivo TAB o TSV a procesar.

    Returns:
        tuple: full_dict (SVs 'full'), split_dict (SVs 'split')
    """
    # Inicializar diccionarios
    full_dict = {}   # SVs 'full'
    split_dict = {}  # SVs 'split'
    
    try:
        with open(file_path, 'rt') as f:
            reader = csv.DictReader(f, delimiter='\t')
            for row in reader:
                try:
                    # Filtrar por calidad
                    if float(row['QUAL']) < QUAL_MIN:
                        continue
                    # Filtrar 'PASS'
                    if row['FILTER'] != 'PASS':
                        continue
                    # Filtrar por frecuencia interna
                    if int(row['Illumina_DRAGEN.exact.counts']) >= INTERNAL_FREQ_MAX:
                        continue
                    
                    sv_id = build_sv_id(row)
                    
                    # Guardar según anotación ('full' o 'split')
                    if row['Annotation_mode'] == 'full':
                        if sv_id not in full_dict:
                            full_dict[sv_id] = {k: v for k, v in row.items() if k != 'sv_id'}
                    elif row['Annotation_mode'] == 'split':
                        split_dict.setdefault(sv_id, []).append(row)
                
                except ValueError as ve:
                    logging.error(f'Error al convertir Qual o Illumina_DRAGEN.exact.counts en {row}: {ve}')

    except csv.Error as e:
        logging.error(f'Error de CSV en el archivo {file_path}: {e}')
    except Exception as e:
        logging.error(f'Error al leer el archivo {file_path}: {e}')
    
    return full_dict, split_dict


def run_preprocess_sv(folder_inputs, folder_outputs, node, family_id, pheno_dict):
    """
    Ejecuta paso 1: Preprocesa SVs germinales de la familia. Según el nodo de
    secuenciación, selecciona los archivos de entrada y aplica filtros. Valida
    la consistencia entre anotaciones full/split y guarda la salida en formato
    pickle. Genera un TSV ordenado por cromosoma e inicio para las SVs 'full'.
    
    Args:
        folder_inputs (str): Ruta al directorio con los archivos (.tab.gz para
                             FPGMX, .tab para NASERTIC y .tsv para CNAG).
        folder_outputs (str): Ruta al directorio de salida.
        node (str): Nodo de secuenciación.
        family_id (str): ID de la familia.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    # Aumentar el límite de tamaño de campo CSV
    set_csv_field_size()

    # Seleccionar archivos de SVs
    input_files = glob.glob(os.path.join(folder_inputs, '*'))
    if node == 'CNAG':
        family_files = [file_path for file_path in input_files if file_path.endswith('.tab.gz')]
        imp_to_cnag, cnag_to_imp = build_cnag_id_maps(pheno_dict)
    elif node == 'FPGMX':
        family_files = [file_path for file_path in input_files if file_path.endswith('full.tab.gz')]
    elif node == 'NASERTIC':
        family_files = [file_path for file_path in input_files if file_path.endswith(('.tab', '.tsv'))]

    # Validar que existen archivos para todas las muestras germinales esperadas
    expected_imp_ids = {
        key for key, val in pheno_dict.items()
        if 'ADN germinal' in val.get('Muestra', '')
    }

    if node == 'CNAG':
        expected_file_ids = {imp_to_cnag[imp_id] for imp_id in expected_imp_ids}
    else:
        expected_file_ids = expected_imp_ids

    found_file_ids = {os.path.basename(file_path).split('.')[0] for file_path in family_files}
    missing_file_ids = sorted(expected_file_ids - found_file_ids)
    if missing_file_ids:
        raise SystemExit(
            '[ ERROR ] - Faltan archivos de SVs germinales de las muestras: '
            f'{", ".join(missing_file_ids)}'
        )
    
    # Procesar cada muestra de la familia
    try:
        for file_path in family_files:
            if node == 'CNAG':
                cnag_id = os.path.basename(file_path).split('.')[0]
                sample_id = cnag_to_imp[cnag_id]

                logging.info(f'Procesando archivo de SVs de la muestra: {cnag_id} / {sample_id}')
                full_dict, split_dict = process_sv_cnag(file_path)

            elif node == 'FPGMX':
                sample_id = os.path.basename(file_path).split('.')[0]
                split_file = file_path.replace('full.tab.gz', 'split.tab.gz')

                logging.info(f'Procesando archivos de SVs de la muestra: {sample_id}')
                full_dict, split_dict = process_sv_fpgmx(file_path, split_file)
            
            elif node == 'NASERTIC':
                sample_id = os.path.basename(file_path).split('.')[0]

                logging.info(f'Procesando archivo de SVs de la muestra: {sample_id}')
                full_dict, split_dict = process_sv_nasertic(file_path)

            # Verificar consistencia entre SVs full y split
            assert_full_split_id_match(full_dict, split_dict, sample_id)
            # assert_full_split_id_match(full_dict, split_dict, sample_id, allow_cleanup=(node == 'CNAG'))

            # Guardar SVs 'full' y 'split'
            output_split = os.path.join(folder_outputs, f'{sample_id}.SVs.annotated_parsed_split.pkl')
            save_pickle(split_dict, output_split)

            output_full = os.path.join(folder_outputs, f'{sample_id}.SVs.annotated_parsed.pkl')
            save_pickle(full_dict, output_full)

            # Guardar SVs 'full' en un archivo TSV (orden por cromosoma y posición de inicio)
            sorted_ids = natsorted(full_dict.keys(), key=lambda id: (full_dict[id].get('SV_chrom', ''), int(full_dict[id].get('SV_start', 0))))
            header = ['AnnotSV_ID'] + list(next(iter(full_dict.values())).keys())

            output_tsv = os.path.join(folder_outputs, f'{sample_id}.SVs.annotated_parsed.tsv')
            with open(output_tsv, 'w', newline='') as tsvfile:
                writer = csv.DictWriter(tsvfile, fieldnames=header, delimiter='\t')
                writer.writeheader()
                for id in sorted_ids:
                    row = {'AnnotSV_ID': id, **full_dict[id]}
                    writer.writerow(row)
        
            logging.info(f'Muestra {sample_id} preprocesada: {len(full_dict)} SVs "full"')

    except FileNotFoundError:
        logging.error(f'Directorio no encontrado: {folder_inputs}')
    except Exception as e:
        logging.error(f'Error inesperado al procesar la familia {family_id}: {e}', exc_info=True)


# -------------------------------------------------
# PASO 2: Priorización y anotación de SVs
# -------------------------------------------------

# ---------- Paso 2.1: Filtrado de SVs solapantes ---------- #
def run_filter_overlapping_sv(family_id, proband, folder_path, folder_outputs):
    """
    Ejecuta paso 2.1: Filtra SVs solapantes del probando, seleccionando la de
    mayor puntuación AnnotSV/longitud y guarda el ID del resto. Guarda el 
    resultado en un archivo pickle.
    
    Args:
        family_id (str): ID familia.
        proband (str): ID completo del probando.
        folder_path (str): Carpeta donde están los archivos.
        folder_outputs (str): Carpeta para guardar el resultado.
    """
    logging.info("PASO 2.1: Filtrando SVs solapantes del probando")

    # Cargar SVs 'full' del probando
    prob_pkl = os.path.join(folder_path, f'{proband}.SVs.annotated_parsed.pkl')
    prob_dict = load_pickle(prob_pkl)

    # Crear índice por cromosoma
    chrom_index = build_chrom_interval_tree(prob_dict, include_score=True)

    # Inicializar diccionario de SVs filtradas
    filtered_dict = {}

    # Filtrar variantes con solapamiento
    for sv_id, sv_data in prob_dict.items():
        sv_type = sv_data.get('SV_type')
        chrom = sv_data.get('SV_chrom')
        if chrom not in chrom_index:
            continue

        try:
            sv_start = int(sv_data['SV_start'])
            sv_end = int(sv_data['SV_end'])
        except ValueError as e:
            logging.error(f'Error en la variante {sv_id}: {e}')
            continue  # Saltar variante problemática
        
        # Obtener IntervalTree (prefiltrar por intervalo y cromosoma)
        query_end = sv_end + 1 if sv_end >= sv_start else sv_start + 1
        overlaps = chrom_index[chrom].overlap(sv_start, query_end)

        # Buscar SV solapantes y seleccionar la de mayor puntuación AnnotSV y, si hay empate, más larga
        best_sv_id, merged_variants = get_best_overlapping(sv_id, sv_data, sv_type, overlaps)

        # Guardar la SV seleccionada
        if best_sv_id not in filtered_dict:
            filtered_dict[best_sv_id] = prob_dict[best_sv_id].copy()
            filtered_dict[best_sv_id]['Merged_variants'] = sorted(merged_variants)

    # Guardar diccionario filtrado
    file_path = f'{folder_outputs}{family_id}_proband_svs_filtered.pkl'
    save_pickle(filtered_dict, file_path)

    logging.info(f'Filtradas SVs del probando: de {len(prob_dict)} a {len(filtered_dict)} SVs')


# ---------- PASO 2.2: Anotación de SVs del probando compartidas con familiares ---------- #
def build_chrom_sv_index(dct):
    """
    Construye un índice IntervalTree de SVs por cromosoma.

    Args:
        dct (dict): Diccionario de SVs {sv_id: sv_data_dict}

    Returns:
        dict: {cromosoma: IntervalTree de SVs}
    """
    chrom_index = {}
    
    for sv_id, sv_data in dct.items():
        try:
            # Extraer datos de la variante
            chrom = normalize_chromosome(sv_data['SV_chrom'])
            start = int(sv_data['SV_start'])
            end = int(sv_data['SV_end'])
            
            # Corregir inserciones
            if start == end:
                end += 1
            
            # Validar coordenadas
            elif start > end:
                raise ValueError(f'Intervalo inválido: start ({start}) > end ({end})')
                
            # Crear el índice para este cromosoma si no existe
            if chrom not in chrom_index:
                chrom_index[chrom] = IntervalTree()

            # Añadir la variante al índice del cromosoma
            chrom_index[chrom][start:end] = (sv_id, sv_data)
            
        except Exception as e:
            logging.error(f'Error procesando la variante {sv_id}: {e}')
    
    return chrom_index


def run_annotate_shared_svs(folder_inputs, folder_path, family_id, node, proband, pheno_dict):
    """
    Ejecuta paso 2.2: Anota SVs del probando compartidas con familiares
    (solapamiento recíproco) y añade en un diccionario las SVs de familiares
    no compartidas con el probando. Guarda los resultados en archivos pickle.

    Args:
        folder_inputs (str): Ruta a la carpeta con las SVs de familiares.
        folder_path (str): Ruta a la carpeta de análisis.
        family_id (str): ID familia/probando.
        node (str): Nodo de secuenciación.
        proband (str): ID completo del probando.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    logging.info("PASO 2.2: Anotando SVs del probando compartidas con familiares")
    
    # Cargar SVs del probando
    proband_file = f'{folder_path}{family_id}_proband_svs_filtered.pkl'
    proband_dict = load_pickle(proband_file)

    # Para CNAG, cambiar claves en el diccionario del probando al ID CNAG
    if node == 'CNAG':
        imp_to_cnag, _ = build_cnag_id_maps(pheno_dict)
        prob_cnag = imp_to_cnag[proband]
        for sv_id in list(proband_dict.keys()):
            sv_dict = proband_dict[sv_id]
            if prob_cnag in sv_dict:
                sv_dict[proband] = sv_dict[prob_cnag]
    
    # Archivos con las SVs familiares
    rels_files = [
        os.path.join(folder_inputs, f) for f in os.listdir(folder_inputs)
        if not f.startswith(family_id) and f.endswith('.SVs.annotated_parsed.pkl')
    ]

    if not rels_files:
        logging.info("No hay SVs de familiares, se omite la anotación de SVs compartidas")
        output_file = proband_file.replace('_svs_filtered.pkl', '_sharedsvs.pkl')
        save_pickle(proband_dict, output_file)
        return
    
    # Crear índice de SVs del probando por cromosoma
    proband_index = build_chrom_sv_index(proband_dict)

    # Inicializar diccionario para guardar SVs únicas de cada familiar
    rels_unique_sv = {}
    
    # Procesar cada familiar
    for rel_file in rels_files:
        rel_dict = load_pickle(rel_file)
        rel_id = os.path.basename(rel_file).split('.')[0]

        # Diccionario para guardar SVs únicas del familiar
        rel_unique = {}

        # Para CNAG, mapear el ID CNAG - IMPaCT
        if rel_id and node == 'CNAG':
            cnag_id = imp_to_cnag[rel_id]
            for sv_id in list(rel_dict.keys()):
                sv_dict = rel_dict[sv_id]
                if cnag_id in sv_dict:
                    sv_dict[rel_id] = sv_dict[cnag_id]
        
        # Comparar cada SV del familiar con el probando
        for rel_sv_id, rel_sv_data in rel_dict.items():
            # Obtener tipo y genotipo
            if node == 'FPGMX':
                sv_type = rel_sv_data['SV_type_original']
                rel_gt = rel_sv_data['Zygosity']
            else:
                sv_type = rel_sv_data['SV_type']
                sv_data = rel_sv_data[rel_id]
                rel_gt = sv_data.split(':')[0]

            # Obtener cromosoma
            chrom = rel_sv_data['SV_chrom']

            # Buscar variantes solapantes en el índice
            found_match = False  # Indicador de si hay solapamiento

            if chrom in proband_index:
                for overlap in proband_index[chrom]:
                    proband_sv_id, proband_sv_data = overlap.data
                    # OJO: 'proband_sv_data' es una referencia al diccionario original
                    # 'proband_dict[proband_sv_id]'. Cambios aquí modifican 'proband_dict' in-place.
                    
                    # Verificar que el tipo de variante coincida
                    if (
                        proband_sv_data['SV_type'] == sv_type
                        or ('SV_type_original' in proband_sv_data and proband_sv_data['SV_type_original'] == sv_type)
                    ):
                        # Verificar la superposición recíproca
                        if check_reciprocal_overlap(rel_sv_data, proband_sv_data):
                            # Si hay, agregar información al probando
                            if rel_id in proband_sv_data:
                                proband_sv_data[rel_id] += f'; GT:{rel_gt}; {rel_sv_id}'
                            else:
                                proband_sv_data[rel_id] = f'GT:{rel_gt}; {rel_sv_id}'
                            found_match = True

            # Si no hay superposición, añadir SV a las únicas del familiar
            if not found_match:
                rel_unique[rel_sv_id] = rel_sv_data

        # Guardar SVs únicas del familiar en el diccionario de todos los familiares
        rels_unique_sv[rel_id] = rel_unique

    # Guardar SVs únicas de los familiares
    pickle_rel_files = f'{folder_path}{family_id}_relatives_unique.pkl'
    save_pickle(rels_unique_sv, pickle_rel_files)

    # Guardar SVs del probando con la información de solapamiento familiar
    output_file = proband_file.replace('_svs_filtered.pkl', '_sharedsvs.pkl')
    save_pickle(proband_dict, output_file)

    logging.info("Anotación de SVs compartidas completada")


# ---------- PASO 2.3: Combinación de SVs de los familiares ----------
def run_merge_relatives_svs(folder_path, family_id, node, pheno_dict):
    """
    Ejecuta paso 2.3: Combina SVs compartidas de familiares (solapamiento
    recíproco), añade información fenotípica y guarda en un archivo pickle.

    Args:
        folder_path (str): Ruta a la carpeta con los archivos.
        family_id (str): ID de la familia.
        node (str): Nodo de secuenciación.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    
    Returns:
        dict: IntervalTree por cromosoma con las SVs de todos los familiares.
    """
    logging.info("PASO 2.3: Combinando SVs de familiares")

    # Cargar SVs únicas de los familiares
    rels_file_path = f'{folder_path}{family_id}_relatives_unique.pkl'
    if not os.path.exists(rels_file_path):
        logging.info("No hay SVs únicas de familiares para combinar")
        return {}

    rels_dict = load_pickle(rels_file_path)

    # Inicializar diccionario para las SVs fusionadas
    merged_svs = {}

    # Procesar el primer familiar como base del índice
    first_rel_id, first_rel_svs = next(iter(rels_dict.items()))
    first_rel_full_id = get_full_id(first_rel_id, pheno_dict)
    
    # Crear índice IntervalTree de SVs
    merged_svs_index = build_chrom_interval_tree(first_rel_svs, include_score=True)
    
    # Agregar SVs del primer familiar
    for sv_id, sv_data in first_rel_svs.items():
        # Obtener genotipo
        if node == 'FPGMX':
            rel_gt = sv_data['Zygosity']
        else:
            rel_data = sv_data.get(first_rel_full_id, '')
            rel_gt = rel_data.split(':')[0] if rel_data else '.'
    
        merged_svs[sv_id] = sv_data.copy()
        merged_svs[sv_id][first_rel_id] = f'GT:{rel_gt}; {sv_id}'

    # Procesar el resto de familiares
    for rel_id, rel_svs in rels_dict.items():
        if rel_id == first_rel_id:
            continue  # Saltar el primero (ya indexado)

        rel_full_id = get_full_id(rel_id, pheno_dict)
        
        for sv_id, sv_data in rel_svs.items():
            try:
                chrom = sv_data['SV_chrom']
                start = int(sv_data['SV_start'])
                end = int(sv_data['SV_end'])
                sv_type = sv_data['SV_type']
                length = sv_data['SV_length']
                score = sv_data['AnnotSV_ranking_score']
                
                # Extraer genotipo
                if node == 'FPGMX':
                    rel_gt = sv_data['Zygosity']
                else:
                    rel_gt = sv_data[rel_full_id].split(':')[0]

                # Buscar solapamientos recíprocos con SVs ya indexadas
                found_match = False

                if chrom in merged_svs_index:
                    chrom_overlaps = merged_svs_index[chrom]
                    for overlap in chrom_overlaps:
                        existing_id, existing_start, existing_end, existing_type, existing_length, existing_score = overlap.data
                        existing_coords = {
                            'SV_start': existing_start,
                            'SV_end': existing_end,
                            'SV_length': existing_length,
                            'SV_type': existing_type
                        }
                    
                        # Verificar si hay solapamiento recíproco
                        if existing_coords['SV_type'] == sv_type and check_reciprocal_overlap(sv_data, existing_coords):

                            # Añadir información del familiar a la SV ya existente
                            if rel_id in merged_svs[existing_id]:
                                merged_svs[existing_id][rel_id] += f'; GT:{rel_gt}; {sv_id}'
                            else:
                                merged_svs[existing_id][rel_id] = f'GT:{rel_gt}; {sv_id}'
                                
                            found_match = True
                            
                # Si no hay solapamiento, añadir como nueva SV
                if not found_match:
                    merged_svs[sv_id] = sv_data.copy()
                    merged_svs[sv_id][rel_id] = f'GT:{rel_gt}; {sv_id}'

                    # Crear entrada en el índice si no existe
                    if chrom not in merged_svs_index:
                        merged_svs_index[chrom] = IntervalTree()
                    
                    # Validar y ajustar coordenadas para inserciones
                    if start == end:
                        end += 1
                    elif start > end:
                        raise ValueError(f'Intervalo inválido: start ({start}) > end ({end})')
            
                    # Añadir la variante al índice del cromosoma
                    merged_svs_index[chrom][start:end] = (sv_id, start, end, sv_type, length, score)

            except Exception as e:
                logging.error(f'Error al procesar variante {sv_id} de {rel_id}: {e}')
                continue

    # Guardar el diccionario con las SVs combinadas
    output_file = f'{folder_path}{family_id}_relatives_mergedsvs.pkl'
    save_pickle(merged_svs, output_file)

    logging.info(f'Combinadas {len(merged_svs)} SVs de familiares')
    return merged_svs_index


# ---------- PASO 2.4: Filtrar SVs solapantes en familiares ---------- #
def run_filter_relatives_overlapping_svs(family_id, chrom_index, folder_path, node, pheno_dict):
    """
    Ejecuta paso 2.4: Filtra SVs solapantes de familiares, seleccionando
    la de mayor puntuación de AnnotSV/longitud y guarda el ID del resto.
    Guarda el resultado en un archivo pickle.

    Args:
        family_id (str): ID familia.
        folder_path (str): Ruta a la carpeta donde están los archivos.
        chrom_index (dict): IntervalTree por cromosoma de SVs de familiares.
        node (str): Nodo de secuenciación.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    logging.info("PASO 2.4: Filtrando SVs solapantes de los familiares")

    # Cargar SVs de los familiares
    rels_file = f'{folder_path}{family_id}_relatives_mergedsvs.pkl'
    if not os.path.exists(rels_file):
        logging.info("No hay SVs de familiares, se omite el filtrado")
        return
    
    rels_dict = load_pickle(rels_file)

    # Inicializar diccionario de SVs filtradas
    filtered_dict = {}

    # Para CNAG, mapeo de IDs (CNAG -> IMPaCT)
    if node == 'CNAG':
        _, cnag_to_imp = build_cnag_id_maps(pheno_dict)

    # Filtrar variantes con solapamiento
    for sv_id, sv_data in rels_dict.items():
        sv_type = sv_data['SV_type']
        chrom = sv_data['SV_chrom']
        if chrom not in chrom_index:
            continue

        try:
            sv_start = int(sv_data['SV_start'])
            sv_end = int(sv_data['SV_end'])
        except ValueError as e:
            logging.error(f'Error en la variante {sv_id}: {e}')
            continue  # Saltar variante problemática
        
        # Obtener IntervalTree (prefiltrar por intervalo y cromosoma)
        query_end = sv_end + 1 if sv_end >= sv_start else sv_start + 1
        overlaps = chrom_index[chrom].overlap(sv_start, query_end)

        # Obtener solapamientos y seleccionar la mejor SV
        best_sv_id, merged_variants = get_best_overlapping(sv_id, sv_data, sv_type, overlaps)

        # Guardar SV seleccionada
        if best_sv_id not in filtered_dict:
            filtered_dict[best_sv_id] = rels_dict[best_sv_id].copy()
            filtered_dict[best_sv_id]['Merged_variants'] = sorted(merged_variants)

            for variant in merged_variants:
                if node == 'CNAG':
                    smpl_cnag = rels_dict[variant].get('Samples_ID')
                    smpl = cnag_to_imp.get(smpl_cnag)
                    gt = rels_dict[variant].get(smpl)
                    filtered_dict[best_sv_id][smpl] = filtered_dict[best_sv_id].get(smpl, '') + f'. {gt}'

                elif node == 'FPGMX':
                    # Buscar muestras que tengan claves con valor en formato 'GT:...'
                    found_any = False
                    for key, val in rels_dict[variant].items():
                        if isinstance(val, str) and val.startswith('GT:'):
                            found_any = True
                            filtered_dict[best_sv_id][key] = (
                                filtered_dict[best_sv_id].get(key, '') + f'. {val}'
                            )
                    if not found_any:
                        logging.warning(f'No se encontró ninguna muestra para la variante {variant}')
                
                elif node == 'NASERTIC':
                    smpl = rels_dict[variant].get('Patient')
                    gt = rels_dict[variant].get(smpl)
                    filtered_dict[best_sv_id][smpl] = filtered_dict[best_sv_id].get(smpl, '') + f'. {gt}'

    # Guardar diccionario filtrado
    file_path = f'{folder_path}{family_id}_relatives_svs_filtered.pkl'
    save_pickle(filtered_dict, file_path)

    logging.info("Filtrado de SVs solapantes en familiares completado")


# ---------- PASO 2.5: Anotación de genes relacionados con cáncer ---------- #
def load_cosmic(cosmic_file=COSMIC_FILE):
    """
    Carga las anotaciones de genes de COSMIC y los clasifica en un diccionario.

    Args:
        cosmic_file (str): Ruta al archivo COSMIC.

    Returns:
        dict: Genes como claves y tipo (PO, TSG, NA) como valores.
    """
    cosmic_dict = {}

    try:
        with open(cosmic_file, mode='r') as file:
            reader = csv.DictReader(file, delimiter='\t')
            for row in reader:
                gene = row.get('GENE', '').strip()
                if not gene:
                    continue

                if row.get('OG', '').strip() == 'Yes':
                    cosmic_dict[gene] = 'PO'
                elif row.get('TSG', '').strip() == 'Yes':
                    cosmic_dict[gene] = 'TSG'
                else:
                    cosmic_dict[gene] = 'NA'

    except FileNotFoundError:
        logging.error(f'Archivo COSMIC no encontrado {cosmic_file}')
    except Exception as e:
        logging.error(f'Error procesando el archivo COSMIC: {e}')

    return cosmic_dict


def load_hereditary_cancer(cosmic_dict, hereditary_file=HEREDITARY_CANCER_FILE):
    """
    Carga una lista de genes relacionados con cáncer hereditario y los asocia
    con el diccionario COSMIC.

    Args:
        cosmic_dict (dict): Genes de COSMIC.
        hereditary_file (str): Ruta al archivo con los genes de cáncer hereditarios.

    Returns:
        dict: {key: gene, value: PO/TSG/'NA'}
    """
    cancer_dict = {}

    try:
        with open(hereditary_file, mode='r', encoding='iso-8859-1') as file:
            reader = csv.reader(file)
            for row in reader:
                gene = row[0].strip()
                cancer_dict[gene] = cosmic_dict.get(gene, 'NA')
    
    except FileNotFoundError:
        logging.error(f'Archivo de cáncer hereditario no encontrado {hereditary_file}')
    except Exception as e:
        logging.error(f'Error procesando el archivo de cáncer hereditario: {e}')

    return cancer_dict


def annotate_cancer_genes(sv_file, cancer_dict, cosmic_dict):
    """Anota genes de cáncer hereditario y genes COSMIC para cada SV."""
    sv_dict = load_pickle(sv_file)
    
    for sv_id, sv_data in sv_dict.items():
        # Inicializar las listas si no existen
        sv_data.setdefault('hereditary cancer gene list', [])
        sv_data.setdefault('cancer cosmic', [])
        
        for gene in get_variant_genes(sv_data):

            # Añadir gen si está en la lista de cáncer hereditario
            if gene in cancer_dict:
                sv_data['hereditary cancer gene list'].append({
                    'cancer gene': gene,
                    'gene type': cancer_dict[gene]
                })

            # Añadir gen si está en COSMIC
            if gene in cosmic_dict:
                sv_data['cancer cosmic'].append({
                    'cancer gene': gene,
                    'gene type': cosmic_dict[gene]
                })
    
    return sv_dict


def run_annotate_cancer_genes(folder_path, family_id):
    """
    Ejecuta paso 2.5: Anota SVs de probando y familiares con genes relacionados
    con cáncer (COSMIC y lista de cáncer hereditario).
    
    Args:
        folder_path (str): Ruta a la carpeta de análisis.
        family_id (str): ID familia.

    Return:
        tuple: (proband_dict, relatives_dict)
    """
    logging.info("PASO 2.5: Anotando SVs de la familia con genes de cáncer")

    # Cargar datos de COSMIC y cáncer hereditario
    cosmic_dict = load_cosmic()
    cancer_dict = load_hereditary_cancer(cosmic_dict)

    # Procesar SVs del probando
    proband_file = os.path.join(folder_path, f'{family_id}_proband_sharedsvs.pkl')
    proband_dict = annotate_cancer_genes(proband_file, cancer_dict, cosmic_dict)
    
    # Procesar SVs de familiares
    rels_file = os.path.join(folder_path, f'{family_id}_relatives_svs_filtered.pkl')
    if os.path.exists(rels_file):
        rels_dict = annotate_cancer_genes(rels_file, cancer_dict, cosmic_dict)
    else:
        rels_dict = {}

    logging.info("Anotación de genes de cáncer completada")
    return proband_dict, rels_dict


# ---------- PASO 2.6: Anotación de pathways ---------- #
def compute_gene_count(dct):
    """Normaliza el número de genes asociados a cada SV."""
    for sv_id, sv_data in dct.items():
        # Si no hay valor en el número de genes
        if sv_data.get('Gene_count') in [None, '']:
            sv_data['Gene_count'] = len(get_variant_genes(sv_data))
            continue
        # Si hay valor, se convierte a número entero
        else:
            try:
                sv_data['Gene_count'] = int(float(sv_data['Gene_count']))
            except ValueError:
                logging.warning(f'No se pudo convertir Gene_count de "{sv_id}" con valor {sv_data["Gene_count"]} a entero.')
                sv_data['Gene_count'] = len(get_variant_genes(sv_data))

    return dct


def load_uniprot_db(uniprot_file=UNIPROT_FILE):
    """Carga anotaciones funcionales de UniProt en un diccionario."""
    uniprot_dict = {}

    with open(uniprot_file, 'r', encoding='utf-8') as f:
        next(f)  # Saltar la cabecera
        for line in f:
            columns = line.strip().split('\t')
            if len(columns) == 3:
                gene_names = columns[0].split()  # Algunos genes tienen nombres alternativos
                uniprot_function = columns[2]

            for gene in gene_names:
                if gene in uniprot_dict:
                    uniprot_dict[gene] += ";" + uniprot_function  # Si ya existe, añadir la nueva ruta
                else:
                    uniprot_dict[gene] = uniprot_function

    return uniprot_dict


def load_kegg_db(kegg_file=KEGG_FILE):
    """Carga un archivo con datos de KEGG en un diccionario."""
    kegg_dict = {}

    with open(kegg_file, "r") as f:
        next(f)
        for line in f:
            columns = line.strip().split("\t")
            if len(columns) == 2:  # Asegurar que la línea tiene dos columnas
                gene_name, kegg_pathway = columns

            if gene_name in kegg_dict:
                kegg_dict[gene_name] += ";" + kegg_pathway  # Si ya existe, añadir la nueva ruta
            else:
                kegg_dict[gene_name] = kegg_pathway

    return kegg_dict


def load_reactome_db(reactome_file=REACTOME_FILE):
    """Carga pathways de Reactome en un diccionario."""
    reactome_dict = {}
    with open(reactome_file, 'r', encoding='utf-8') as f:
        next(f)
        for line in f:
            columns = line.strip().split('\t')
            if len(columns) == 2:
                gene_name, reactome_pathway = columns

                # Agregar el pathway al diccionario
                if gene_name in reactome_dict:
                    reactome_dict[gene_name] += ";" + reactome_pathway
                else:
                    reactome_dict[gene_name] = reactome_pathway

    return reactome_dict


def add_pathway_annotations(sv_dict, uniprot_dict, kegg_dict, reactome_dict):
    """
    Añade descripcion funcional de UniProt y pathways KEGG/Reactome de los
    genes asociados a cada SV.
    
    Args:
        sv_dict (dict): Diccionario con los genes asociados a cada SV.
        uniprot_dict (dict): Anotaciones funcionales de UniProt.
        kegg_dict (dict): Pathways de KEGG.
        reactome_dict (dict): Pathways de Reactome.
    
    Returns:
        dict: Diccionario actualizado con los campos 'Uniprot_Functions',
              'KEGG_pathways' y 'Reactome_pathways'.
    """
    for sv_data in sv_dict.values():
        if sv_data.get('Gene_count') > 10:
            continue

        genes = get_variant_genes(sv_data)
        
        uniprot_function = []
        kegg_pathways = []
        reactome_pathways = []
        
        for gene in genes:
            # Obtener la descripción de Uniprot, KEGG y Reactome
            function_desc = uniprot_dict.get(gene, 'None')
            function_desc = function_desc.replace('FUNCTION: ', '')
            uniprot_function.append(f'{gene}: {function_desc}')

            kegg = kegg_dict.get(gene, 'NA')
            kegg_pathways.append(f'{gene}: {kegg}')

            reactome = reactome_dict.get(gene, 'NA')
            reactome_pathways.append(f'{gene}: {reactome}')

        # Añadir la información al diccionario
        sv_data['UniProt_Functions'] = '; '.join(uniprot_function)
        sv_data['KEGG_pathways'] = '; '.join(kegg_pathways)
        sv_data['Reactome_pathways'] = '; '.join(reactome_pathways)

    return sv_dict


def run_annotate_pathways(proband_dict, relatives_dict):
    """
    Ejecuta paso 2.6: Anota genes asociados a las SVs de probando y familiares
    con información funcional de UniProt y rutas de KEGG/Reactome (< 10 genes).
    
    Args:
        proband_dict (dict): SVs del probando.
        relatives_dict (dict): SVs de los familiares.
    
    Returns:
        tuple: (proband_dict, relatives_dict) con la anotación añadida.
    """
    logging.info("PASO 2.6: Anotando genes con Uniprot, KEGG y Reactome")

    # Cargar las bases de datos
    uniprot_dict = load_uniprot_db()
    kegg_dict = load_kegg_db()
    reactome_dict = load_reactome_db()

    # Calcular el número de genes asociados a cada SV
    proband_dict = compute_gene_count(proband_dict)
    relatives_dict = compute_gene_count(relatives_dict)
    
    # Añadir anotación a probando y familiares
    proband_dict = add_pathway_annotations(proband_dict, uniprot_dict, kegg_dict, reactome_dict)
    relatives_dict = add_pathway_annotations(relatives_dict, uniprot_dict, kegg_dict, reactome_dict)

    logging.info("Anotación de genes completada")
    return proband_dict, relatives_dict


# ---------- Paso 2.7: Anotación con gnomAD ---------- #
def load_gnomad_db(gnomad_file=GNOMAD_FILE):
    """
    Carga los datos de gnomAD en un diccionario indexado por IDs únicos.
    
    Args:
        gnomad_file (str): Ruta al archivo TSV de gnomAD.
    
    Returns:
        dict: Diccionario donde cada clave es un ID único de SV en gnomAD y
              el valor es un diccionario con sus datos.
    """
    try:
        gnomad_dict = {}

        with open(gnomad_file, 'r') as f:
            reader = csv.DictReader(f, delimiter='\t')
            for row in reader:
                sv_id = f"{row['Chrom']}_{row['Pos']}_{row['END']}_{row['Alt']}"
                gnomad_dict[sv_id] = {k: v for k, v in row.items() if k != 'sv_id'}

        return gnomad_dict
    
    # Devuelve un diccionario vacío si no encuentra el archivo o no se puede cargar
    except FileNotFoundError:
        logging.error(f'Archivo gnomAD no encontrado: {gnomad_file}')
        return {}  
    except Exception as e:
        logging.error(f'Error al cargar gnomAD data: {e}')
        return {}


def build_gnomad_index(gnomad_dict, gnomad_index_path):
    """
    Crea un índice de las SVs de gnomAD por cromosoma usando Numpy arrays.

    Args:
        gnomad_dict (dict): SVs de gnomAD indexadas por ID.
        gnomad_index_path (str): Ruta para guardar el índice como pickle.

    Returns:
        dict: Clave un cromosoma y valor un diccionario con arrays:
              'start', 'end', 'id' (IDs de gnomAD), y 'data' (lista con los
              diccionarios de datos).
    """
    chrom_index = {}

    for gnomad_id, gnomad_data in gnomad_dict.items():
        try:
            chrom = normalize_chromosome(gnomad_data['Chrom'])
            start = int(gnomad_data['Pos'])
            end = int(gnomad_data['END'])
            
            # Corregir inserciones
            if start == end:
                end += 1
                
            # Validar coordenadas
            elif start > end:
                raise ValueError(f'Intervalo inválido: start ({start}) > end ({end})')
                
            # Crear el índice para este cromosoma si no existe
            if chrom not in chrom_index:
                chrom_index[chrom] = {
                    'start': [],
                    'end': [],
                    'id': [],
                    'data': []
                }

            # Añadir SV al índice del cromosoma
            chrom_index[chrom]['start'].append(start)
            chrom_index[chrom]['end'].append(end)
            chrom_index[chrom]['id'].append(gnomad_id)
            chrom_index[chrom]['data'].append(gnomad_data)

        except Exception as e:
            logging.error(f'Error procesando la variante de gnomAD {gnomad_id}: {e}')

    # Convertir listas a arrays NumPy
    for chrom, values in chrom_index.items():
        starts = np.array(values['start'], dtype=np.int32)
        ends = np.array(values['end'], dtype=np.int32)
        ids = np.array(values['id'])
        data = np.array(values['data'], dtype=object)

        # Ordenar por posición de inicio
        order = np.argsort(starts)
        chrom_index[chrom] = {
            'start': starts[order],
            'end': ends[order],
            'id': ids[order],
            'data': data[order]
        }
            
    # Guardar el archivo indexado
    save_pickle(chrom_index, gnomad_index_path)
    logging.info(f'Índice Numpy de gnomAD creado para {len(chrom_index)} cromosomas')

    return chrom_index


def check_reciprocal_overlap_gnomad(sv_data, gnomad_data, threshold=OVERLAP_THRESHOLD):
    """
    Comrprueba si hay olapamiento recíproco entre una SV dada y otra de gnomAD.
    
    Args:
        sv_data (dict): Datos de la SV a analizar.
        gnomad_data (dict): Datos de la SV de gnomAD.
        threshold (float): Umbral de solapamiento recíproco.
        
    Returns:
        bool: True si hay solapamiento recíproco ≥ umbral, False si no.
    """
    # Posiciones de las SVs
    start1, end1 = int(sv_data['SV_start']), int(sv_data['SV_end'])
    start2, end2 = int(gnomad_data['Pos']), int(gnomad_data['END'])
    
    # Calcular la longitud de la superposición
    overlap = max(0, min(end1, end2) - max(start1, start2) + 1)

    # Obtener la longitud de las variantes
    len_sv = int(sv_data['SV_length']) if sv_data.get('SV_length') and sv_data['SV_length'].isdigit() else (end1 - start1 + 1)
    len_gnomad = int(gnomad_data['SVLEN']) if gnomad_data.get('SVLEN') and gnomad_data['SVLEN'].isdigit() else (end2 - start2 + 1)

    # Ajustar la longitud si es 0
    len_sv = max(len_sv, 1)
    len_gnomad = max(len_gnomad, 1)

    # Comprobar si el solapamiento es al menos el umbral
    return (overlap / len_sv >= threshold and overlap / len_gnomad >= threshold)


def map_gnomad_type(sv_type):
    """Mapea los tipos de variantes a sus equivalentes en gnomAD y devuelve
    una lista con los tipos."""
    type_mapping = {
        'DEL': ['DEL', 'CNV', 'DEL:ME:LINE1', 'DEL:ME:HERVK'],
        'DUP': ['DUP', 'CNV'],
        'INS': ['INS', 'INS:ME:ALU', 'INS:ME:LINE1'],
        'TRA': ['BND', 'CPX'],
        'INV': ['INV']
    }
    return type_mapping.get(sv_type, [])


def find_overlaps_numpy(sv_start, sv_end, chrom_data):
    """
    Devuelve los índices de SVs que solapan con el intervalo [sv_start, sv_end].
    """
    starts = chrom_data['start']
    ends = chrom_data['end']
    mask = (starts <= sv_end) & (ends >= sv_start)
    return np.where(mask)[0]


def annotate_gnomad(sv_dict, gnomad_index):
    """
    Anota SVs con información de gnomAD (solapamiento recíproco y eligiendo
    la de mayor AF y FREQ_HOMALT).
    
    Args:
        sv_dict (str): Diccionario con las SVs.
        gnomad_index (dict): Índice de gnomAD por cromosoma (Numpy arrays).
    """
    for sv_id, sv_data in sv_dict.items():
        chrom = sv_data['SV_chrom']
        start = int(sv_data['SV_start'])
        end = int(sv_data['SV_end'])
        sv_type = sv_data['SV_type']
        
        matches = []  # Lista para las coincidencias con gnomAD
        
        # Buscar SVs solapantes
        if chrom in gnomad_index:
            chrom_data = gnomad_index[chrom]
            idx_overlaps = find_overlaps_numpy(start, end, chrom_data)

            # Obtener los tipos equivalentes en gnomAD para el tipo de SV actual
            mapped_types = map_gnomad_type(sv_type)

            # Buscar la mejor coincidencia entre las SVs solapantes 
            for idx in idx_overlaps:
                gnomad_data = chrom_data['data'][idx]
                
                # Verificar que el tipo de variante coincide
                if gnomad_data['Alt'] in mapped_types:
                    # Verificar solapamiento recíproco
                    if check_reciprocal_overlap_gnomad(sv_data, gnomad_data):
                        gnomad_id = chrom_data['id'][idx]
                        matches.append((gnomad_id, gnomad_data))

        # Guardar número de coincidencias encontradas
        sv_data['gnomAD'] = len(matches)

        # Si hay coincidencia, agregar los datos de gnomAD
        if matches:
            # Seleccionar coincidencia con mayor AF y, si hay empate, mayor FREQ_HOMALT
            best_match = max(matches, key=lambda x: (
                float(x[1].get('AF', 0.0) or 0.0),
                float(x[1].get('FREQ_HOMALT', 0.0) or 0.0)
            ))

            gnomad_id, gnomad_data = best_match

            sv_data['gnomad_id'] = gnomad_id
            sv_data['gnomAD_SVLEN'] = gnomad_data['SVLEN']
            sv_data['gnomAD_AC'] = gnomad_data.get('AC', None)
            sv_data['gnomAD_AF'] = gnomad_data.get('AF', None)
            sv_data['gnomAD_AN'] = gnomad_data.get('AN', None)
            sv_data['gnomAD_FREQ_HOMALT'] = gnomad_data.get('FREQ_HOMALT', None)
        
        # Si no hay coincidencia, asignar None
        else:
            sv_data['gnomad_id'] = None
            sv_data['gnomAD_SVLEN'] = None
            sv_data['gnomAD_AC'] = None
            sv_data['gnomAD_AF'] = None
            sv_data['gnomAD_AN'] = None
            sv_data['gnomAD_FREQ_HOMALT'] = None
    
    return sv_dict


def run_annotate_gnomad(output_folder, family_id, proband_dict, relatives_dict):
    """
    Ejecuta paso 2.7: Procesa en paralelo los archivos de SVs de una familia,
    añadiendo información de gnomAD.

    Args:
        output_folder (str): Ruta a la carpeta de salida.
        family_id (str): ID de la familia.
        proband_dict (dict): SVs del probando.
        relatives_dict (dict): SVs de los familiares.

    Returns:
        dict: Diccionario actualizado de SVs del probando.
    """
    logging.info("PASO 2.7: Anotando SVs con gnomAD")

    # Cargar datos de gnomAD
    gnomad_folder = GNOMAD_FILE.split('gnomad.v4')[0]
    gnomad_index_path = f'{gnomad_folder}gnomad_index_sv.pkl'
    
    # Cargar el archivo existe, crear y guardar si no
    if os.path.exists(gnomad_index_path):
        logging.info("Cargando índice de gnomAD")
        gnomad_index = load_pickle(gnomad_index_path)
    else:
        logging.info("Creando índice de gnomAD")
        gnomad_dict = load_gnomad_db()
        gnomad_index = build_gnomad_index(gnomad_dict, gnomad_index_path)

    # Ejecutar probando y familiares en paralelo
    with ThreadPoolExecutor(max_workers=2) as executor:
        fut_prob = executor.submit(annotate_gnomad, proband_dict, gnomad_index)
        fut_rels = executor.submit(annotate_gnomad, relatives_dict, gnomad_index)

        proband_annot = fut_prob.result()
        relatives_annot = fut_rels.result()
    
    # Guardar SVs de familiares
    if relatives_annot:
        relatives_output = os.path.join(output_folder, f'{family_id}_relatives_svs_annotated.pkl')
        save_pickle(relatives_annot, relatives_output)

    logging.info("Anotación con gnomAD completada")
    return proband_annot


# ---------- PASO 2.8: Procesamiento de SVs somáticas ---------- #
def add_germline_column(somatic_dict, germinal_dict):
    """
    Anota SVs somáticas indicando coincidencias con alguna SV germinal
    (superposición recíproca ≥ umbral y mismo tipo).
    
    Args:
        somatic_dict (dict): SVs somáticas.
        germinal_dict (dict): SVs germinales.
    
    Return:
        dict: SVs somáticas con campo 'germline' ('yes'/'no'/'error')
    """
    # Construir IntervalTree para las SV germinales
    germinal_index = build_chrom_interval_tree(germinal_dict)

    # Guardar cromosomas somáticos no encontrados en germinal
    missing_chr = set()
    
    # Recorrer cada SV somática
    for somatic_id, somatic_data in somatic_dict.items():
        try:
            found_match = False # Para indicar si la SV está en germinal
            tumor_data = {
                'SV_start': int(somatic_data['SV_start']),
                'SV_end': int(somatic_data['SV_end']),
                'SV_length': somatic_data['SV_length'],
                'SV_type': somatic_data['SV_type']
            }

            # Comprobar si el cromosoma está en germinal
            somatic_chrom = normalize_chromosome(somatic_data['SV_chrom'])
            if somatic_chrom in germinal_index:
                overlaps = germinal_index[somatic_chrom]
                
                # Revisar cada SV germinal solapante
                for overlap in overlaps:
                    germ_id, germ_start, germ_end, germ_type, germ_length = overlap.data
                    germline_data = {
                        'SV_start': int(germ_start),
                        'SV_end': int(germ_end),
                        'SV_length': germ_length,
                        'SV_type': germ_type
                    } 

                    # Comprobar SVs del mismo tipo
                    if somatic_data['SV_type'] == germline_data['SV_type']:
                        if check_reciprocal_overlap(tumor_data, germline_data):
                            found_match = True
                            break  # Detener búsqueda: coincidencia encontrada
            else:
                missing_chr.add(somatic_chrom)

            # Asignar resultado a la SV
            somatic_dict[somatic_id]['germline'] = 'yes' if found_match else 'no'

        except Exception as e:
            logging.error(f'Error procesando SV somática {somatic_id}: {e}')
            somatic_dict[somatic_id]['germline'] = 'error'
    
    # Mostrar cromosomas somáticos no encontrados en germinal
    if missing_chr:
        missing_str = ', '.join(sorted(missing_chr))
        logging.warning(f'Cromosomas de somático no presentes en germinal: {missing_str}')
    
    return somatic_dict


def process_somatic_sv_file(file_path, sample_id, proband_dict,  exclude_chr_y=False):
    """
    Procesa un archivo de SVs somáticas (TSV/TAB comprimido o no), selecciona las
    'full', filtra por calidad (no hay frecuencia interna) y añade anotación germinal.

    Args:
        file_path (str): Ruta al archivo SV somático.
        sample_id (str): ID muestra para los logs.
        proband_dict (dict): SVs germinales del probando.
        exclude_chr_y (bool): Excluir cromosoma Y en muestras de mujeres.

    Returns:
        dict | None: SVs procesadas, o None si no hay archivo de somático.
    """
    if not os.path.exists(file_path):
        logging.warning(f'No se encontró el archivo: {file_path}.')
        return None
    
    # Cargar archivo somático (TAB, TSV o TSV.GZ)
    opener = gzip.open if file_path.endswith('.gz') else open
    with opener(file_path, 'rt') as f:
        df = pd.read_csv(f, sep='\t', low_memory=False)

    # Validar columnas necesarias
    required_cols = {'AnnotSV_ID', 'Annotation_mode', 'FILTER'}
    missing = required_cols - set(df.columns)
    if missing:
        raise KeyError(f'El archivo {file_path} no contiene las columnas necesarias: {missing}')

    somatic_dict = {}

    for _, row in df.iterrows():
        # Eliminar variantes 'split'
        if row['Annotation_mode'] == 'split':
            continue
        
        # Filtrar por calidad mínima
        try:
            if (float(row['QUAL']) if 'QUAL' in row else float(row['Qual'])) < QUAL_MIN:
                continue
        except (TypeError, ValueError):
            pass # conservar valores no numéricos ('.')
        
        # Filtrar 'PASS'
        if row['FILTER'] != 'PASS':
            continue
        
        # Quitar sufijo numérico de 'AnnotSV_ID' y guardar
        base_annot_id = '_'.join(row['AnnotSV_ID'].split('_')[:-1])
        somatic_dict[base_annot_id] = row.to_dict()

    # Excluir cromosoma mitocondrial
    somatic_dict = {
        k: v for k, v in somatic_dict.items()
        if normalize_chromosome(v.get('SV_chrom', '')) != 'M'
    }

    # Excluir cromosoma Y en muestras de mujeres
    if exclude_chr_y:
        somatic_dict = {
            k: v for k, v in somatic_dict.items()
            if normalize_chromosome(v.get('SV_chrom', '')) != 'Y'
        }

    somatic_dict = add_germline_column(somatic_dict, proband_dict)
    logging.info(f"Procesadas {len(somatic_dict)} SVs somáticas para {sample_id}")
    return somatic_dict


def check_sv_in_somatic(germ_sv_data, somatic_index):
    """
    Comprueba si una SV germinal solapa reciprocamente con SVs somáticas. 
    
    Args:
        germ_sv_data (dict): Datos de la SV germinal.
        somatic_index (dict): IntervalTree por cromosoma de SVs somáticas.
    
    Returns:
        bool: True si hay superposición recíproca, False si no.
    """
    germ_chrom = normalize_chromosome(germ_sv_data['SV_chrom'])

    if germ_chrom not in somatic_index:
        return False  # No hay SVs somáticas en este cromosoma
    
    germ_sv_type = germ_sv_data['SV_type']
    germ_data = {
        'SV_start': germ_sv_data['SV_start'],
        'SV_end': germ_sv_data['SV_end'],
        'SV_length': germ_sv_data['SV_length'],
        'SV_type': germ_sv_data['SV_type']
    }

    overlaps = somatic_index[germ_chrom]       
    for overlap in overlaps:
        som_id, som_start, som_end, som_type, som_length = overlap.data
        som_data = {
            'SV_start': som_start,
            'SV_end': som_end,
            'SV_length': som_length,
            'SV_type': som_type
        }

        if germ_sv_type != som_type:
            continue

        if check_reciprocal_overlap(germ_data, som_data):
            return True
    
    return False


def collect_somatic_files(node, somatic_path, pheno_dict):
    """
    Construye la lista de archivos somáticos por muestra y tipo de comparación.

    Args:
        node (str): Nodo de secuenciación.
        somatic_path (str): Ruta a la carpeta de SVs somáticas.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    
    Returns:
        list: Lista de tuplas (tumor_id, {kind: path, ...})
    """
    somatic_files = os.listdir(somatic_path)
    samples = []

    if node == 'CNAG':
        # Obtener correspondencias de IDs de las muestras
        _, cnag_to_imp = build_cnag_id_maps(pheno_dict)

        to_files = sorted([f for f in somatic_files if f.endswith('NA.SVCalls.ann.tag.tsv') or f.endswith('AB.SVCalls.ann.tag.tsv')])
        for to_file in to_files:
            cnag_id = to_file.split('_vs_')[0]  # CNAG ID: 'ADN tumor' (ej. ABA02219)
            impact_id = cnag_to_imp[cnag_id]    # ID completo muestra (ej. 1943-1943-4impact-02)
            tumor_id = '-'.join([impact_id.split('-')[0], impact_id.split('-')[-1]])  # ID muestra y tumor (ej. 1943-02)
            
            # Ruta a los archivos somáticos disponibles (TN y TO)
            to_path = os.path.join(somatic_path, to_file)
            tn_candidates = [
                fname for fname in somatic_files
                if fname.startswith(f'{cnag_id}_vs_ABA') and fname.endswith('.SVCalls.ann.tag.tsv')
            ]
            tn_path = os.path.join(somatic_path, sorted(tn_candidates)[0]) if tn_candidates else None
            samples.append((tumor_id, {'TN': tn_path, 'TO': to_path}))

    elif node == 'NASERTIC':
        to_files = sorted([f for f in os.listdir(somatic_path) if f.endswith('TO.annotated.PASSfiltered.tsv.gz')])
        for to_file in to_files:
            impct_id = to_file.split('.')[0]  # ID completo muestra (ej. 2270-2270-4impact-02)
            tumor_id = "-".join([impct_id.split("-")[0], impct_id.split("-")[-1]])  # ID muestra y tumor (ej. 2270-02)

            # Ruta a archivos de somático disponibles (TN y TO)
            to_path = os.path.join(somatic_path, to_file) 
            tn_path = to_path.replace('TO.annotated.PASSfiltered.tsv.gz', 'TN.annotated.tsv.gz')
            samples.append((tumor_id, {'TN': tn_path, 'TO': to_path}))
    
    elif node == 'FPGMX':
        to_files = sorted([
            f for f in os.listdir(somatic_path)
            if f.endswith('DTO_.sv.filtered_annotated.tsv.gz')
        ])
        for to_file in to_files:
            impact_id = to_file.split('.')[0]  # ID completo muestra (ej. 2270-2270-4impact-02)
            tumor_id = "-".join([impact_id.split("-")[0], impact_id.split("-")[-1]]) # ID muestra y tumor (ej. 2270-02)

            # Ruta a los 3 archivos de somático
            to_path = os.path.join(somatic_path, to_file)
            tn_path = to_path.replace('DTO_.sv.filtered_annotated.tsv.gz', 'DTN_.sv.annotated.tsv.gz')
            pon_path = to_path.replace('_*DTO_.sv.filtered_annotated.tsv.gz', '_D_.sv.annotated.tsv.gz')
            samples.append((tumor_id, {'TN': tn_path, 'PON': pon_path, 'TO': to_path}))
    
    return samples


def run_process_somatic(folder_path, folder_output, family_id, node, proband_dict, pheno_dict):
    """
    Ejecuta paso 2.8: Procesa archivos de SVs somáticas que guarda en pickles.
    Anota SVs germinales con la presencia/ausencia en somático y las guarda.

    Args:
        folder_path (str): Ruta a la carpeta que contiene los archivos.
        folder_output (str): Ruta a la carpeta para guardar el resultado.
        family_id (str): ID de la familia/probando.
        node (str): Nodo de secuenciación.
        proband_dict (dict): SVs del probando.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    logging.info("PASO 2.8: Anotando SVs germinales del probando con SVs somáticas")

    proband_output = os.path.join(folder_output, f'{family_id}_{family_id}_proband_tumour.pkl')
    
    # Carpeta con los archivos de SVs somáticas
    if node == 'FPGMX':
        parent_folder = os.path.dirname(os.path.normpath(folder_path))
        somatic_path = os.path.join(parent_folder, 'INPUTS', 'SOMATIC_SV')
    else:
        somatic_path = os.path.join(folder_path, 'INPUTS', 'SOMATIC')

    # Obtener las muestras tumorales esperadas
    expected_tumor_ids = {
        '-'.join([key.split('-')[0], key.split('-')[-1]])
        for key, val in pheno_dict.items()
        if 'ADN tumor' in val.get('Muestra', '')
    }

    if not expected_tumor_ids and not os.path.exists(somatic_path):
        logging.info("Esta familia no tiene muestras tumorales")
        # Asignar 'NA' a la columna 'Tumour' del probando y guardar
        for sv_id in proband_dict:
            proband_dict[sv_id]['Tumour'] = 'NA'
        save_pickle(proband_dict, proband_output)
        return
    
    if expected_tumor_ids and not os.path.exists(somatic_path):
        logging.error(
            f'Se esperan archivos somáticos pero no existe la carpeta: {somatic_path}. '
            f'Se detiene la ejecución.'
        )
        raise SystemExit(1)

    if not expected_tumor_ids and os.path.exists(somatic_path):
        logging.error('Hay carpeta de somático pero no se esperan muestras tumorales. '
                      'Se detiene la ejecución.')
        raise SystemExit(1)

    # Mapa de sexo por ID corto de muestra (1: male, 2: female)
    ped_file = os.path.join(os.path.dirname(folder_path.rstrip('/')), f"{family_id}.ped")
    ped_dict = parse_ped_file(ped_file)
    sex_by_short_id = {
        sample_id.split('-')[0]: sample_info.get('sex')
        for sample_id, sample_info in ped_dict.items()
    }
    
    # Rutas a los archivos por muestra somática
    samples = collect_somatic_files(node, somatic_path, pheno_dict)
    
    # Comprobar que todas las muestras tumorales esperadas tienen archivos
    found_tumor_ids = {tumor_id for tumor_id, _ in samples}
    missing_tumors = expected_tumor_ids - found_tumor_ids
    if missing_tumors:
        logging.error(
            f'No se encontraron archivos de las muestras tumorales {", ".join(missing_tumors)}. '
            f'Se detiene la ejecución.'
        )
        raise SystemExit(1)

    # Comprobar que existen todos los archivos somáticos
    required_types_by_node = {
        'CNAG': {'TO', 'TN'},
        'NASERTIC': {'TO', 'TN'},
        'FPGMX': {'TO', 'TN', 'PON'}
    }
    required_types = required_types_by_node.get(node, {'TO', 'TN', 'PON'})

    missing_paths = []
    for tumor_id, paths in samples:
        for tumor_type in required_types:
            path = paths.get(tumor_type)
            if not path or not os.path.exists(path):
                missing_paths.append(f'{tumor_id}:{tumor_type}')

    if missing_paths:
        logging.error(
            'Faltan los archivos somáticos: %s. Se detiene la ejecución.',
            ', '.join(sorted(missing_paths))
        )
        raise SystemExit(1)
    
    # Iniciar diccionarios para IntervalTrees de SVs somáticas
    to_trees = {}
    tn_trees = {}
    pon_trees = {}

    # Procesar archivos somáticos
    somatic_dict = {}

    for tumor_id, paths in samples:
        somatic_dict[tumor_id] = {}
        tumor_short_id = tumor_id.split('-')[0]
        is_female = sex_by_short_id.get(tumor_short_id) == 2
        for tumor_type, path in paths.items():
            # Procesar cada archivo de somático
            if path and os.path.exists(path):
                somatic_data = process_somatic_sv_file(path,f'{tumor_id}_{tumor_type}',
                                                       proband_dict, exclude_chr_y=is_female)

                # Añadir al diccionario
                somatic_dict[tumor_id][f'{tumor_id}_{tumor_type}'] = somatic_data
        
                # Construir IntervalTree
                if somatic_data:
                    if tumor_type == 'TO':
                        to_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
                    elif tumor_type == 'TN':
                        tn_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
                    elif tumor_type == 'PON':
                        pon_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
    
    # Anotar cada SV germinal con la presencia/ausencia en somático
    for sv_id, sv_data in proband_dict.items():
        for tumor_id, tree in to_trees.items():
            proband_dict[sv_id][f'{tumor_id}_TO'] = 'yes' if check_sv_in_somatic(sv_data, tree) else 'no'
        for tumor_id, tree in tn_trees.items():
            proband_dict[sv_id][f'{tumor_id}_TN'] = 'yes' if check_sv_in_somatic(sv_data, tree) else 'no'
        for tumor_id, tree in pon_trees.items():
            proband_dict[sv_id][f'{tumor_id}_PON'] = 'yes' if check_sv_in_somatic(sv_data, tree) else 'no'
    
    # Guardar SVs anotadas del probando y CVNs somáticas
    save_pickle(proband_dict, proband_output)

    somatic_file = os.path.join(folder_output, f'{family_id}_family_tumours.pkl')
    save_pickle(somatic_dict, somatic_file)

    logging.info("Anotación somática completada")


# ---------- PASO 2.9: Anotación con CNVs ---------- #
def run_annotate_cnvs_overlap(folder_path, folder_outputs, family_id):
    """
    Ejecuta paso 2.9: Anota SVs con solapamiento recíproco de CNVs en el
    probando. Guarda el diccionario actualizado (campo 'CNV match').

    Args:
        folder_path (str): Ruta a la carpeta que contiene los archivos.
        folder_outputs (str): Ruta a la carpeta para guardar el resultado.
        family_id (str): ID de la familia/probando.
    """
    logging.info("PASO 2.9: Anotando SVs del probando con CNVs")

    # Cargar SVs del probando
    proband_file = f'{folder_outputs}{family_id}_{family_id}_proband_tumour.pkl'
    proband_dict = load_pickle(proband_file)

    # Cargar CNVs del probando (archivo preprocesado en el script 'CNV_parse')
    folder_up = os.path.dirname(os.path.normpath(folder_path))
    pattern = f'{folder_up}/CNV/INTERMEDIATES/{family_id}-{family_id}-4impact-*.CNVs.annotated_parsed.pkl'
    matches = glob.glob(pattern)
    cnvs_file = matches[0] if matches else None

    cnvs_dict = {}
    if cnvs_file:
        try:
            cnvs_dict = load_pickle(cnvs_file)
            logging.info("Archivo de CNVs del probando cargado")
        except Exception:
            logging.error(f"Error al cargar el archivo de CNVs del probando desde {cnvs_file}")
    else:
        logging.warning(f"No se encontró el archivo de CNVs del probando en {pattern}")
    
    # Crear índice de CNVs con IntervalTree
    cnvs_index = build_chrom_interval_tree(cnvs_dict)
    
    # Anotar SVs con solapamiento de CNVs
    for sv_id, data in proband_dict.items():
        chrom = data.get('SV_chrom')

        match_ids = []
        
        # Comprobar si la SV está en el archivo de CNVs
        if chrom in cnvs_index:
            overlaps = cnvs_index[chrom]
            for overlap in overlaps:
                cnv_id, cnv_start, cnv_end, cnv_type, cnv_length = overlap.data
                
                # Crear diccionario para la CNV
                cnv_data = {
                    'SV_start': cnv_start,
                    'SV_end': cnv_end,
                    'SV_length': cnv_length,
                    'SV_type': cnv_type
                }

                # Verificar si hay solapamiento recíproco por encima del umbral
                if check_reciprocal_overlap(data, cnv_data):
                    match_ids.append(cnv_id)

        # Añadir resultado al diccionario y guardar
        if not match_ids:
            data['CNV match'] = None
        elif len(match_ids) == 1:
            data['CNV match'] = match_ids[0]
        else:
            data['CNV match'] = '; '.join(match_ids)

    output_file = f'{folder_outputs}{family_id}_{family_id}_proband_svs_annotated.pkl'
    save_pickle(proband_dict, output_file)
    logging.info("Anotación con CNVs completada")


# -------------------------------------------------
# PASO 3: GENERACIÓN DEL INFORME FINAL
# -------------------------------------------------

def translate_status(status):
    """Traduce el estado numérico del estatus a un caracter."""
    status_map = {1: 'unaffected', 2: 'affected'}
    return status_map.get(status, 'NA')


def translate_sex(sex):
    """Traduce el código numérico de la variable sexo a caracter."""
    sex_map = {1: 'male', 2: 'female'}
    return sex_map.get(sex, 'NA')


def length_to_int(dct):
    """
    Comprueba o calcula la longitud de cada variante en el diccionario y la
    convierte a tipo entero. Los valores vacíos ('') se convierten en None.
    
    Args:
        dct (dict): Diccionario con las variantes (value = dict o list[dict]).
    
    Returns:
        dict: Diccionario actualizado con la longitud correcta como entero.
    """
    for variant in dct.values():
        if isinstance(variant, list):
            variant = variant[0]
        if variant['SV_length'] == '':
            variant['SV_length'] = None
        else:
            try:
                variant['SV_length'] = int(float(variant['SV_length']))
            except (TypeError, ValueError, KeyError):
                try:
                    start = int(variant['SV_start'])
                    end = int(variant['SV_end'])
                    variant['SV_length'] = max(0, end - start)
                except (TypeError, ValueError, KeyError):
                    logging.warning(f'Valor inesperado {variant["SV_length"]}.')
    return dct


def dict_to_excel(wb, data_dict, sheet_name):
    """Inserta un diccionario de SVs en una hoja de Excel nueva con encabezados."""
    if not data_dict:
        return  # No hacer nada si el diccionario de SVs está vacío
    
    # Crear nueva hoja de Excel
    ws = wb.create_sheet(title=sheet_name)

    # Obtener las claves del diccionario para usarlas como encabezados
    headers = list(next(iter(data_dict.values())).keys())
    headers = headers[1:] # Quitar 'AnnotSV_ID'
    ws.append(['SV ID'] + headers)  # Agregar encabezado

    # Agregar filas con los valores
    for annot_id, values in data_dict.items():
        row = [annot_id] + [', '.join(map(str, values[col])) if isinstance(values[col], list) else values[col] for col in headers]
        ws.append(row)


def build_sv_row(sv_id, value, node, sample_headers, sv_dict, proband=None, pheno_dict=None, somatic_headers=None, acmg_dict=ACMG_CLASSIFICATION):
    """
    Construye la fila de Excel para una SV de un probando o familiar.
    
    Args:
        sv_id (str): ID de la SV.
        value (dict): Datos de la SV.
        node (str): Nodo de secuenciación.
        sample_headers (list): Encabezados para los genotipos.
        sv_dict (dict): SVs para acceder a genotipos y datos somáticos.
        proband (str, optional): ID del probando (Default None).
        pheno_dict (dict, optional): Fenotipos de las muestras (Default None).
        somatic_headers (list, optional): Encabezados somáticos (Default None).
        acmg_dict (dict): Anotación de ACMG para la clasificación de AnnotSV.
    
    Returns:
        list: Datos de la SV formateados para una fila de Excel.
    """
    chrom = value.get('SV_chrom', '')
    start = value.get('SV_start', '')
    end = value.get('SV_end', '')
    length = value.get('SV_length', '')
    sv_type = value.get('SV_type', '')
    
    try:
        annot_classification = int(value.get('ACMG_class', ''))
    except (ValueError, TypeError):
        annot_classification = None
    annot_classification = acmg_dict.get(annot_classification, 'VUS')
    
    try:
        annot_score = float(value.get('AnnotSV_ranking_score', 0.0))
    except (ValueError, TypeError):
        annot_score = 0.0
    
    annot_criteria = value.get('AnnotSV_ranking_criteria', '')
    n_genes = value.get('Gene_count')
    genes = value.get('Gene_name')
    
    # Convertir lsita de genes en una cadena de texto
    hereditary = ', '.join([f"{item['cancer gene']} ({item['gene type']})" for item in value.get('hereditary cancer gene list', [])])
    cancer_cosmic = ', '.join([f"{item['cancer gene']} ({item['gene type']})" for item in value.get('cancer cosmic', [])])
    
    omim_ids = value.get('OMIM_ID', '')
    uniprot = value.get('UniProt_Functions', '')
    kegg = value.get('KEGG_pathways', '')
    reactome = value.get('Reactome_pathways', '')
    
    # Genotipo de muestras
    sample_details = []
    for header in sample_headers:
        smpl = header.split(' ')[0]
        # Probando
        if proband:
            if smpl == proband.split('-')[0]:
                if node == 'FPGMX':
                    gt_prob = sv_dict[sv_id]['Zygosity']
                else:
                    prob = sv_dict[sv_id][proband]
                    gt_prob = prob.split(':')[0]
                sample_details.append(f'GT:{gt_prob}')
            else:
                smpl_full = get_full_id(smpl, pheno_dict)
                if smpl_full in sv_dict[sv_id]:
                    sample_info = sv_dict[sv_id][smpl_full]
                    sample_details.append(sample_info)
                else:
                    sample_details.append('')

        # Familiares
        else:
            smpl_full = get_full_id(smpl, pheno_dict)
            if smpl_full in sv_dict[sv_id]:
                sample_info = sv_dict[sv_id][smpl_full]
                sample_details.append(sample_info)
            else:
                sample_details.append('')
    
    # Datos somáticos
    somatic_details = []
    if somatic_headers:
        for sh in somatic_headers:
            if sh in sv_dict[sv_id]:
                somatic_details.append(sv_dict[sv_id][sh])
            else:
                somatic_details.append('')
    
    exact_counts = value.get('Illumina.exact.counts' if node == 'CNAG' else 'Illumina_DRAGEN.exact.counts', '')
    similar_counts = value.get('Illumina.similar.counts' if node == 'CNAG' else 'Illumina_DRAGEN.similar.counts', '')
    other_counts = value.get('Illumina.other.counts' if node == 'CNAG' else 'Illumina_DRAGEN.other.counts', '')
    gnomad = value.get('gnomAD', '')
    gnomad_id = value.get('gnomad_id', '')
    gnomad_AC = value.get('gnomAD_AC', '')
    gnomad_AF = value.get('gnomAD_AF', '')
    # gnomad_AN = value.get('gnomAD_AN', '')
    gnomad_FREQ_HOMALT = value.get('gnomAD_FREQ_HOMALT', '')

    row = [sv_id, chrom, start, end, length, sv_type, annot_classification, annot_score, annot_criteria, 
           n_genes, genes, hereditary, cancer_cosmic, omim_ids, uniprot, kegg, reactome] + sample_details + \
           somatic_details + [exact_counts, similar_counts, other_counts, gnomad, gnomad_id, gnomad_AC, gnomad_AF, gnomad_FREQ_HOMALT]
    
    if proband:
        row.append(value.get('CNV match'))

        overlapped_svs = value.get('Merged_variants')
        overlapped_svs = ','.join(map(str, overlapped_svs)) if isinstance(overlapped_svs, list) else overlapped_svs
        row.append(overlapped_svs)
    
    return row


def apply_conditional_format(ws):
    """Aplica formato condicional a las columnas con los resultados de
    AnnotSV para resaltar SVs usando reglas de Excel."""
    max_row = ws.max_row

    col_rules = {
        # Clasificación de AnnotSV (columna G)
        'G': [("G2=\"P\"", EXCEL_COLORS['red']),
              ("G2=\"LP\"", EXCEL_COLORS['orange']),
              ("G2=\"VUS\"", EXCEL_COLORS['yellow'])],
        # Puntuación de AnnotSV (columna H)
        'H': [("H2>=0.99", EXCEL_COLORS['red']),
              ("AND(H2>=0.9,H2<0.99)", EXCEL_COLORS['orange']),
              ("AND(H2>=-0.89,H2<=0.89)", EXCEL_COLORS['yellow'])]
    }
 
    for col, rules in col_rules.items():
        for formula, fill in rules:
            ws.conditional_formatting.add(f'{col}2:{col}{max_row}', 
                                          FormulaRule(formula=[formula], fill=fill))


def adjust_columns_width(ws, custom_widths=None, max_columns=None):
    """
    Ajusta automáticamente el ancho de las columnas de un worksheet.
    
    Args:
        ws: Worksheet de openpyxl.
        custom_widths: Diccionario opcional {header_name: width}.
        max_columns: Número máximo de columnas a ajustar.
    """
    for idx, col in enumerate(ws.columns):
        if max_columns is not None and idx >= max_columns:
            break

        max_length = 0
        column_letter = get_column_letter(col[0].column)
        header = col[0].value

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                continue

        # Ancho por defecto con un margen adicional
        adjusted_width = (max_length + 2)

        # Sobreescribir anchos específicos si se proporcionan
        if custom_widths and header in custom_widths:
            adjusted_width = custom_widths[header]
        
        # Ajustar columnas de genotipos por patrón
        if header and ('Status' in header):
            adjusted_width = 60

        # Aplicar cambios a la hoja de Excel
        ws.column_dimensions[column_letter].width = adjusted_width


def annotate_split_svs_hereditary(split_dict, hereditary_file=HEREDITARY_CANCER_FILE):
    """Añade 'Hereditary cancer genes' a cada fila de SVs 'split'."""
    with open(hereditary_file, mode='r', encoding='iso-8859-1') as f:
        hereditary_genes = {row[0].strip() for row in csv.reader(f) if row and row[0].strip()}

    for annot_id, rows in split_dict.items():
        new_rows = []
        for row in rows:
            gene = row['Gene_name'].strip()
            new_row = {}
            for key, value in row.items():
                new_row[key] = value
                if key == 'Gene_name':
                    new_row['Hereditary cancer genes'] = gene if gene in hereditary_genes else ''
            new_rows.append(new_row)
        split_dict[annot_id] = new_rows

    return split_dict


def create_excel_report(folder_path, node, family_id, ped_dict, pheno_dict, 
                        proband_dict, relatives_dict, prob_split_dict, somatic_dict, output_file):
    """
    Genera un archivo Excel con SVs anotadas en varias hojas: 'Probad' (SVs
    probando), 'Proband_split' (SVs 'split' probando, 'Relatives' (SVs 
    familiares), '{IDrelative}_split' (SVs 'split' familiares) y datos somáticos.

    Args:
        folder_path (str): Ruta a la carpeta que contiene los archivos.
        node (str): Nodo de secuenciación de procedencia.
        family_id (str): ID de la familia.
        ped_dict (str): Diccionario PED de las muestras.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
        proband_dict (dict): SVs de probando.
        relatives_dict (dict): SVs únicas de familiares.
        somatic_dict (dict): SVs somáticas.
        output_file (str): Nombre del archivo Excel.
    """
    # ---------- CREAR Y GUARDAR ARCHIVO EXCEL ----------
    # Crear un nuevo libro de Excel con el glosario (mantiene formato y estilos)
    wb = openpyxl.load_workbook(GLOSSARY_FILE)
    logging.info("Archivo Excel con el glosario creado")

    # Añadir hoja con el resumen
    ws_summary = add_summary_sheet(wb, node, family_id, ped_dict, pheno_dict,
                                   proband_dict, prob_split_dict, relatives_dict, somatic_dict,
                                   variant_label='SV')
    logging.info("Hoja de resumen añadida al Excel")

    # Crear nueva hoja para el probando
    ws_proband = wb.create_sheet('Proband')

    # ---------- 'Proband': ENCABEZADOS ----------
    # Obtener IDs de las muestras para las columnas de genotipo
    sample_headers = []
    for key in ped_dict.keys():
        smpl = key.split('-')[0]
        smpl_id = get_full_id(smpl, pheno_dict)

        if not smpl_id:
            logging.warning(f'No hay muestra germinal para {key}')
            continue
        
        # Añadir información de fenotipo
        status = translate_status(ped_dict[key].get('status', 'NA'))
        if status == 'affected':
            pheno = pheno_dict[smpl_id]['Subpatología']
            status = f'{status} ({pheno})'
        sex = translate_sex(ped_dict[key].get('sex', 'NA'))
        kinship = pheno_dict[smpl_id]['Relación familiar']

        if smpl == family_id:
            sample_headers.append((int(smpl), f'{smpl} - proband (Status: {status}, Sex: {sex})'))
        else:
            sample_headers.append((int(smpl), f'{smpl} (Status: {status}, Sex: {sex}, Kinship: {kinship})'))
    
    # Ordenar los encabezados de las muestras
    sample_headers = sorted(sample_headers, key=lambda x: x[0])
    sample_headers = [header[1] for header in sample_headers]

    # Eliminar duplicados manteniendo el orden
    sample_headers = list(dict.fromkeys(sample_headers))
    
    # Obtener IDs de las muestras tumorales
    somatic_headers = []
    for smpl in somatic_dict:
        somatic_headers.extend(somatic_dict[smpl].keys())
    
    # Encabezados del probando
    common_headers = ['SV ID', 'Chrom', 'Start', 'End', 'Length', 'Type',
                      'AnnotSV classification', 'AnnotSV score', 'AnnotSV criteria',
                      'Gene count', 'Genes', 'Hereditary cancer genes', 'Cancer Cosmic',
                      'OMIM', 'UniProt functions', 'KEGG pathways', 'Reactome pathways']
    proband_headers = somatic_headers + ['Exact counts', 'Similar counts', 'Other counts', 'gnomAD', 'gnomAD_ID', 'gnomAD_AC', 'gnomAD_AF', 'gnomAD_FREQ_HOMALT', 'Match in CNVs', 'Overlapped SVs']
    ws_proband.append(common_headers + sample_headers + proband_headers)

    # ---------- 'Proband': INFORMACIÓN SVs DEL PROBANDO ----------
    # Ordenar SVs del probando
    sorted_svs = sorted(
        proband_dict.keys(),
        key=lambda k: (
            -float(proband_dict[k].get('AnnotSV_ranking_score', '0') or '0'),  # Ordenar por AnnotSV score
            -float(proband_dict[k].get('SV_length', '0') or '0')               # Ordenar por longitud
        )
    )
    
    # Agregar datos del probando al archivo Excel
    proband = get_proband_id(family_id, pheno_dict)
    for sv_id in sorted_svs:
        value = proband_dict[sv_id]
        row = build_sv_row(sv_id=sv_id, value=value, node=node, sample_headers=sample_headers,
                           sv_dict=proband_dict, proband=proband, pheno_dict=pheno_dict, somatic_headers=somatic_headers)
        ws_proband.append(row)

    logging.info("Hoja del probando añadida al Excel")

    # ---------- 'Proband_split': CREAR HOJA Y AÑADIRENCABEZADOS ----------
    # Obtener los encabezados
    excluded_fields = ['AnnotSV_ID', 'Samples_ID', 'Closest_left', 'Closest_right', 'Gene_count',
                       'RE_gene', 'po_P_gain_phen', 'po_P_gain_hpo', 'po_P_gain_source', 
                       'po_P_gain_coord', 'po_P_gain_percent', 'po_P_loss_phen', 'po_P_loss_hpo', 
                       'po_P_loss_source', 'po_P_loss_coord', 'po_P_loss_percent', 
                       'po_B_gain_allG_source', 'po_B_gain_allG_coord', 'po_B_gain_someG_source', 
                       'po_B_gain_someG_coord', 'po_B_loss_allG_source', 'po_B_loss_allG_coord', 
                       'po_B_loss_someG_source', 'po_B_loss_someG_coord', 'GC_content_left', 
                       'GC_content_right', 'Repeat_coord_left', 'Repeat_type_left', 'Repeat_coord_right', 
                       'Repeat_type_right', 'Gap_left', 'Gap_right', 'SegDup_left', 'SegDup_right', 
                       'ENCODE_blacklist_left', 'ENCODE_blacklist_characteristics_left', 
                       'ENCODE_blacklist_right', 'ENCODE_blacklist_characteristics_right', 
                       'AnnotSV_ranking_score', 'AnnotSV_ranking_criteria', 'ACMG_class']
    
    prob_split_dict = annotate_split_svs_hereditary(prob_split_dict)
    sample_split_headers = list(next(iter(prob_split_dict.values()))[0].keys())
    split_headers = ['AnnotSV ID'] + [key for key in sample_split_headers if key not in excluded_fields]
    
    # Crear nueva hoja en el Excel y agregar encabezados
    ws_prob_split = wb.create_sheet(title='Proband_split')
    ws_prob_split.append(split_headers)
    
     # ---------- 'Proband_split': SVs 'SPLIT' ----------
    # Ordenar SVs por cromosoma y posición
    sorted_prob_split = {}
    for sv_id in natsorted(prob_split_dict.keys(), key=lambda id: (prob_split_dict[id][0]['SV_chrom'], int(prob_split_dict[id][0]['SV_start']))):
        sorted_prob_split[sv_id] = prob_split_dict[sv_id]
    
    # Añadir los datos al archivo Excel
    for sv_id, rows in sorted_prob_split.items():
        for row in rows:
            row_data = [sv_id] + [value for key, value in row.items() if key not in excluded_fields]
            ws_prob_split.append(row_data)
    logging.info("Hoja con las SVs 'split' del probando añadida al Excel")

    # ---------- 'Relatives': CREAR HOJA Y AÑADIR ENCABEZADOS ----------
    ws_relatives = None
    sample_headers_rel = sample_headers[1:]  # Excluir probando
    if relatives_dict:
        ws_relatives = wb.create_sheet('Relatives')
        rel_headers = common_headers + sample_headers_rel + ['Exact counts', 'Similar counts', 'Other counts', 'gnomAD', 'gnomAD_ID', 'gnomAD_AC', 'gnomAD_AF', 'gnomAD_FREQ_HOMALT']
        ws_relatives.append(rel_headers)

        # ---------- 'Relatives': INFORMACIÓN SVs ----------
        # Ordenar SVs
        rel_sorted_svs = sorted(
            relatives_dict.keys(),
            key=lambda k: (
                -float(relatives_dict[k].get('AnnotSV_ranking_score', '0') or '0'),  # Ordenar por AnnotSV score
                -float(relatives_dict[k].get('SV_length', '0') or '0')               # Ordenar por longitud
            )
        )
    
        # Agregar datos de cada familiar al Excel
        for sv_id in rel_sorted_svs:
            value = relatives_dict[sv_id]
            row = build_sv_row(sv_id=sv_id, value=value, node=node, sample_headers=sample_headers_rel,
                            sv_dict=relatives_dict, pheno_dict=pheno_dict)
            ws_relatives.append(row)
        logging.info("Hoja de los familiares añadida al Excel")
    
        # ---------- 'Relatives_split': SVs 'split' ----------
        for rel in ped_dict:
            rel_smpl = rel.split('-')[0]
            if rel_smpl == family_id:
                continue
            # Cargar archivos 'split' si existen
            pattern = os.path.join(folder_path, f'{rel_smpl}-{family_id}-4impact-*.SVs.annotated_parsed_split.pkl')
            matches = glob.glob(pattern)
            if not matches:
                logging.warning(f"Falta el archivo de SVs 'split' de {rel_smpl}")
                continue

            rel_split_dict = load_pickle(matches[0])
            rel_split_dict = annotate_split_svs_hereditary(rel_split_dict)
            
            # Crear nueva hoja en el Excel y agregar encabezados
            ws_rel_split = wb.create_sheet(title=f'{rel_smpl}_split')
            ws_rel_split.append(split_headers)
            
            # Ordenar SVs por cromosoma y posición
            sorted_rels_split = {}
            for sv_id in natsorted(rel_split_dict.keys(), key=lambda id: (rel_split_dict[id][0]['SV_chrom'], int(rel_split_dict[id][0]['SV_start']))):
                sorted_rels_split[sv_id] = rel_split_dict[sv_id]
            
            # Añadir datos al Excel
            for sv_id, rows in sorted_rels_split.items():
                for row in rows:
                    row_data = [sv_id] + [value for key, value in row.items() if key not in excluded_fields]
                    ws_rel_split.append(row_data)

        logging.info("Hojas con las SVs 'split' de los familiares añadidas al Excel")
    else:
        logging.info("No hay familiares, no se crea la hoja 'Relatives'")

    # ----------------- SOMATIC: SVs -----------------
    for smpl, sample_somatic_dict in somatic_dict.items():
        for somatic_id, somatic_data in sample_somatic_dict.items():
            if somatic_data is not None:
                sheet_name = ''
                if 'TO' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_Only'
                elif 'TN' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_vs_Normal'
                elif 'PON' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_vs_PON'
    
                if sheet_name:
                    sorted_somatic = {}
                    for sv_id in natsorted(
                        somatic_data.keys(),
                        key=lambda id: (
                            somatic_data[id]['SV_chrom'], 
                            int(somatic_data[id]['SV_start']))
                    ):
                        sorted_somatic[sv_id] = somatic_data[sv_id]

                    dict_to_excel(wb, sorted_somatic, sheet_name)
    if somatic_dict:
        logging.info("Hojas con las SVs somáticas añadidas al Excel")
    
    # ---------- FORMATO CONDICIONAL ----------
    # Resaltar SVs según su puntuación de AnnotSV
    apply_conditional_format(ws_proband)
    if ws_relatives is not None:
        apply_conditional_format(ws_relatives)

    # ---------- AJUSTAR ANCHO DE COLUMNA ----------
    # Definir anchos personalizados para columnas específicas
    custom_widths = {
        'Length': 11,
        'AnnotSV classification': 20,
        'AnnotSV criteria': 18,
        'Genes': 20,
        'Hereditary cancer genes': 15.3,
        'Cancer Cosmic': 15,
        'OMIM': 15,
        'UniProt functions': 25, 'KEGG pathways': 25, 'Reactome pathways': 25,
        'Other counts': 25,
        'gnomAD_FREQ_HOMALT': 22,
        'Match in CNVs': 26, 'Overlapped SVs': 26
    }

   # Ajustar ancho de las columnas de 'Proband' y 'Relatives'
    adjust_columns_width(ws_proband, custom_widths)
    if ws_relatives is not None:
        adjust_columns_width(ws_relatives, custom_widths)

    # Ajustar las primeras 5 columnas en las hojas 'split' y somáticas
    for ws in wb.worksheets:
        if ws.title.endswith('_split'):
            adjust_columns_width(ws, max_columns=5)
        if any(suffix in ws.title for suffix in ['_Tumour_Only', '_Tumour_vs_Normal', '_Tumour_vs_PON']):
            adjust_columns_width(ws, max_columns=5, custom_widths={'SV_length': 15})
    
    # Ajustar ancho de las primeras columnas del resumen
    adjust_columns_width(ws_summary, max_columns=5)

    # ---------- GUARDAR EL ARCHIVO EXCEL ----------
    wb.save(output_file)
    logging.info(f'Archivo Excel completado y guardado en: {output_file}')


def run_generate_excel_report(folder_intermediates, output_folder, node, family_id, ped_dict, pheno_dict):
    """
    Ejecuta paso 3: Genera un informe Excel con SVs anotadas del probando y
    familiares. Integra datos somáticos, germinales, genes de cáncer y otras
    anotaciones con glosario.
    """
    # Cargar SVs del probando y familiares
    proband_file = os.path.join(output_folder, f'{family_id}_{family_id}_proband_svs_annotated.pkl')
    proband_dict = load_pickle(proband_file)

    relatives_file = os.path.join(output_folder, f'{family_id}_relatives_svs_annotated.pkl')
    if os.path.exists(relatives_file):
        relatives_dict = load_pickle(relatives_file)
    else:
        relatives_dict = {}

    # Obtener la longitud como número entero
    proband_dict = length_to_int(proband_dict)
    relatives_dict = length_to_int(relatives_dict)

    # Cargar SVs 'split' del probando
    pattern = os.path.join(folder_intermediates, f'{family_id}-{family_id}-4impact-*.SVs.annotated_parsed_split.pkl')
    matches = glob.glob(pattern) 
    prob_split_dict = load_pickle(matches[0])

    # Archivo de SVs somáticas (ausencia implica que no hay tumor, ya se validó en el paso 2.8)
    somatic_file = os.path.join(output_folder, f'{family_id}_family_tumours.pkl')
    if os.path.exists(somatic_file):
        somatic_dict = load_pickle(somatic_file)
    else:
        somatic_dict = {}
        logging.info("Familia sin muestras tumorales. No se incluirán hojas somáticas en el Excel.")

    # Crear y guardar el archivo Excel
    output_file = os.path.join(output_folder, f'{family_id}_prioritized_SVs.xlsx')
    create_excel_report(folder_intermediates, node, family_id, ped_dict, pheno_dict,
                        proband_dict, relatives_dict, prob_split_dict, somatic_dict, output_file)


# ---------- FUNCIÓN GLOBAL ----------
def main(folder_path):
    """
    Función principal que orquesta la ejecucion completa del algoritmo de SVs.

    Args:
        folder_path (str): Ruta a la carpeta de análisis de la familia.
    """
    # ---------- Preparar variables y archivos ----------
    # Nodo de secuenciación e ID del probando/familia
    node, family_id = split_analysis_folder(folder_path)

    # Archivo correspondencia muestras-fenotipo
    pheno_dict = load_phenotype_data(SAMPLE_PHENO_FILE, family_id)
    
    # Archivo PED
    ped_file = os.path.join(os.path.dirname(folder_path.rstrip('/')), f"{family_id}.ped")
    ped_dict = parse_ped_file(ped_file)

    # ID del probando
    proband = get_proband_id(family_id, pheno_dict)

    # ---------- Configurar carpetas de entrada y salida ----------
    folder_root = os.path.dirname(os.path.normpath(folder_path)) if node == 'FPGMX' else folder_path
    folder_inputs = os.path.join(folder_root, 'INPUTS/')
    folder_intermediates = os.path.join(folder_path, 'INTERMEDIATES/')
    folder_outputs = os.path.join(folder_path, 'OUTPUTS/')

    for folder in [folder_intermediates, folder_outputs]:
        os.makedirs(folder, exist_ok=True)
        if os.path.isdir(folder):
            logging.info(f"La carpeta '{folder}' ya existe")
        else:
            logging.info(f"Creada carpeta '{folder}'")

    # Paso 1: Procesamiento de SVs de probando y familiares
    logging.info("INICIANDO EL PREPROCESAMIENTO DE SVs (PASO 1)")
    run_preprocess_sv(folder_inputs, folder_intermediates, node, family_id, pheno_dict)
    logging.info("PREPROCESAMIENTO DE SVs COMPLETADO")

    # Paso 2: Priorización y anotación de SVs
    logging.info("INICIANDO EL PROCESO DE ANOTACIÓN Y PRIORIZACIÓN DE SVs (PASO 2)")
    
    # Paso 2.1: Filtrar SVs solapantes del probando
    run_filter_overlapping_sv(family_id, proband, folder_intermediates, folder_outputs)
    # Paso 2.2: Anotación de SVs del probando compartidas con familiares
    run_annotate_shared_svs(folder_intermediates, folder_outputs, family_id, node, proband, pheno_dict)
    # Paso 2.3: Combinación de SVs de los familiares
    rels_svs_index = run_merge_relatives_svs(folder_outputs, family_id, node, pheno_dict)
    # Paso 2.4: Filtrar SVs que solapan en los familiares
    run_filter_relatives_overlapping_svs(family_id, rels_svs_index, folder_outputs, node, pheno_dict)
    # Paso 2.5: Anotación de genes relacionados con cáncer
    proband_dict, rels_dict = run_annotate_cancer_genes(folder_outputs, family_id)
    # Paso 2.6: Anotación de pathways
    proband_dict, rels_dict = run_annotate_pathways(proband_dict, rels_dict)
    # Paso 2.7: Anotación con gnomAD
    proband_dict = run_annotate_gnomad(folder_outputs, family_id, proband_dict, rels_dict)
    # Paso 2.8: Procesamiento y anotación de SVs somáticas
    run_process_somatic(folder_path, folder_outputs, family_id, node, proband_dict, pheno_dict)
    # Paso 2.9: Anotación con CNVs
    run_annotate_cnvs_overlap(folder_path, folder_outputs, family_id)

    # PASO 3: Generar informe final
    logging.info("INICIO DE LA GENERACIÓN DEL INFORME FINAL (PASO 3)")
    run_generate_excel_report(folder_intermediates, folder_outputs, node, family_id, ped_dict, pheno_dict)
    logging.info("GENERACIÓN DEL INFORME FINAL COMPLETADA")

    logging.info("ANÁLISIS SVs FINALIZADO")
    logging.info(f"FAMILIA {family_id} PROCESADA")


if __name__ == '__main__':
    # folder_path = 'C:/Users/edurne.urrutia/Desktop/VUSCAN_PIPELINES/DATA/SAMPLES/CNAG/1282/SV/'
    WORKDIR = sys.argv[1]
    folder_path = f'{WORKDIR}SV/'

    if not os.path.isdir(folder_path):
        logging.error(f"La ruta introducida '{folder_path}' no es válida o no existe.")
        sys.exit(1)
    else:
        main(folder_path)
