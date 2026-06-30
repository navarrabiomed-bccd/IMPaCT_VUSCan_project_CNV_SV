# -*- coding: utf-8 -*-
"""Helpers to build the shared Excel summary sheet for CNV and SV reports."""

from datetime import datetime
from openpyxl.styles import Font, Alignment


TUMOUR_TYPES = ('TO', 'TN', 'PON')
STATUS_LABELS = {1: 'unaffected', 2: 'affected'}


def sort_member_ids(member_ids):
    """Sort numeric-like IDs numerically, preserving non-numeric IDs at the end."""
    numeric_ids = []
    other_ids = []
    for member_id in member_ids:
        if str(member_id).isdigit():
            numeric_ids.append(str(member_id))
        else:
            other_ids.append(str(member_id))

    numeric_ids = sorted(numeric_ids, key=lambda x: int(x))
    other_ids = sorted(other_ids)
    return numeric_ids + other_ids


def count_split_variant_gene_pairs(split_dict):
    """Count 'split' variants (variant-gene pairs)."""
    pairs = 0
    for rows in split_dict.values():
        if isinstance(rows, list):
            pairs += len(rows)
        elif rows is not None:
            pairs += 1
    return pairs


def get_member_status(ped_dict, family_id, member_id):
    """Return the affection status for a family member."""
    ped_key = f'{member_id}-{family_id}-4impact'
    status = ped_dict.get(ped_key, {}).get('status')
    return STATUS_LABELS.get(status, 'NA')


def get_member_status_counts(member_ids, ped_dict, family_id):
    """Return (affected_count, unaffected_count) for family members."""
    affected_count = 0
    unaffected_count = 0

    for member_id in member_ids:
        member_status = get_member_status(ped_dict, family_id, member_id)
        if member_status == 'affected':
            affected_count += 1
        elif member_status == 'unaffected':
            unaffected_count += 1

    return affected_count, unaffected_count


def get_member_samples(pheno_dict, family_id, member_id):
    """Return non-tumour samples for one family member."""
    return sorted([
        sample_id for sample_id in pheno_dict.keys()
        if sample_id.startswith(f'{member_id}-{family_id}-')
        and 'ARN tumor' not in pheno_dict[sample_id].get('Muestra', '')
    ])


def get_subpathology(sample_info, affected_status):
    """Return subpathology value for summary rows."""
    if affected_status != 'affected':
        return '-'
    return sample_info.get('Subpatologia', sample_info.get('Subpatología', 'NA'))


def count_rel_not_shared_by_member(relatives_dict, rel_id):
    """Count non-shared variants of a relative in the dictionary."""
    rel_not_shared = 0
    for variant_data in relatives_dict.values():
        if any(str(key).startswith(f'{rel_id}-') or str(key) == rel_id for key in variant_data.keys()):
            rel_not_shared += 1
    return rel_not_shared


def append_bold_row(worksheet, values):
    """Append one row and apply bold style on non-empty cells."""
    worksheet.append(values)
    row_idx = worksheet.max_row
    for col_idx, value in enumerate(values, start=1):
        if value not in [None, '']:
            worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)


def add_summary_sheet(wb, node, family_id, ped_dict, pheno_dict,
                      proband_dict, prob_split_dict, relatives_dict,
                      somatic_dict, variant_label):
    """Add the analysis summary sheet for CNV/SV pipelines."""
    member_ids = sort_member_ids({key.split('-')[0] for key in ped_dict.keys()})
    relative_ids = [member_id for member_id in member_ids if not member_id.startswith(family_id)]

    # Affected / unaffected member counts
    affected_count, unaffected_count = get_member_status_counts(member_ids, ped_dict, family_id)

    # Proband variants 'full' and 'split'
    full_variants = set(proband_dict.keys())
    split_variants = set(prob_split_dict.keys())
    split_variant_gene_pairs = count_split_variant_gene_pairs(prob_split_dict)

    # Variants 'full' without genes associated
    only_full = full_variants - split_variants
    prob_no_genes = set()
    for variant_id in only_full:
        prob_gene_count = proband_dict.get(variant_id, {}).get('Gene_count', None)
        if float(prob_gene_count) == 0:
            prob_no_genes.add(variant_id)

    # Shared proband variants by relatives
    shared_with_relatives = 0
    shared_by_relative = {rel_id: 0 for rel_id in relative_ids}
    for variant_data in proband_dict.values():
        shared_in_this_variant = False
        for rel_id in relative_ids:
            if any(str(key).startswith(f'{rel_id}-{family_id}-') or str(key) == rel_id for key in variant_data.keys()):
                shared_in_this_variant = True
                shared_by_relative[rel_id] += 1
        if shared_in_this_variant:
            shared_with_relatives += 1

    # Shared proband variants in tumour samples
    shared_with_tumour = 0
    shared_tumour_by_type = {tumour_type: 0 for tumour_type in TUMOUR_TYPES}
    if somatic_dict:
        for variant_data in proband_dict.values():
            tumour_type_found = False
            for tumour_type in TUMOUR_TYPES:
                if any(key.endswith(tumour_type) and variant_data.get(key) == 'yes' for key in variant_data):
                    if not tumour_type_found:
                        shared_with_tumour += 1
                        tumour_type_found = True
                    shared_tumour_by_type[tumour_type] += 1

    # Somatic variants
    tumour_variants = {}
    for sample, sample_somatic_dict in somatic_dict.items():
        tumour_variants[sample] = {tumour_type: 0 for tumour_type in TUMOUR_TYPES}
        for somatic_id, somatic_data in sample_somatic_dict.items():
            if somatic_data is not None:
                for tumour_type in TUMOUR_TYPES:
                    if somatic_id.endswith(f'_{tumour_type}'):
                        tumour_variants[sample][tumour_type] += len(somatic_data)
                        break

    # Variants detected in the other variant type
    prob_with_other_variant_type = sum(
        1 for v in proband_dict.values() if v.get('sv_match') or v.get('CNV match')
    )

    # Create new sheet for summary
    ws_summary = wb.create_sheet(title='Summary', index=1)

    # Add summary content to the sheet
    if variant_label == 'CNV':
        append_bold_row(ws_summary, ['CNVs SUMMARY'])
    elif variant_label == 'SV':
        append_bold_row(ws_summary, ['SVs SUMMARY'])
    
    ws_summary.append([])
    append_bold_row(ws_summary, ['FAMILY INFORMATION'])
    ws_summary.append(['Node', node])
    ws_summary.append(['Family ID', family_id])
    ws_summary.append(['Family members', str(len(member_ids))])
    ws_summary.append(['Affected members', str(affected_count)])
    ws_summary.append(['Unaffected members', str(unaffected_count)])
    ws_summary.append(['Date of analysis', datetime.now().strftime('%d-%m-%Y')])
    
    ws_summary.append([])
    append_bold_row(ws_summary, ['FAMILY SAMPLES ANALYZED'])
    append_bold_row(ws_summary, ['Member', 'Member ID', 'Affected status', 'Sample ID', 'Sample type', 'Subpathology'])
    for member_id in member_ids:
        affected_status = get_member_status(ped_dict, family_id, member_id)
        member_samples = get_member_samples(pheno_dict, family_id, member_id)

        if not member_samples:
            ws_summary.append([member_id, member_id, affected_status, 'NA', 'NA', 'NA'])
            continue

        for sample_id in member_samples:
            sample_info = pheno_dict.get(sample_id, {})
            ws_summary.append([
                sample_info.get('Relacion familiar', sample_info.get('Relación familiar', 'NA')),
                member_id,
                affected_status,
                sample_id,
                sample_info.get('Muestra', 'NA'),
                get_subpathology(sample_info, affected_status),
            ])
    
    ws_summary.append([])
    append_bold_row(ws_summary, [f'PROBAND {variant_label}s SUMMARY'])
    ws_summary.append([f'Total proband {variant_label}s', len(proband_dict)])
    if relative_ids:
        ws_summary.append([f'Proband {variant_label}s shared with >=1 relative', shared_with_relatives])
    if somatic_dict:
        if shared_with_tumour > 0:
            tumour_type_summary = [
                f'{tumour_type}: {count}'
                for tumour_type, count in shared_tumour_by_type.items()
                if count > 0
            ]
            ws_summary.append([
                f'Proband {variant_label}s detected in tumour samples',
                f'{shared_with_tumour} ({" | ".join(tumour_type_summary)})'
            ])
            # Alinear a la derecha la celda con el resumen de tumores
            row_idx = ws_summary.max_row
            ws_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
        else:
            ws_summary.append([f'Proband {variant_label}s detected in tumour samples', shared_with_tumour])
    if variant_label == 'CNV':
        other_variant_type = 'SV'
    else:
        other_variant_type = 'CNV'
    ws_summary.append([f'Proband {variant_label}s with {other_variant_type} match', prob_with_other_variant_type])
    ws_summary.append([f'Proband {variant_label}s without genes associated', len(prob_no_genes)])
    ws_summary.append([f'Proband {variant_label}s with gene(s) associated', len(proband_dict) - len(prob_no_genes)])
    ws_summary.append([f'Proband {variant_label}s "split" ("Proband_split")', len(split_variants)])
    ws_summary.append([f'Proband {variant_label}s "split" by gene (variant-gene pairs)', split_variant_gene_pairs])

    if relative_ids:
        ws_summary.append([])
        append_bold_row(ws_summary, [f'RELATIVES {variant_label}s SUMMARY'])
        append_bold_row(ws_summary, [f'- Relatives {variant_label}s not shared with proband'])
        ws_summary.append([f'Total relatives {variant_label}s not shared with proband', len(relatives_dict)])
        for rel_id in relative_ids:
            rel_not_shared = count_rel_not_shared_by_member(relatives_dict, rel_id)
            ws_summary.append([f'{rel_id} - Not shared proband {variant_label}s', rel_not_shared])

        append_bold_row(ws_summary, [f'- Relatives {variant_label}s shared with proband'])
        ws_summary.append([f'Total relatives {variant_label}s shared with proband', shared_with_relatives])
        for rel_id in relative_ids:
            ws_summary.append([f'{rel_id} - Shared proband {variant_label}s', shared_by_relative.get(rel_id, 0)])

    if somatic_dict:
        ws_summary.append([])
        append_bold_row(ws_summary, [f'TUMOUR {variant_label}s SUMMARY'])
        for tumour_sample, tumour_counts in sorted(tumour_variants.items()):
            ws_summary.append([f'{tumour_sample} - Tumour Only', tumour_counts['TO']])
            ws_summary.append([f'{tumour_sample} - Tumour vs Normal', tumour_counts['TN']])
            ws_summary.append([f'{tumour_sample} - Tumour vs PON', tumour_counts['PON']])
    else:
        ws_summary.append([])
        append_bold_row(ws_summary, ['NOTES'])
        ws_summary.append([f'No tumour samples available for this family'])

    return ws_summary
