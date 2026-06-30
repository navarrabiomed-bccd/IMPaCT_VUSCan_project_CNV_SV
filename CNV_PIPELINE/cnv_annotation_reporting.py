# -*- coding: utf-8 -*-
"""
CNV pipeline - Steps 2 and 3: Priorization and annotation of CNV data to
generate Excel reports.

This pipeline processes CNV (Copy Number Variant) data through multiple
annotation and prioritization steps, comparing results from different tools
(AnnotSV and ClassifyCNV) and integrating multiple data sources to generate
a final Excel report.

Last modification: April 2026
"""


import glob
import logging
import os
import pickle
import csv
import gzip
import re
import sys
from natsort import natsorted
import numpy as np
import openpyxl
import pandas as pd
from intervaltree import IntervalTree
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# Obtener la raíz del proyecto 'VUSCAN' para permitir imports relativos entre carpetas
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

# Importar la utilidad que genera la hoja 'Execution_summary'
from general_scripts.execution_summary_excel import add_summary_sheet


# -------------------------------------------------------------------------
# CONFIGURACIÓN DE CONSTANTES
# -------------------------------------------------------------------------

# Rutas a archivos
BASE_PATH = '/mnt/c/Users/bioinfor/Desktop/VUSCAN_PIPELINES/DATA'
COSMIC_FILE = f'{BASE_PATH}/cancer_genes/tsg_og.tsv'
HEREDITARY_CANCER_FILE = f'{BASE_PATH}/cancer_genes/lista_CH_genes.txt'
#tiers_file = 'C:/Users/edurne.urrutia/Desktop/VUSCAN_PIPELINES/DATA/PropuestaNuevosGenes_VUSCan.xlsx'   ### EHI GENES CANCER
GNOMAD_FILE = f'{BASE_PATH}/gnomad/gnomad.v4.1.cnv.non_neuro_controls_filtered.tsv'
UNIPROT_FILE = f'{BASE_PATH}/pathway_annotation_db/uniprot.txt'
KEGG_FILE = f'{BASE_PATH}/pathway_annotation_db/kegg_pathways.tsv'
REACTOME_FILE = f'{BASE_PATH}/pathway_annotation_db/Gene2Reactome.txt'
SAMPLE_PHENO_FILE = f'{BASE_PATH}/muestras_fenotipo.txt'
GLOSSARY_FILE = f'{BASE_PATH}/glosarios/glosario_cnv.xlsx'
CNAG_FREQ_FILE =  f'{BASE_PATH}/SAMPLES/IMPaCT_families/CNAG/Frequency_VUSCan/cnag_5kb_ref_freq.tsv'
CNAG_FREQ_COLUMN = 'IMPaCT exact counts (5 kbp)'


# Parametros de filtrado
OVERLAP_THRESHOLD = 0.7   # 70% reciprocal overlap
QUAL_MIN = 30             # Minimum quality QUAL
INTERNAL_FREQ_MAX = 10    # Maximum internal frequency count

# Colores para formatos condicionales de Excel
EXCEL_COLORS = {
    'red': PatternFill(bgColor="FFC7CE"),
    'orange': PatternFill(bgColor="FFEB9C"),
    'yellow': PatternFill(bgColor="FFFFCC")
}

# Anotación de ACMG para la clasificación de AnnotSV
ACMG_CLASSIFICATION = {1: 'B', 2: 'LB', 3: 'VUS', 4: 'LP', 5: 'P'}

# Correspondencia de puntuaciones ACMG entre AnnotSV y ClassifyCNV
SCORE_MAPPING = {
    # 1A y 1B comparan con la misma columna '1A-B'
    '1A': '1A-B', '1B': '1A-B',
    # 2A comparación directa
    '2A': '2A',
    # 2C-1 y 2C-2 mapean a la columna '2C'
    '2C-1': '2C', '2C-2': '2C',
    # 2D y subclaves mapean con '2D'
    '2D': '2D', '2D-1': '2D', '2D-3': '2D', '2D-4': '2D',
    # 2E con sufijos comparan con '2E'
    '2E-1': '2E', '2E-2': '2E', '2E-3': '2E', '2E-4': '2E',
    # 2H y 2H-2 mapean con '2H'
    '2H': '2H', '2H-2': '2H',
    # 2F/4O y 40 comparan con '4O'
    '2F/4O': '4O', '4O': '4O',
    # 3A, 3B y 3C comparan con '3'
    '3A': '3', '3B': '3', '3C': '3'
    # 2I-* no existe en ClassifyCNV (discrepancia automática)
    # 5F y 5F-01 se ignoran
}


# -------------------------------------------------------------------------
# LOGGING SETUP
# -------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,  # Nivel mínimo que se mostrará
    format='%(asctime)s - [ %(levelname)s ] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


# -------------------------------------------------------------------------
# CARGA Y GUARDADO DE ARCHIVOS
# -------------------------------------------------------------------------

def save_pickle(data, file_path):
    """Guarda un archivo en formato pickle."""
    with open(file_path, 'wb') as f:
        pickle.dump(data, f)


def load_pickle(file_path):
    """Carga un archivo en formato pickle."""
    with open(file_path, 'rb') as f:
        return pickle.load(f)


def set_csv_field_size():
    """Aumenta de forma segura el limite maximo de tamaño de campo CSV."""
    max_size = sys.maxsize
    while True:
        try:
            csv.field_size_limit(max_size)
            break
        except OverflowError:
            max_size = int(max_size / 10)


# --------------------------------------------------------------
# PROCESAMIENTO DE DATOS DE LAS MUESTRAS Y MAPEO DE IDS
# --------------------------------------------------------------

def split_analysis_folder(folder_path):
    """Extrae nodo y ID de la familia a partir de la ruta de analisis."""
    parts = [p for p in os.path.normpath(folder_path).split(os.sep) if p]
    if len(parts) < 3:
        raise ValueError(f'Ruta de analisis invalida: {folder_path}')

    if parts[-1].upper() != 'CNV':
        raise ValueError(f"La ruta debe terminar en 'CNV': {folder_path}")

    node = parts[-3]
    family_id = parts[-2]
    return node, family_id


def normalize_chromosome(chrom):
    """Normaliza cromosomas para obtener en formato base (sin 'chr')."""
    chrom_str = str(chrom)
    base = chrom_str[3:] if chrom_str.startswith('chr') else chrom_str
    return base


def load_phenotype_data(folder_path, file_path):
    """
    Carga fenotipos de la familia en un diccionario por ID IMPaCT.
    
    Args:
        folder_path (str): Ruta a la carpeta de análisis.
        file_path (str): Ruta al archivo de fenotipos.
    
    Returns:
        dict: Diccionario con los fenotipos de las muestras.
    """
    _, family_id = split_analysis_folder(folder_path)
    
    try:
        df = pd.read_csv(file_path, sep='\t', encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, sep='\t', encoding='ISO-8859-1')

    pattern = f"-{family_id}-"
    df = df[df['Muestra ID'].astype(str).str.contains(pattern)]

    if df.empty:
        raise ValueError('No se encontraron muestras para esta familia')
    
    # Convertir a diccionario (clave: ID IMPaCT)
    pheno_dict = df.set_index('Muestra ID').T.to_dict()
    return pheno_dict


def parse_ped_file(folder_path):
    """
    Lee el archivo PED de la familia y devuelve sexo/estatus por muestra.

    Args:
        folder_path (str): Ruta a la carpeta de la familia.

    Returns:
        dict: Diccionario con la información del PED.
    """
    # Ruta al archivo PED de la familia
    _, family_id = split_analysis_folder(folder_path)
    ped_file = os.path.join(os.path.dirname(os.path.normpath(folder_path)), f"{family_id}.ped")
    
    # Leer archivo PED filtrando filas con 'impact' en el ID
    try:
        ped_data = pd.read_csv(ped_file, sep='\t', encoding='utf-8', dtype=str)
        ped_data = ped_data[ped_data['id'].str.contains('impact', na=False)]
        
        # Crear diccionario con información de sexo y estatus
        ped_dict = {}
        for _, row in ped_data.iterrows():
            try:
                sample_id = row['id'].replace('\u200b', '')  # Elimina caracteres invisibles
                ped_dict[sample_id] = {
                    'sex': int(row['sex']),
                    'status': int(row['aff'])
                }
            except ValueError as ve:
                logging.warning(f'Error al procesar el PED para {sample_id}: {ve}')

        return ped_dict

    except FileNotFoundError:
        logging.error(f'El archivo {ped_file} no se encuentra')
        raise
    except Exception as e:
        logging.error(f'Error al leer el archivo PED: {e}')
        raise


def build_cnag_id_maps(pheno_dict):
    """
    Construye los mapas de conversión entre IDs de muestras de CNAG e IMPaCT.

    Args:
        pheno_dict (dict): Diccionario de fenotipos de las muestras.

    Returns:
        dict (imp_to_cnag): De IMPaCT a CNAG (key: IMPaCT ID, value: CNAG ID)
        dict (cnag_to_imp): De CNAG a IMPaCT (key: CNAG ID, value: IMPaCT ID)
    """
    imp_to_cnag, cnag_to_imp = {}, {}

    for imp_id, data in pheno_dict.items():
        cnag_id = data.get('ID_CNAG')
        # Construir diccionarios con el mapa de IDs
        if cnag_id:
            imp_to_cnag[imp_id] = cnag_id
            cnag_to_imp[cnag_id] = imp_id

    return imp_to_cnag, cnag_to_imp


def get_full_sample_id(short_id, pheno_dict):
    """Obtiene el ID completo de una muestra germinal a partir del ID corto."""
    matches = [
        key for key, val in pheno_dict.items()
        if key.startswith(short_id) and 'ADN germinal' in val.get('Muestra', '')
    ]
    return matches[0] if matches else None


def get_proband_id(folder_path, pheno_dict):
    """Devuelve el ID IMPaCT del probando conn el diccionario de fenotipos."""
    _, family_id = split_analysis_folder(folder_path)
    proband_prefix = f'{family_id}-{family_id}-4impact'
    return get_full_sample_id(proband_prefix, pheno_dict)


# -------------------------------------------------------
# Funciones de solapamiento y comparación entre variantes
# -------------------------------------------------------

def check_reciprocal_overlap(a, b, threshold=OVERLAP_THRESHOLD):
    """
    Comprueba si hay superposición recíproca ≥ umbral entre dos variantes.
    
    Args:
        a, b (dict): Diccionarios con 'SV_start', 'SV_end' y 'SV_length'.
        threshold (float): Umbral mínimo de solapamiento recíproco.
    
    Returns:
        bool: True si hay superposición recíproca ≥ threshold, False si no.
    """
    # Obtener coordenadas de las variantes
    start1, end1 = int(a['SV_start']), int(a['SV_end'])
    start2, end2 = int(b['SV_start']), int(b['SV_end'])
    
    # Calcular la longitud de la superposición
    overlap = max(0, min(end1, end2) - max(start1, start2) + 1)

    # Obtener longitud de las variantes
    len_a = max(abs(float(a.get('SV_length', end1 - start1 + 1))), 1)
    len_b = max(abs(float(b.get('SV_length', end2 - start2 + 1))), 1)

    # Verificar si la superposición recíproca es al menos el umbral
    return (overlap / len_a >= threshold and overlap / len_b >= threshold)


def build_chrom_interval_tree(cnv_dict):
    """
    Construye un índice IntervalTree por cromosoma para CNVs.

    Args:
        cnv_dict (dict): Diccionario de CNVs.

    Returns:
        dict: {cromosoma: IntervalTree con CNVs}
    """
    chrom_index = {}

    for cnv_id, cnv_data in cnv_dict.items():
        try:
            chrom = normalize_chromosome(cnv_data['SV_chrom'])
            start = int(cnv_data['SV_start'])
            end = int(cnv_data['SV_end'])
            length = cnv_data['SV_length']
            cnv_type = cnv_data['SV_type']

            # Corregir inserciones
            if start == end:
                end += 1

            # Crear el índice para el cromosoma, si no existe
            if chrom not in chrom_index:
                chrom_index[chrom] = IntervalTree()

            # Añadir CNV al índice del cromosoma
            chrom_index[chrom][start:end] = (cnv_id, start, end, cnv_type, length)

        except (KeyError, TypeError, ValueError) as e:
            logging.error(f'Error procesando la variante {cnv_id}: {e}')
    
    return chrom_index


# ----------------------------------------------------------------
# PASO 2.1: COMPARACIÓN DE PRIORIZACIÓN (AnnotSV/ClassifyCNV)
# ----------------------------------------------------------------

def find_score_discrepancies(annotsv_scores, classify_scores):
    """Compara criterios ACMG de AnnotSV y ClassifyCNV y devuelve las discrepancias."""
    discrepancies = []
    
    for key, annotsv_score in annotsv_scores.items():
        annotsv_val = float(annotsv_score)
        
        # Tratar no correspondencias
        # 2I-* no existe en ClassifyCNV: discrepancia automática
        if key in ['2I-1', '2I-3']:
            discrepancies.append(key)
            continue
        # 5F y 5F-01 se ignoran
        if key in ['5F', '5F-01']:
            continue
        
        # Mapear a la clasificación de ClassifyCNV
        classify_key = SCORE_MAPPING.get(key, key)
        
        if classify_key not in classify_scores:
            logging.warning(f'Puntuación de AnnotSV pendiente de añadir: {key}, {classify_key} ({annotsv_val})')
            continue
        
        classify_val = float(classify_scores[classify_key])
        
        if classify_val != annotsv_val:
            discrepancies.append(key)
    
    return discrepancies


def compare_cnv_scores(annotsv_file, classify_file, node, sample_id, pheno_dict):
    """
    Compara resultados de AnnotSV y ClassifyCNV, integra campos relevantes y
    registra discrepancias de ACMG.
    
    Args:
        annotsv_file (str): Ruta al archivo preprocesado de AnnotSV.
        classify_file (str): Ruta al archivo de ClassifyCNV.
        node (str): Nodo de secuenciación ('CNAG', 'FPGMX' o 'NASERTIC').
        sample_id (str): ID muestra.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    # Para CNAG, obtener correspondencia de IDs de las muestras
    if node == 'CNAG':
        imp_to_cnag, _ = build_cnag_id_maps(pheno_dict)
    
    # Cargar CNVs con AnnotSV
    annotsv_dict = load_pickle(annotsv_file)

    # Diccionario para guardar la comparativa
    comparison_dict = {}
    
    # Cargar CNVs con ClassifyCNV
    with open(classify_file, 'r') as csvfile:
        reader = csv.DictReader(csvfile, delimiter='\t')
        for row in reader:
            cnv_id = row['VariantID'].lstrip('chr')
            comparison_dict[cnv_id] = {
                'SV_chrom': row['Chromosome'].lstrip('chr'),
                'SV_start': row['Start'],
                'SV_end': row['End'],
                'SV_type': row['Type'],
                'classify_classification': row['Classification'],
                'classify_score': row['Total score'],
                '1A-B': row['1A-B'], '2A': row['2A'], '2B': row['2B'],
                '2C': row['2C'], '2D': row['2D'], '2E': row['2E'],
                '2F': row['2F'], '2G': row['2G'], '2H': row['2H'],
                '2J': row['2J'], '2L': row['2L'], '3': row['3'], '4O': row['4O'],
                'Known or predicted dosage-sensitive genes': row['Known or predicted dosage-sensitive genes']
            }

            # Buscar ID en el diccionario de AnnotSV
            if cnv_id in annotsv_dict:
                entry = annotsv_dict[cnv_id]
    
                # Obtener genotipo (GT) y número de copias (CN)
                if node == 'CNAG':
                    # prob_cnag = imp_to_cnag[sample_id]
                    # smpl_gt = entry[prob_cnag]   # EHI: REVISAR, EN VUSCAN NO HAY ESTA COLUMNA
                    smpl_gt = 'NA'
                    smpl_cn = entry['ControlFreeC_CopyNumber']
                elif node == 'FPGMX':
                    smpl_gt = entry['Zygosity']
                    smpl_cn = entry['CopyNumber']
                elif node == 'NASERTIC':
                    smpl_info = entry.get(sample_id, 'NA')
                    smpl_gt = smpl_info.split(':')[0]
                    smpl_cn = smpl_info.split(':')[2]
                
                comparison_dict[cnv_id][sample_id] = f'GT:{smpl_gt}; CN:{smpl_cn}'
                
                comparison_dict[cnv_id].update({
                    'SV_length': entry['SV_length'],
                    'Genes': entry['Gene_name'],
                    'Gene_count': entry['Gene_count'],
                    'ACMG_annotSV': entry['ACMG_class'],  # ACMG según AnnotSV
                    'OMIM id': entry['OMIM_ID'],
                    'AnnotSV_ranking_score': entry['AnnotSV_ranking_score'],
                    'AnnotSV_ranking_criteria': entry['AnnotSV_ranking_criteria'],
                    'FILTER': 'PASS' if node in ['FPGMX', 'CNAG'] else entry['FILTER'],
                    'QUAL': entry['Qual'] if node == 'FPGMX' else entry['QUAL'] if node == 'NASERTIC' else 'NA',
                    'CNV_window': entry['CNV_window'] if node == 'CNAG' else None
                })

                # Obtener frecuencias internas
                count_prefix = 'Illumina' if node == 'CNAG' else 'Illumina_DRAGEN'
                comparison_dict[cnv_id].update({
                    'Similar counts': entry[f'{count_prefix}.similar.counts'],
                    'Exact counts': entry[f'{count_prefix}.exact.counts'],
                    'Other counts': entry[f'{count_prefix}.other.counts']
                })

                # Añadir los criterios de ClassifyCNV que se cumplen
                classify_keys = ['1A-B', '2A', '2B', '2C', '2D', '2E', '2F',
                                 '2G', '2H', '2I', '2J', '2K', '2L', '3',
                                 '4A', '4B', '4C', '4D', '4E', '4F-H',
                                 '4I', '4J', '4K', '4L', '4M', '4N', '4O',
                                 '5A', '5B', '5C', '5D', '5E', '5F', '5G', '5H']
                
                classify_criteria = []
                for ck in classify_keys:
                    val = row.get(ck, '')
                    try:
                        if float(val) != 0:
                            classify_criteria.append(f'{ck} ({float(val):+.2f})')
                    except ValueError:
                        pass

                comparison_dict[cnv_id]['ClassifyCNV_criteria'] = ";".join(classify_criteria)

                # Buscar discrepancias para cada criterio ACMG
                # 1. Extraer puntuaciones ACMG de AnnotSV
                annotsv_scores = {}

                # Expresión para buscar el carácter + o - seguido de otros 4
                pattern = re.compile(r'([+-]\d\.\d{2})')

                criteria = entry['AnnotSV_ranking_criteria']
                for criterion in criteria.split(';'):
                    match = pattern.search(criterion)
                    if match:
                        key = criterion.split(' ')[0]
                        annotsv_scores[key] = match.group(0)

                # 2. Extraer puntuaciones ACMG de ClassifyCNV
                classify_scores = {k: row.get(k, '0') for k in comparison_dict[cnv_id].keys()
                                   if k in ['1A-B', '2A', '2B', '2C', '2D', '2E', '2F', '2G',
                                            '2H', '2J', '2L', '3', '4O']}
            
                # 3. Comparar puntuaciones de AnnotSV y ClassifyCNV
                discrepancies = find_score_discrepancies(annotsv_scores, classify_scores)
                comparison_dict[cnv_id]['Discrepancies'] = discrepancies
    
    # Reportar solo si hay diferentes CNVs
    annot_ids = set(annotsv_dict.keys())
    classify_ids = set(comparison_dict.keys())
    missing_in_annotsv = classify_ids - annot_ids
    missing_in_classify = annot_ids - classify_ids

    if missing_in_annotsv:
        logging.warning(f"[{sample_id}] {len(missing_in_annotsv)} CNVs en ClassifyCNV pero NO en AnnotSV")
    if missing_in_classify:
        logging.warning(f"[{sample_id}] {len(missing_in_classify)} CNVs en AnnotSV pero NO en ClassifyCNV")

    return comparison_dict


def run_compare_prioritization(folder_path, output_folder, pheno_dict, proband, ped_dict):
    """
    Ejecuta paso 2.1: compara priorización AnnotSV/ClassifyCNV del probando y
    familiares. Guarda los resultados en formato pickle.
    
    Args:
        folder_path (str): Ruta a la carpeta de análisis de la familia (CNV).
        output_folder (str): Ruta a la carpeta para guardar los resultados.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
        proband (str): ID completo del probando.
        ped_dict (dict): Diccionario PED de las muestras.
    
    Returns:
        dict: Diccionario con la comparación de criterios.
    """
    logging.info("PASO 2.1: Comparando priorización (ClassifyCNV vs AnnotSV)")

    node, family_id = split_analysis_folder(folder_path)

    # Procesar probando y guardar
    annotsv_file = os.path.join(folder_path, 'INTERMEDIATES', f'{proband}.CNVs.annotated_parsed.pkl')
    classify_file = os.path.join(folder_path, 'INTERMEDIATES', 'ClassifyCNV', f'{family_id}_Scoresheet.txt')
    proband_comparison = compare_cnv_scores(annotsv_file, classify_file, node, proband, pheno_dict)

    proband_file = os.path.join(output_folder, f'{family_id}_proband_priorization.pkl')
    save_pickle(proband_comparison, proband_file)

    # Procesar familiares
    relatives_dict = {}

    for key in ped_dict.keys():
        if key.split('-')[0] != family_id:
            smpl = key.split('-')[0]
            fam_pattern = os.path.join(folder_path, 'INTERMEDIATES', f'{key}-*.CNVs.annotated_parsed.pkl')
            matches = glob.glob(fam_pattern)

            if not matches:
                logging.warning(f'Falta el archivo de CNVs germinal de la muestra: {key}')
                continue
            
            annotsv_file = matches[0]
            classify_file = os.path.join(folder_path, 'INTERMEDIATES', 'ClassifyCNV', f'{smpl}_Scoresheet.txt')
            rel_id = os.path.basename(annotsv_file).split('.')[0]
            
            relatives_dict[smpl] = compare_cnv_scores(annotsv_file, classify_file, node, rel_id, pheno_dict)
        
    # Guardar CNVs de los familiares solo si existen
    if relatives_dict:
        relatives_file = os.path.join(output_folder, f'{family_id}_relatives_priorization.pkl')
        save_pickle(relatives_dict, relatives_file)
    
    logging.info("Comparación de priorización completada")


# ----------------------------------------------------------------
# PASO 2.2: ANOTACIÓN DE CNVs DEL PROBANDO COMPARTIDAS CON FAMILIARES
# ----------------------------------------------------------------

def add_overlapping_relatives(ped_dict, family_id, folder_path, proband_dict, node, pheno_dict):
    """
    Anota CNVs del probando con solapamientos en familiares, añadiendo
    información de genotipo y estado de cada familiar.

    Args:
        ped_dict (dict): Diccionario PED de las muestras.
        family_id (str): ID de la familia.
        folder_path (str): Ruta a la carpeta de análisis.
        proband_dict (dict): Diccionario con las CNVs del probando.
        node (str): Nodo de secuenciación
        pheno_dict (dict): Diccionario de fenotipos de las muestras.

    Returns:
        dict: Diccionario actualizado con las CNVs de los familiares.
    """
    cnvs_common_path = os.path.join(folder_path, 'INTERMEDIATES', f'{family_id}.CNVs.relatives_overlap_proband.pkl')
    if not os.path.exists(cnvs_common_path):
        logging.info("No hay CNVs de familiares para anotar el probando")
        return proband_dict
    
    cnvs_common = load_pickle(cnvs_common_path)

    # Correspondencia IDs para CNAG
    if node == 'CNAG':
        imp_to_cnag, _ = build_cnag_id_maps(pheno_dict)

    # CNVs de cada familiar
    for rel_full_id, rel_data in cnvs_common.items():
        rel_id = rel_full_id.split('-')[0]

        # Información del familiar
        sex = ped_dict[f'{rel_id}-{family_id}-4impact']['sex']
        status = ped_dict[f'{rel_id}-{family_id}-4impact']['status']

        # Procesar cada CNV
        for row in rel_data.values():
            # Obtener genotipo (GT) y número de copias (CN)
            if node == 'CNAG':
                # rel_cnag = imp_to_cnag[rel_full_id]
                # rel_gt = row.get(rel_cnag, 'NA')   # EHI: REVISAR, EN VUSCAN NO HAY ESTA COLUMNA
                rel_gt = 'NA'
                rel_cn = row['ControlFreeC_CopyNumber']
            elif node == 'FPGMX':
                rel_gt = row['Zygosity']
                rel_cn = row['CopyNumber']
            elif node == 'NASERTIC':
                rel_info = row.get(rel_full_id, 'NA')
                rel_gt = rel_info.split(':')[0]
                rel_cn = rel_info.split(':')[2]

            # Obtener CNV solapante del probando y guardar en el diccionario
            for cnv_id, cnv_data in proband_dict.items():
                if row['SV_chrom'] == cnv_data['SV_chrom'] and row['SV_type'] == cnv_data['SV_type'] and check_reciprocal_overlap(cnv_data, row):
                    proband_dict[cnv_id][rel_id] = {
                        'gt': rel_gt,
                        'cn': rel_cn,
                        'sex': sex,
                        'status': status,
                        'cnv': f"{row['SV_chrom']}_{row['SV_start']}_{row['SV_end']}"
                    }

    return proband_dict


def run_annotate_shared_cnvs(folder_path, output_folder, pheno_dict, ped_dict):
    """
    Ejecuta paso 2.2: Anota CNVs del probando compartidas con familiares
    (solapamiento recíproco). Guarda el resultado en un archivo pickle.

    Args:
        folder_path (str): Ruta a la carpeta de análisis de la familia.
        output_folder (str): Ruta a la carpeta para guardar los resultados.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
        ped_dict (dict): Diccionario PED de las muestras.
    """
    logging.info("PASO 2.2: Anotando CNVs del probando compartidas con familiares")

    node, family_id = split_analysis_folder(folder_path)

    # Cargar CNVs del probando
    proband_file = os.path.join(output_folder, f'{family_id}_proband_priorization.pkl')
    proband_dict = load_pickle(proband_file)

    # Añadir solapamiento con familiares y guardar
    proband_dict = add_overlapping_relatives(ped_dict, family_id, folder_path, proband_dict, node, pheno_dict)

    pickle_file = os.path.join(output_folder, f'{family_id}_proband_sharedcnvs.pkl')
    save_pickle(proband_dict, pickle_file)

    logging.info("Anotación de CNVs compartidas completada")


# ----------------------------------------------------------------
# PASO 2.3: COMBINACIÓN DE CNVs DE LOS FAMILIARES
# ----------------------------------------------------------------

def merge_family_cnvs(rels_dict, family_id, pheno_dict):
    """
    Combina las CNVs de los familiares, agrupando las que solapan entre sí.

    Args:
        rels_dict (dict): CNVs de familiares {sample: {CNV ID: CNV data}}
        family_id (str): ID de la familia.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.

    Returns:
        dict: Diccionario con CNVs solapantes agrupadas por ID.
    """
    merged_cnvs = {}

    for sample_id, cnv_dict in rels_dict.items():
        short_id = f'{sample_id}-{family_id}'
        rel_full_id = get_full_sample_id(short_id, pheno_dict)

        for cnv_id, cnv_data in cnv_dict.items():
            existing_id = None

            # Buscar si esta CNV solapa con alguna ya añadida
            for ref_cnv_id, ref_cnv_data in merged_cnvs.items():
                if (
                    cnv_data.get('SV_chrom') == ref_cnv_data.get('SV_chrom')
                    and cnv_data.get('SV_type') == ref_cnv_data.get('SV_type')
                    and check_reciprocal_overlap(cnv_data, ref_cnv_data)
                ):
                    existing_id = ref_cnv_id
                    break

            # Si no hay solapamiento, añadir la CNV como nueva entrada
            if existing_id is None:
                merged_cnvs[cnv_id] = cnv_data.copy()

            # Si solapa, añadir genotipo/CN del familiar a la CNV existente
            else:
                if rel_full_id in cnv_data:
                    sample_info = cnv_data[rel_full_id]
                    gt = sample_info.split('; ')[0]
                    cn = sample_info.split('; ')[1]
                    merged_cnvs[existing_id][rel_full_id] = f'{gt}; {cn}; CNV:{cnv_id}'

    return merged_cnvs


def run_merge_relatives_cnvs(folder_path, output_folder, pheno_dict):
    """
    Ejecuta paso 2.3: Combina CNVs compartidas de familiares (solapamiento
    recíproco), añade información fenotípica y guarda en un archivo pickle.

    Args:
        folder_path (str): Ruta a la carpeta de analisis de la familia.
        output_folder (str): Ruta a la carpeta para guardar los resultados.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    logging.info("PASO 2.3: Combinando CNVs de familiares")

    _, family_id = split_analysis_folder(folder_path)

    # Cargar CNVs únicas de los familiares
    rels_file_path = os.path.join(output_folder, f'{family_id}_relatives_priorization.pkl')
    if not os.path.exists(rels_file_path):
        logging.info("No hay archivos de familiares, se omite la combinación de CNVs")
        return
    
    rels_dict = load_pickle(rels_file_path)
    
    # Combinar CNVs solapantes entre familiares y guardar solo si hay datos
    merged_cnvs = merge_family_cnvs(rels_dict, family_id, pheno_dict)

    output_file = os.path.join(output_folder, f'{family_id}_relatives_mergedcnvs.pkl')
    save_pickle(merged_cnvs, output_file)
    logging.info(f"Combinadas {len(merged_cnvs)} CNVs de familiares")


# ----------------------------------------------------------------
# PASO 2.4: ANOTACIÓN DE GENES RELACIONADOS CON CÁNCER
# ----------------------------------------------------------------

def load_cosmic(cosmic_file=COSMIC_FILE):
    """
    Carga las anotaciones de genes de COSMIC y los clasifica en un diccionario.

    Args:
        cosmic_file (str): Ruta al archivo COSMIC.

    Returns:
        dict: Genes como claves y tipo (PO, TSG, NA) como valores.
    """
    cosmic_dict = {}

    with open(cosmic_file, mode='r') as file:
        reader = csv.DictReader(file, delimiter='\t')
        for row in reader:
            gene = row['GENE']
            if row['OG'] == 'Yes':
                cosmic_dict[gene] = 'PO'
            elif row['TSG'] == 'Yes':
                cosmic_dict[gene] = 'TSG'
            else:
                cosmic_dict[gene] = 'NA'
    return cosmic_dict


def load_hereditary_cancer(cosmic_dict, hereditary_file=HEREDITARY_CANCER_FILE):
    """
    Carga una lista de genes relacionados con cáncer hereditario y los asocia
    con el diccionario COSMIC.

    Args:
        cosmic_dict (dict): Genes de COSMIC.
        hereditary_file (str): Ruta al archivo con los genes de cáncer hereditario.

    Returns:
        dict: {key: gene, value: PO/TSG/'NA'}
    """
    cancer_dict = {}
    with open(hereditary_file, mode='r', encoding='iso-8859-1') as file:
        reader = csv.reader(file)
        for row in reader:
            gene = row[0]
            cancer_dict[gene] = cosmic_dict.get(gene, 'NA')

    return cancer_dict


def annotate_cancer_genes(cnv_file, cancer_dict, cosmic_dict):
    """Anota genes de cancer hereditario y genes COSMIC para cada CNV."""
    cnv_dict = load_pickle(cnv_file)

    for cnv_id, cnv_data in cnv_dict.items():
        # Inicializar listas si no existen
        cnv_data.setdefault('hereditary cancer gene list', [])
        cnv_data.setdefault('cancer cosmic', [])
        
        # Obtener los genes asociados a la CNV
        genes = cnv_data.get('Genes', '').split(';')

        for gene in genes:
            gene = gene.strip()
            if not gene:
                continue

            # Añadir gen si está en la lista de cáncer hereditario
            if gene in cancer_dict:
                cnv_data['hereditary cancer gene list'].append({
                    'cancer gene': gene,
                    'gene type': cancer_dict[gene]
                })

            # Añadir gen si está en COSMIC
            if gene in cosmic_dict:
                cnv_data['cancer cosmic'].append({
                    'cancer gene': gene,
                    'gene type': cosmic_dict[gene]
                })
                
    return cnv_dict


def run_annotate_cancer_genes(folder_path, output_folder):
    """
    Ejecuta paso 2.4: Anota CNVs de probando y familiares con:
    - Genes anotados como oncogenes (OG) o supresores tumorales (TSG)
    del repositorio COSMIC.
    - Lista de genes relacionados con cancer hereditario (367 genes).

    Args:
        folder_path (str): Ruta a la carpeta de analisis de la familia.
        output_folder (str): Ruta a la carpeta para guardar los resultados.

    Returns:
        tupple: Diccionario probando y diccionario familiares.
    """
    logging.info("PASO 2.4: Anotando CNVs de la familia con genes de cáncer")

    _, family_id = split_analysis_folder(folder_path)

    # Cargar datos de COSMIC y cáncer hereditario
    cosmic_dict = load_cosmic()
    cancer_dict = load_hereditary_cancer(cosmic_dict)

    # Procesar CNVs del probando
    proband_file = os.path.join(output_folder, f'{family_id}_proband_sharedcnvs.pkl')
    proband_dict = annotate_cancer_genes(proband_file, cancer_dict, cosmic_dict)
    
    # Procesar CNVs de familiares
    rels_file = os.path.join(output_folder, f'{family_id}_relatives_mergedcnvs.pkl')
    if os.path.exists(rels_file):
        rels_dict = annotate_cancer_genes(rels_file, cancer_dict, cosmic_dict)
    else:
        rels_dict = {}

    logging.info("Anotación de genes de cáncer completada")
    return proband_dict, rels_dict


# ----------------------------------------------------------------
# PASO 2.5: ANOTACIÓN DE PATHWAYS
# ----------------------------------------------------------------

def load_uniprot_db(uniprot_file=UNIPROT_FILE):
    """Carga anotaciones funcionales de UniProt en un diccionario."""
    uniprot_dict = {}

    with open(uniprot_file, 'r', encoding='utf-8') as f:
        next(f)  # Saltar la cabecera
        for line in f:
            columns = line.strip().split('\t')  # Separar por tabulación
            if len(columns) != 3:
                continue

            gene_names = columns[0].split()  # Algunos genes tienen nombres alternativos
            uniprot_function = columns[2]

            for gene in gene_names:
                if gene in uniprot_dict:
                    uniprot_dict[gene] += ";" + uniprot_function  # Si existe, añadir la nueva ruta
                else:
                    uniprot_dict[gene] = uniprot_function

    return uniprot_dict


def load_kegg_db(kegg_file=KEGG_FILE):
    """Carga un archivo con datos de KEGG en un diccionario."""
    kegg_dict = {}

    with open(kegg_file, "r") as f:
        next(f)
        for line in f:
            columns = line.strip().split("\t")
            if len(columns) != 2:  # Asegurar que la línea tiene dos columnas
                continue

            gene_name, kegg_pathway = columns

            if gene_name in kegg_dict:
                kegg_dict[gene_name] += ";" + kegg_pathway  # Si existe, añadir la nueva ruta
            else:
                kegg_dict[gene_name] = kegg_pathway

    return kegg_dict


def load_reactome_db(reactome_file=REACTOME_FILE):
    """Cargar pathways de Reactome en un diccionario."""
    reactome_dict = {}
    
    with open(reactome_file, 'r', encoding='utf-8') as f:
        next(f)
        for line in f:
            columns = line.strip().split('\t')
            if len(columns) == 2:
                gene_name, reactome_pathway = columns

                if gene_name in reactome_dict:
                    reactome_dict[gene_name] += ";" + reactome_pathway
                else:
                    reactome_dict[gene_name] = reactome_pathway

    return reactome_dict


def add_pathway_annotations(cnv_dict, uniprot_dict, kegg_dict, reactome_dict):
    """
    Añade descripcion funcional de UniProt y pathways KEGG/Reactome de los
    genes asociados a cada CNV.
    
    Args:
        cnv_dict (dict): Diccionario con los genes asociados a cada variante.
        uniprot_dict (dict): Anotaciones funcionales de UniProt.
        kegg_dict (dict): Pathways de KEGG.
        reactome_dict (dict): Pathways de Reactome.
    
    Returns:
        dict: Diccionario actualizado con los campos 'Uniprot_Functions',
              'KEGG_pathways' y 'Reactome_pathways'.
    """
    for cnv_id, cnv_data in cnv_dict.items():
        genes = cnv_data.get('Genes', '').split(';')

        uniprot_function = []
        kegg_pathways = []
        reactome_pathways = []
        
        for gene in genes:
            gene = gene.strip()
            if not gene:
                continue
            
            # Obtener la información de Uniprot, KEGG y Reactome
            function_desc = uniprot_dict.get(gene, 'NA')
            function_desc = function_desc.replace('FUNCTION: ', '')
            uniprot_function.append(f'{gene}: {function_desc}')

            kegg = kegg_dict.get(gene, 'NA')
            kegg_pathways.append(f'{gene}: {kegg}')

            reactome = reactome_dict.get(gene, 'NA')
            reactome_pathways.append(f'{gene}: {reactome}')

        # Añadir la información al diccionario como nuevo campo
        cnv_data['UniProt_Functions'] = '; '.join(uniprot_function)
        cnv_data['KEGG_pathways'] = '; '.join(kegg_pathways)
        cnv_data['Reactome_pathways'] = '; '.join(reactome_pathways)

    return cnv_dict


def run_annotate_pathways(proband_dict, relatives_dict):
    """
    Ejecuta paso 2.5: Anota genes asociados a las CNVs de probando y familiares
    con información funcional de Uniprot y rutas de KEGG/Reactome. Devuelve
    los diccionarios actualizados.

    Args:
        proband_dict (dict) : CNVs del prbando.
        relatives_dict (dict): CNVs de familiares.

    Returns:
        tuple: (proband_dict, relatives_dict) con la anotación añadida.
    """
    logging.info("PASO 2.5: Anotando genes con Uniprot, KEGG y Reactome")

    # Cargar las bases de datos
    uniprot_dict = load_uniprot_db()
    kegg_dict = load_kegg_db()
    reactome_dict = load_reactome_db()

    # Añadir anotación a probando y familiares
    proband_dict = add_pathway_annotations(proband_dict, uniprot_dict, kegg_dict, reactome_dict)
    relatives_dict = add_pathway_annotations(relatives_dict, uniprot_dict, kegg_dict, reactome_dict)

    logging.info("Anotación de genes completada")
    return proband_dict, relatives_dict


# ----------------------------------------------------------------
# PASO 2.6: ANOTACIÓN CON GNOMAD
# ----------------------------------------------------------------

def parse_cnv_id(cnv_id):
    """
    Extrae coordenadas y tipo de CNV a partir del ID de la CNV.
    
    Args:
        cnv_id (str): ID CNV con formato 'chr_start_end_type' (ej. 'chr1_12345_67890_DEL')
    
    Returns:
        tuple: (chrom (str), start (int), end (int), cnvtype (str))
    """
    chrom, start, end, cnvtype = cnv_id.split('_')
    return str(chrom), int(start), int(end), cnvtype


def find_gnomad_match(cnv_id, cnv_data, threshold=OVERLAP_THRESHOLD):
    """
    Verifica si una CNV se solapa de forma recíproca (≥ threshold) con CNVs
    del archivo proporcionado de gnomAD.
    
    Args:
        cnv_id (str): ID de la CNV.
        cnv_data (pd.DataFrame): DataFrame de CNVs de gnomAD.
        threshold (float): Umbral mínimo de solapamiento recíproco.

    Returns:
        dict: Diccionario actualizado con los campos:
            - gnomAD (int): número de CNVs coincidentes
            - gnomAD_ID (str | None): ID de la mejor coincidencia
            - gnomAD_freq (float | None): frecuencia poblacional (SF)
            - gnomAD_length (int | None): longitud de la CNV
    """
    chrom_g, start_g, end_g, cnvtype_g = parse_cnv_id(cnv_id)

    # Filtrar CNVs de gnomAD por cromosoma y tipo
    subset = cnv_data[
        (cnv_data['Chrom'].astype(str) == chrom_g) &
        (cnv_data['Alt'].astype(str) == cnvtype_g)
    ].copy()

    # Calcular solapamiento (vectorizado)
    overlap_start = np.maximum(start_g, subset['Pos'].astype(int))
    overlap_end = np.minimum(end_g, subset['END'].astype(int))
    overlap = np.maximum(0, overlap_end - overlap_start + 1)

    # Calcular el porcentaje de solape recíproco
    length_g = max(1, end_g - start_g + 1)
    length_t = subset['END'].astype(int) - subset['Pos'].astype(int) + 1
    reciprocal = ((overlap / length_g) >= threshold) & ((overlap / length_t) >= threshold)
    
    # Seleccionar coincidencias
    matches = subset[reciprocal]

    if matches.empty:
        return {"gnomAD": 0, "gnomAD_ID": None, "gnomAD_freq": None, "gnomAD_length": None}

    # Filtrar mejor coincidencia (mayor frecuencia poblacional y mayor longitud)
    best_match = matches.sort_values(by=["SF", "SVLEN"], ascending=[False, False]).iloc[0]
    return {
        "gnomAD": len(matches),
        "gnomAD_ID": f"{best_match['Chrom']}_{best_match['Pos']}_{best_match['END']}_{best_match['Alt']}",
        "gnomAD_freq": best_match.get("SF"),
        "gnomAD_length": best_match.get("SVLEN")
    }


def run_annotate_gnomad(folder_path, output_folder, proband_dict, relatives_dict):
    """
    Ejecuta paso 2.6: Anota CNVs de probando y familiares con información
    de gnomAD (solapamiento recíproco). Devuelve el diccionario del probando
    actualizado y guarda el de los familiares.

    Args:
        folder_path (str): Ruta a la carpeta de análisis de la familia.
        output_folder (str): Ruta a la carpeta de salida.
        proband_dict (dict): CNVs del probando.
        relatives_dict (dict): CNVs de los familiares.

    Returns:
        dict: Diccionario actualizado de CNVs del probando
    """
    logging.info("PASO 2.6: Anotando CNVs con gnomAD")

    _, family_id = split_analysis_folder(folder_path)

    # Cargar CNVs de gnomAD
    gnomad_data = pd.read_csv(GNOMAD_FILE, sep='\t')
    
    # Anotar CNVs con la información de gnomAD
    for cnv_id in proband_dict:
        match_info = find_gnomad_match(cnv_id, gnomad_data)
        proband_dict[cnv_id].update(match_info)
    
    for cnv_id in relatives_dict:
        match_info = find_gnomad_match(cnv_id, gnomad_data)
        relatives_dict[cnv_id].update(match_info)

    # Guardar CNVs de familiares
    if relatives_dict:
        relatives_output = os.path.join(output_folder, f'{family_id}_relatives_cnvs_annotated.pkl')
        save_pickle(relatives_dict, relatives_output)

    logging.info("Anotación con gnomAD completada")
    return proband_dict


# ----------------------------------------------------------------
# PASO 2.7: PROCESAMIENTO DE CNVs SOMÁTICAS
# ----------------------------------------------------------------

def add_germline_column(somatic_dict, germinal_dict):
    """
    Anota CNVs somáticas indicando coincidencias con alguna CNV germinal
    (superposición recíproca ≥ umbral y mismo tipo).
    
    Args:
        somatic_dict (dict): CNVs somáticas.
        germinal_dict (dict): CNVs germinales.

    Returns:
        dict: CNVs somáticas con campo 'germline' ('yes'/'no'/'error')
    """
    # Construir IntervalTree para las CNVs germinales
    germinal_index = build_chrom_interval_tree(germinal_dict)

    # Guardar cromosomas somáticos no encontrados en germinal
    missing_chr = set()
    
    # Recorrer cada CNV somática
    for somatic_id, somatic_data in somatic_dict.items():
        try:
            found_match = False  # Para indicar si la CNV está en germinal
            tumor_data = {
                'SV_start': int(somatic_data['SV_start']),
                'SV_end': int(somatic_data['SV_end']),
                'SV_length': float(somatic_data['SV_length']),
                'SV_type': somatic_data['SV_type']
            }

            # Comprobar si el cromosoma está en germinal
            somatic_chrom = normalize_chromosome(somatic_data['SV_chrom'])
            if somatic_chrom in germinal_index:
                overlaps = germinal_index[somatic_chrom]
                
                # Revisar cada CNV germinal que solapa
                for overlap in overlaps:
                    germ_id, germ_start, germ_end, germ_type, germ_length = overlap.data
                    germline_data = {
                        'SV_start': int(germ_start),
                        'SV_end': int(germ_end),
                        'SV_length': float(germ_length),
                        'SV_type': germ_type
                    }

                    # Comprobar CNVs del mismo tipo
                    if somatic_data['SV_type'] == germline_data['SV_type']:
                        if check_reciprocal_overlap(tumor_data, germline_data):
                            found_match = True
                            break  # Detener búsqueda: coincidencia encontrada
            else:
                missing_chr.add(somatic_chrom)
            
            # Asignar resultado a la CNV
            somatic_dict[somatic_id]['germline'] = 'yes' if found_match else 'no'

        except Exception as e:
            logging.error(f"Error procesando CNV somática {somatic_id}: {e}")
            somatic_dict[somatic_id]['germline'] = 'error'
    
    # Mostrar cromosomas somáticos no encontrados en germinal
    if missing_chr:
        logging.warning(f'Cromosomas de somático no presentes en germinal: {", ".join(sorted(missing_chr))}')

    return somatic_dict


def process_somatic_cnv_file(file_path, sample_id, proband_dict, node, exclude_chr_y=False):
    """
    Procesa un archivo somático de CNVs (TSV comprimido o no), selecciona las
    'full', filtra por calidad (NASERTIC y FPGMX) y añade anotación germinal.

    Args:
        file_path (str): Ruta al archivo TSV de CNVs somático (comprimido o no).
        sample_id (str): ID muestra para los logs.
        proband_dict (dict): CNVs germinales del probando.
        node (str): Nodo de secuenciación.
        exclude_chr_y (bool): Excluir cromosoma Y en muestras de mujeres.

    Returns:
        dict | None: CNVs procesadas, o None si el archivo no existe.
    """
    if not os.path.exists(file_path):
        logging.warning(f'Archivo no encontrado: {file_path}.')
        return None

    # Cargar archivo sómatico
    if file_path.endswith('.gz'):
        with gzip.open(file_path, 'rt') as f:
            df = pd.read_csv(f, sep='\t', low_memory=False)
    else:
        df = pd.read_csv(file_path, sep='\t', low_memory=False)

    # Validar columnas necesarias
    required_cols = {'AnnotSV_ID', 'Annotation_mode'}
    missing = required_cols - set(df.columns)
    if missing:
        raise KeyError(f'El archivo {file_path} no contiene las columnas necesarias: {missing}')

    somatic_dict = {}
    for _, row in df.iterrows():
        # Eliminar anotaciones 'split'
        if row['Annotation_mode'] == 'split':
            continue
        
        # Quitar sufijo numérico de 'AnnotSV_ID' y guardar
        base_annot_id = '_'.join(row['AnnotSV_ID'].split('_')[:-1])
        somatic_dict[base_annot_id] = row.to_dict()

    # Filtrar CNVs de NASERTIC y FPGMX por calidad
    if node != 'CNAG':
        somatic_dict = {
            k: v for k, v in somatic_dict.items()
            if v.get('FILTER') == 'PASS' and float(v.get('QUAL', 0)) >= QUAL_MIN
        }

    # Excluir cromosoma mitocondrial
    somatic_dict = {
        k: v for k, v in somatic_dict.items()
        if normalize_chromosome(v.get('SV_chrom', '')) != 'M'
    }

    # Excluir cromosoma Y en muestras de mujeres
    if exclude_chr_y:
        somatic_dict = {
            k: v for k, v in somatic_dict.items()
            if normalize_chromosome(v.get('SV_chrom', '')) != 'Y'
        }
    
    # Anotar si alguna CNV germinal coincide (columna 'germline')
    somatic_dict = add_germline_column(somatic_dict, proband_dict)
    logging.info(f"Procesadas {len(somatic_dict)} CNVs somáticas para {sample_id}")
    return somatic_dict


def check_cnv_in_somatic(germ_cnv_data, somatic_index):
    """
    Comprueba si una CNV germinal solapa recíprocamente con CNVs somáticas.
    
    Args:
        germ_cnv_data (dict): Datos de la CNV germinal.
        somatic_index (dict): IntervalTree por cromosoma de CNVs somáticas.
    
    Returns:
        bool: True si hay superposición recíproca, False si no.
    """
    germ_chrom = normalize_chromosome(germ_cnv_data['SV_chrom'])
    if germ_chrom not in somatic_index:
        return False

    try:
        germ_start = int(germ_cnv_data['SV_start'])
        germ_end = int(germ_cnv_data['SV_end'])
    except (TypeError, ValueError):
        return False

    if germ_start == germ_end:
        germ_end += 1

    germ_cnv_type = germ_cnv_data['SV_type']
    germ_data = {
        'SV_start': germ_start,
        'SV_end': germ_end,
        'SV_length': float(germ_cnv_data.get('SV_length', germ_end - germ_start + 1)),
        'SV_type': germ_cnv_type
    }

    overlaps = somatic_index[germ_chrom].overlap(germ_start, germ_end)
    for overlap in overlaps:
        _, som_start, som_end, som_type, som_length = overlap.data
        if germ_cnv_type != som_type:
            continue

        som_data = {
            'SV_start': som_start,
            'SV_end': som_end,
            'SV_length': som_length,
            'SV_type': som_type
        }

        if check_reciprocal_overlap(germ_data, som_data):
            return True

    return False


def collect_somatic_files(node, somatic_path, pheno_dict):
    """
    Construye la lista de archivos somaticos por muestra y tipo de comparacion.

    Args:
        node (str): Nodo de secuenciación.
        somatic_path (str): Ruta a la carpeta de CNVs somáticos.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.

    Returns:
        list: Lista de tuplas (tumor_id, {kind: path, ...}).
    """
    somatic_files = os.listdir(somatic_path)
    samples = []
    
    if node == 'CNAG':
        _, cnag_to_imp = build_cnag_id_maps(pheno_dict)

        to_files = sorted([f for f in somatic_files if f.endswith('NA.CNVCalls.ann.tag.tsv')])
        for to_file in to_files:
            cnag_id = to_file.split('_vs_')[0]
            impact_id = cnag_to_imp.get(cnag_id)
            tumor_id = '-'.join([impact_id.split('-')[0], impact_id.split('-')[-1]])  # ID muestra y tumor (ej. 1943-02)

            to_path = os.path.join(somatic_path, to_file)
            tn_candidates = [
                fname for fname in somatic_files
                if fname.startswith(f'{cnag_id}_vs_ABA') and fname.endswith('.CNVCalls.ann.tag.tsv')
            ]
            tn_path = os.path.join(somatic_path, sorted(tn_candidates)[0]) if tn_candidates else None
            pon_path = to_path.replace('NA.CNVCalls.ann.tag.tsv', 'Baseline.CNVCalls.ann.tag.tsv')
            samples.append((tumor_id, {'TO': to_path, 'TN': tn_path, 'PON': pon_path}))

    elif node == 'NASERTIC':
        tn_files = sorted([f for f in somatic_files if f.endswith('TN.annotated.tsv.gz')])
        for tn_file in tn_files:
            impact_id = tn_file.split('.')[0]
            tumor_id = "-".join([impact_id.split("-")[0], impact_id.split("-")[-1]])

            tn_path = os.path.join(somatic_path, tn_file)
            to_path = tn_path.replace('TN.annotated.tsv.gz', 'TO.annotated.PASSfiltered.tsv.gz')
            pon_path = tn_path.replace('TN.annotated.tsv.gz', 'PON.annotated.PASSfiltered.tsv.gz')
            samples.append((tumor_id, {'TN': tn_path, 'PON': pon_path, 'TO': to_path}))

    elif node == 'FPGMX':
        to_files = sorted([
            f for f in somatic_files
            if f.endswith('DTO_.nonref.cnv.filtered_annotated.tsv.gz')
        ])
        for to_file in to_files:
            impact_id = to_file.split('.')[0]
            tumor_id = "-".join([impact_id.split("-")[0], impact_id.split("-")[-1]])

            to_path = os.path.join(somatic_path, to_file)
            tn_path = to_path.replace('DTO_.nonref.cnv.filtered_annotated.tsv.gz', 'DTN_.nonref.cnv.annotated.tsv.gz')
            pon_path = to_path.replace('_*DTO_.nonref.cnv.filtered_annotated.tsv.gz', '_D_.nonref.cnv.annotated.tsv.gz')
            samples.append((tumor_id, {'TN': tn_path, 'PON': pon_path, 'TO': to_path}))

    return samples


def run_process_somatic(folder_path, output_folder, proband_dict, pheno_dict):
    """
    Ejecuta paso 2.7: Procesa archivos de CNVs somáticas que guarda en pickles.
    Anota CNVs germinales con la presencia/ausencia en somático y las guarda.

    Args:
        folder_path (str): Ruta a la carpeta con el subdirectorio 'SOMATIC'.
        output_folder (str): Ruta a la carpeta de salida.
        proband_dict (dict): CNVs germinales del probando
        pheno_dict (dict): Diccionario con fenotipos de las muestras.
    """
    logging.info("PASO 2.7: Anotando CNVs germinales del probando con CNVs somáticas")

    node, family_id = split_analysis_folder(folder_path)

    proband_output = os.path.join(output_folder, f'{family_id}_proband_tumour.pkl')

    # Carpeta con los archivos de CNVs somáticas
    if node == 'FPGMX':
        parent_folder = os.path.dirname(os.path.normpath(folder_path))
        somatic_path = os.path.join(parent_folder, 'INPUTS', 'SOMATIC_CNV')
    else:
        somatic_path = os.path.join(folder_path, 'INPUTS', 'SOMATIC')

    # Detectar muestras tumorales esperadas
    expected_tumor_ids = {
        '-'.join([key.split('-')[0], key.split('-')[-1]])
        for key, val in pheno_dict.items()
        if 'ADN tumor' in val.get('Muestra', '')
    }

    if not expected_tumor_ids and not os.path.exists(somatic_path):
        logging.info("Esta familia no tiene muestras tumorales")
        # Asignar 'NA' a la columna 'Tumour' del probando y guardar
        for cnv_id in proband_dict:
            proband_dict[cnv_id]['Tumour'] = 'NA'
        save_pickle(proband_dict, proband_output)
        return

    if expected_tumor_ids and not os.path.exists(somatic_path):
        logging.error(
            f'Se esperan archivos somáticos pero no existe la carpeta: {somatic_path}. '
            f'Se detiene la ejecución.'
        )
        raise SystemExit(1)

    if not expected_tumor_ids and os.path.exists(somatic_path):
        logging.error('Hay carpeta de somático pero no se esperan muestras tumorales. '
                      'Se detiene la ejecución.')
        raise SystemExit(1)

    # Mapa de sexo por ID corto de muestra (1: male, 2: female)
    ped_dict = parse_ped_file(folder_path)
    sex_by_short_id = {
        sample_id.split('-')[0]: sample_info.get('sex')
        for sample_id, sample_info in ped_dict.items()
    }

    # Rutas a los archivos por muestra somática
    samples = collect_somatic_files(node, somatic_path, pheno_dict)

    # Comprobar que todas las muestras tumorales esperadas tienen archivos
    found_tumor_ids = {tumor_id for tumor_id, _ in samples}
    missing_tumors = expected_tumor_ids - found_tumor_ids
    if missing_tumors:
        logging.error(
            f'No se encontraron archivos de las muestras tumorales {", ".join(missing_tumors)}. '
            f'Se detiene la ejecución.'
        )
        raise SystemExit(1)

    # Comprobar que existen todos los archivos somáticos
    required_types = {'TO', 'TN', 'PON'}

    missing_paths = []
    for tumor_id, paths in samples:
        for tumor_type in required_types:
            path = paths.get(tumor_type)
            if not path or not os.path.exists(path):
                missing_paths.append(f'{tumor_id}:{tumor_type}')

    if missing_paths:
        logging.error(
            'Faltan los archivos somáticos: %s. Se detiene la ejecución.',
            ', '.join(sorted(missing_paths))
        )
        raise SystemExit(1)

    # Iniciar diccionarios para IntervalTrees de CNVs somáticas
    to_trees = {}
    tn_trees = {}
    pon_trees = {}
    
    # Procesar archivos somáticos
    somatic_dict = {}

    for tumor_id, paths in samples:
        somatic_dict[tumor_id] = {}
        tumor_short_id = tumor_id.split('-')[0]
        is_female = sex_by_short_id.get(tumor_short_id) == 2
        for tumor_type, path in paths.items():
            # Procesar cada archivo de somático
            if path and os.path.exists(path):
                somatic_data = process_somatic_cnv_file(path, f'{tumor_id}_{tumor_type}', proband_dict,
                                                        node, exclude_chr_y=is_female)

                # Añadir al diccionario
                somatic_dict[tumor_id][f'{tumor_id}_{tumor_type}'] = somatic_data
        
                # Construir IntervalTree
                if somatic_data:
                    if tumor_type == 'TO':
                        to_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
                    elif tumor_type == 'TN':
                        tn_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
                    elif tumor_type == 'PON':
                        pon_trees[tumor_id] = build_chrom_interval_tree(somatic_data)
    
    # Anotar cada CNV germinal con la presencia/ausencia en somático
    for cnv_id, cnv_data in proband_dict.items():
        for tumor_id, tree in to_trees.items():
            proband_dict[cnv_id][f'{tumor_id}_TO'] = 'yes' if check_cnv_in_somatic(cnv_data, tree) else 'no'
        for tumor_id, tree in tn_trees.items():
            proband_dict[cnv_id][f'{tumor_id}_TN'] = 'yes' if check_cnv_in_somatic(cnv_data, tree) else 'no'
        for tumor_id, tree in pon_trees.items():
            proband_dict[cnv_id][f'{tumor_id}_PON'] = 'yes' if check_cnv_in_somatic(cnv_data, tree) else 'no'
    
    # Guardar CNVs anotadas del probando y CNVs somáticas
    save_pickle(proband_dict, proband_output)

    somatic_file = os.path.join(output_folder, f'{family_id}_family_tumours.pkl')
    save_pickle(somatic_dict, somatic_file)

    logging.info("Anotación somática completada")


# ----------------------------------------------------------------
# PASO 2.8: ANOTACIÓN CON SVs
# ----------------------------------------------------------------

def build_chrom_interval_tree_sv(sv_file, node):
    """
    Carga un archivo de SVs, aplica filtros de calidad/frecuencia y
    crea un índice IntervalTree por cromosoma. Filtra las SVs según:
        - Calidad (Qual/QUAL y 'PASS')
        - Frecuencia interna
        - Anotación 'full'
        - Se excluyen CNVs ('GAIN' o 'LOSS') para FPGMX
    
    Args:
        sv_file (str): Ruta al archivo de SVs del probando.
        node (str): Nodo de secuenciación.

    Returns:
        dict: {chrom: IntervalTree con las SVs}
    """
    sv_index = {}

    # Aumentar el límite de tamaño de campo CSV
    set_csv_field_size()

    # Abrir archivo de entrada según el formato
    opener = gzip.open if sv_file.endswith('.gz') else open
    try:
        with opener(sv_file, 'rt') as f:
            reader = csv.DictReader(f, delimiter='\t')

            for row in reader:
                # Filtrar por calidad
                if (float(row['Qual']) if node == 'FPGMX' else float(row['QUAL'])) < QUAL_MIN:
                    continue

                if node == 'CNAG':
                    # Filtrar por 'PASS'
                    if row['FILTER'] != 'PASS':
                        continue
                    # Filtrar por frecuencia interna
                    try:
                        if int(row['Illumina.exact.counts']) >= INTERNAL_FREQ_MAX:
                            continue
                    except ValueError:
                        continue  # Conservar valores no numéricos ('#')
                    # Filtrar CNVs 'split'
                    if row.get('Annotation_mode') == 'split':
                        continue

                elif node == 'FPGMX':
                    # Seleccionar SVs
                    if row['SV_type_original'] == 'GAIN' or row['SV_type_original'] == 'LOSS':
                        continue
                    # Filtrar por frecuencia interna
                    if int(row['Illumina_DRAGEN.exact.counts']) >= INTERNAL_FREQ_MAX:
                        continue

                elif node == 'NASERTIC':
                    # Filtrar por 'PASS'
                    if row['FILTER'] != 'PASS':
                        continue
                    # Filtrar por frecuencia interna
                    if int(row['Illumina_DRAGEN.exact.counts']) >= INTERNAL_FREQ_MAX:
                        continue
                    # Filtrar variantes 'split'
                    if row.get('Annotation_mode') == 'split':
                        continue
                
                chrom = normalize_chromosome(row['SV_chrom'])
                start = int(row['SV_start'])
                end = int(row['SV_end'])
                
                # Corregir casos de intersecciones
                if start == end:
                    end += 1
                # Validar coordenadas
                elif start > end:
                    raise ValueError(f'Intervalo inválido: start > end ({start} > {end})')

                # Crear índice por cromosoma, si no existe
                if chrom not in sv_index:
                    sv_index[chrom] = IntervalTree()

                sv_index[chrom][start:end] = {
                    'SV_start': start,
                    'SV_end': end,
                    'SV_type': row['SV_type'],
                    'sv_id': f"{chrom}_{start}_{end}_{row['SV_type']}"
                }

        return sv_index

    except Exception as e:
        logging.error(f'Error procesando el archivo de SVs: {e}')
        return None


def annotate_cnvs_with_svs(cnv_file, sv_index):
    """
    Anota cada CNV con la primera SV que cumpla solapamiento reciproco.

    Args:
        cnv_file (str): Ruta al archivo con las CNVs.
        sv_index (dict): IntervalTree por cromosoma de SVs.

    Returns:
        dict: CNVs con la anotación de solapamiento con SVs.
    """
    # Cargar CNVs del probando
    cnv_dict = load_pickle(cnv_file)

    match_count = 0

    for cnv_id, cnv_data in cnv_dict.items():
        chrom = normalize_chromosome(cnv_data['SV_chrom'])
        start = int(cnv_data['SV_start'])
        end = int(cnv_data['SV_end'])

        cnv_data['sv_match'] = None  # valor por defecto

        # Buscar SVs en el cromosoma
        if chrom in sv_index:
            overlaps = sv_index[chrom][start:end]
            for overlap in overlaps:
                sv_data = overlap.data

                # Si hay superposición recíproca (≥ umbral), anotar variante
                if check_reciprocal_overlap(cnv_data, sv_data):
                    cnv_data['sv_match'] = sv_data['sv_id']
                    match_count += 1
                    break
    
    logging.info(f"Encontradas {match_count} CNVs con coincidencias de SVs")
    return cnv_dict


def run_annotate_svs_overlap(folder_path, output_folder, pheno_dict):
    """
    Ejecuta paso 2.8: Anota CNVs con solapamiento recíproco de SVs en el
    probando. Guarda el resultado en un archivo pickle.

    Args:
        folder_path (str): Ruta al directorio principal de la familia.
        output_folder (str): Ruta a la carpeta para guardar los resultados.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
    """
    logging.info("PASO 2.8: Anotando CNVs del probando con SVs")

    node, family_id = split_analysis_folder(folder_path)
    
    # Cargar SVs del probando
    folder_up = os.path.dirname(os.path.normpath(folder_path))
    if node == 'CNAG':
        imp_to_cnag, _ = build_cnag_id_maps(pheno_dict)
        matches = [
            key for key, val in pheno_dict.items()
            if key.startswith(f'{family_id}-{family_id}-4impact') and
            'Probando' in val.get('Relación familiar', '') and
            'ADN germinal' in val.get('Muestra', '')
        ]
        prob_impct = matches[0]
        prob_cnag = imp_to_cnag[prob_impct]
        
        pattern_1 = os.path.join(folder_up, 'SV', 'INPUTS', f'{prob_cnag}.ANNOTSV.sv.fixed.tab.gz')
        pattern_2 = os.path.join(folder_up, 'SV', 'INPUTS', f'{prob_cnag}.annotsv.IntFreq.tab.gz')
        pattern_3 = os.path.join(folder_up, 'SV', 'INPUTS', f'{prob_cnag}.manta.annotsv.IntFreq.tab.gz')
        pattern_4 = os.path.join(folder_up, 'SV', 'INPUTS', f'{prob_cnag}.annotsv.fixedV2.tab.gz')  # archivo corregido por CNAG de VUSCAN_03_04 y 06_07
        matches = glob.glob(pattern_1) or glob.glob(pattern_2) or glob.glob(pattern_3) or glob.glob(pattern_4)
        sv_file = matches[0] if matches else None

    elif node == 'FPGMX':
        pattern = os.path.join(folder_up, 'INPUTS', f'{family_id}-{family_id}-4impact-*.split_sv.full.tab.gz')
        matches = glob.glob(pattern)  # mismo archivo para CNVs y SVs
        sv_file = matches[0] if matches else None
        
    elif node == 'NASERTIC':
        pattern_updated = os.path.join(folder_up, 'SV', 'INPUTS', f'{family_id}-{family_id}-4impact-*.SVs.annotated.updated.tsv')
        pattern = os.path.join(folder_up, 'SV', 'INPUTS', f'{family_id}-{family_id}-4impact-*.SVs.annotated.tab')
        matches = glob.glob(pattern_updated) or glob.glob(pattern)
        sv_file = matches[0] if matches else None

    if not sv_file or not os.path.exists(sv_file):
        raise SystemExit(f'[ ERROR ] - Archivo de SV no encontrado: {sv_file}')
    
    # Crear índice IntervalTree por cromosoma de SVs
    sv_index = build_chrom_interval_tree_sv(sv_file, node)
    if not sv_index:
        raise SystemExit(f'[ ERROR ] - No se pudo construir el indice de SVs para {sv_file}')
    
    # Anotar CNVs con las SVs y guardar
    cnv_file = os.path.join(output_folder, f'{family_id}_proband_tumour.pkl')
    annotated_cnvs = annotate_cnvs_with_svs(cnv_file, sv_index)

    output_file = os.path.join(output_folder, f'{family_id}_proband_cnvs_annotated.pkl')
    save_pickle(annotated_cnvs, output_file)

    logging.info("Anotación con SVs completada")


# ----------------------------------------------------------------
# PASO 3: GENERACIÓN DEL INFORME EXCEL
# ----------------------------------------------------------------

def convert_gene_count_to_int(cnv_dict):
    """Normaliza 'Gene_count' a entero; si falta, lo calcula desde 'Genes'."""
    for variant in cnv_dict.values():
        gene_count = variant.get('Gene_count')
        genes = variant.get('Genes', '')

        if gene_count in [None, '']:
            variant['Gene_count'] = 0 if genes == '' else len(genes.split(';'))
            continue

        try:
            variant['Gene_count'] = int(float(gene_count))
        except (TypeError, ValueError):
            annot_id = variant.get('AnnotSV_ID', 'unknown')
            logging.warning(f"No se pudo convertir Gene_count a entero ('{annot_id}')")
            variant['Gene_count'] = 0

    return cnv_dict

def convert_field_to_int(cnv_dict, field):
    """
    Convierte un campo específico a entero de cada variante del diccionario.
    
    Args:
        cnv_dict (dict): Diccionario con las variantes.
        field (str): Nombre del campo para convertir a entero.
    
    Returns:
        dict: Diccionario actualizado con el campo convertido a entero.
    """
    for cnv in cnv_dict.values():
        if field in cnv:
            try:
                cnv[field] = int(float(cnv[field]))
            except ValueError:
                logging.warning(f"No se pudo convertir '{field}' a entero: '{cnv[field]}'")
    return cnv_dict


def translate_status(status):
    """Traduce el estado numérico del estatus a un caracter."""
    status_map = {2: 'affected', 1: 'unaffected'}
    return status_map.get(status, 'NA')


def translate_sex(sex):
    """Traduce el código numérico de la variable sexo a caracter."""
    sex_map = {2: 'female', 1: 'male'}
    return sex_map.get(sex, 'NA')


def load_freq_cnag_5kbp(file_path=CNAG_FREQ_FILE):
    """Carga las frecuencias internas calculadas con las CNVs de CNAG
    de la ventana de 5 kbp de la cohorte de IMPaCT de referencia."""
    if not os.path.exists(file_path):
        raise SystemExit(
            f'[ ERROR ] - No existe el archivo de frecuencia de CNAG (CNVs 5 kbp): {file_path}'
        )

    ref_freq = {}
    with open(file_path, newline='', encoding='utf-8') as freq_file:
        reader = csv.DictReader(freq_file, delimiter='\t')
        for row in reader:
            cnv_id = row.get('CNV_ID', '')
            ref_value = row.get('Reference_exact_count', 'NA')
            total_n = row.get('Reference_total_samples', 'NA')
            ref_freq[cnv_id] = f'{ref_value}/{total_n}'

    return ref_freq


def add_freq_cnag_5kbp_to_cnvs(cnv_dict, ref_freq, column_name=CNAG_FREQ_COLUMN):
    """Añade a cada CNV (CNAG ventana 5 kbp) la frecuencia interna."""
    for cnv_id, cnv_data in cnv_dict.items():
        if cnv_data['CNV_window'] == '5 kbp':
            cnv_data[column_name] = ref_freq.get(cnv_id, 'NA')
        else:
            cnv_data[column_name] = '-'

    return cnv_dict


def build_cnv_row(cnv_id, value, node, sample_headers, cnv_dict, proband=None, pheno_dict=None, somatic_headers=None, acmg_dict=ACMG_CLASSIFICATION):
    """
    Construye la fila de Excel para una CNV de un probando o familiar.
    
    Args:
        cnv_id (str): ID de la CNV.
        value (dict): Datos de la CNV.
        node (str): Nodo de secuenciación.
        sample_headers (list): Encabezados para los genotipos.
        cnv_dict (dict): CNVs para acceder a genotipos y datos somáticos.
        proband (str, optional): ID del probando (Default None).
        pheno_dict (dict, optional): Fenotipos de las muestras (Default None).
        somatic_headers (list, optional): Encabezados somáticos (Default None).
        acmg_dict (dict): Anotación de ACMG para la clasificación de AnnotSV.

    Returns:
        list: Datos de la CNV formateados para una fila de Excel.
    """
    chrom = value.get('SV_chrom', '').lstrip('chr')
    start = value.get('SV_start', '')
    end = value.get('SV_end', '')
    length = value.get('SV_length', '')
    cnv_type = value.get('SV_type', '')
    classify_classification = value.get('classify_classification', '')

    try:
        annot_classification = int(value.get('ACMG_annotSV', ''))
    except (ValueError, TypeError):
        annot_classification = None
    annot_classification = acmg_dict.get(annot_classification, 'VUS')

    discrepancies = ', '.join(value.get('Discrepancies', ''))

    try:
        classify_score = float(value.get('classify_score', 0.0))
    except ValueError:
        classify_score = 0.0

    try:
        annot_score = float(value.get('AnnotSV_ranking_score', 0.0))
    except (ValueError, TypeError):
        annot_score = 0.0
    
    classify_criteria = value.get('ClassifyCNV_criteria', '')
    annot_criteria = value.get('AnnotSV_ranking_criteria', '')
    n_genes = value.get('Gene_count')
    genes = value.get('Genes', '')
    
    # Convertir lista de genes en una cadena de texto
    hereditary = ', '.join([f"{item['cancer gene']} ({item['gene type']})" for item in value.get('hereditary cancer gene list', [])])
    cancer_cosmic = ', '.join([f"{item['cancer gene']} ({item['gene type']})" for item in value.get('cancer cosmic', [])])
    
    dosage_genes = value.get('Known or predicted dosage-sensitive genes', '')
    omim_ids = value.get('OMIM id', '')
    uniprot = value.get('UniProt_Functions', '')
    kegg = value.get('KEGG_pathways', '')
    reactome = value.get('Reactome_pathways', '')
    
    # Genotipo de muestras
    sample_details = []
    for header in sample_headers:
        smpl = header.split(' ')[0]
        # Probando
        if proband:
            if smpl == proband.split('-')[0]:
                sample_details.append(cnv_dict[cnv_id][proband])
            else:
                if smpl in cnv_dict[cnv_id]:
                    sample_info = cnv_dict[cnv_id][smpl]
                    rel = f"GT:{sample_info.get('gt', 'NA')}; CN:{sample_info.get('cn', 'NA')}; CNV:{sample_info.get('cnv', 'NA')}"
                    sample_details.append(rel)
                else:
                    sample_details.append('')

        # Familiares
        else:
            smpl_full = get_full_sample_id(smpl, pheno_dict)
            if not smpl_full:
                sample_details.append('')
                continue

            if smpl_full in cnv_dict[cnv_id]:
                sample_details.append(cnv_dict[cnv_id][smpl_full])
            else:
                sample_details.append('')
    
    # Datos somáticos
    somatic_details = []
    if somatic_headers:
        for sh in somatic_headers:
            if sh in cnv_dict[cnv_id]:
                somatic_details.append(cnv_dict[cnv_id][sh])
            else:
                somatic_details.append('')
    
    exact_counts = value.get('Exact counts', '')
    similar_counts = value.get('Similar counts', '')
    other_counts = value.get('Other counts', '')
    gnomAD = value.get('gnomAD', '')
    gnomAD_ID = value.get('gnomAD_ID', '')
    gnomAD_freq = value.get('gnomAD_freq', '')
    gnomAD_length = value.get('gnomAD_length', '')
    
    row = [cnv_id, chrom, start, end, length, cnv_type, classify_classification, classify_score, classify_criteria,
           annot_classification, annot_score, annot_criteria, discrepancies,
           n_genes, genes, hereditary, cancer_cosmic, dosage_genes, omim_ids,
           uniprot, kegg, reactome] + sample_details + somatic_details + [exact_counts, similar_counts, other_counts]

    if node == 'CNAG':
        row.append(value.get('CNV_window', ''))
        row.append(value.get(CNAG_FREQ_COLUMN, 'NA'))

    row += [gnomAD, gnomAD_ID, gnomAD_freq, gnomAD_length]
    
    if proband:
        row.append(value.get('sv_match', ''))
    
    return row


def apply_conditional_format(ws):
    """Aplica formato condicional a las columnas con los resultados de
    ClassifyCNV y AnnotSV para resaltar CNVs usando reglas de Excel."""
    max_row = ws.max_row

    col_rules = {
        # Clasificación de ClassifyCNV (columna G)
        'G': [("G2=\"Pathogenic\"", EXCEL_COLORS['red']),
              ("G2=\"Likely pathogenic\"", EXCEL_COLORS['orange']),
              ("G2=\"Uncertain significance\"", EXCEL_COLORS['yellow'])],
        # Puntuación de ClassifyCNV (columna H)
        'H': [("H2>=0.99", EXCEL_COLORS['red']),
              ("AND(H2>=0.9,H2<0.99)", EXCEL_COLORS['orange']),
              ("AND(H2>=-0.89,H2<=0.89)", EXCEL_COLORS['yellow'])],
        # Clasificación de AnnotSV (columna J)
        'J': [("J2=\"P\"", EXCEL_COLORS['red']),
              ("J2=\"LP\"", EXCEL_COLORS['orange']),
              ("J2=\"VUS\"", EXCEL_COLORS['yellow'])],
        # Puntuación de AnnotSV (columna K)
        'K': [("K2>=0.99", EXCEL_COLORS['red']),
              ("AND(K2>=0.9,K2<0.99)", EXCEL_COLORS['orange']),
              ("AND(K2>=-0.89,K2<=0.89)", EXCEL_COLORS['yellow'])]
    }
    
    for col, rules in col_rules.items():
        for formula, fill in rules:
            ws.conditional_formatting.add(f'{col}2:{col}{max_row}',
                                          FormulaRule(formula=[formula], fill=fill))


def adjust_columns_width(ws, custom_widths=None, max_columns=None):
    """
    Ajusta automáticamente el ancho de columnas en una hoja Excel.
    
    Args:
        ws: Hoja de openpyxl.
        custom_widths: Diccionario opcional {header_name: width}.
        max_columns: Número máximo de columnas a ajustar.
    """
    for idx, col in enumerate(ws.columns):
        if max_columns is not None and idx >= max_columns:
            break

        max_length = 0
        column_letter = get_column_letter(col[0].column)
        header = col[0].value

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                continue

        # Ancho por defecto con un margen adicional
        adjusted_width = (max_length + 2)

        # Sobreescribir anchos específicos si se proporcionan
        if custom_widths and header in custom_widths:
            adjusted_width = custom_widths[header]
        
        # Ajustar columnas de genotipos por patrón
        if header and ('Status' in header):
            adjusted_width = 60

        # Aplicar cambios a la hoja de Excel
        ws.column_dimensions[column_letter].width = adjusted_width


def add_dict_to_excel_sheet(wb, data_dict, sheet_name):
    """Inserta un diccionario de CNVs como nueva hoja de Excel."""
    if not data_dict:
        return
    
    # Crear nueva hoja de Excel
    ws = wb.create_sheet(title=sheet_name)

    # Obtener las claves del diccionario para usarlas como encabezados
    headers = list(next(iter(data_dict.values())).keys())
    headers = headers[1:]  # Quitar 'AnnotSV_ID'
    ws.append(['CNV ID'] + headers)  # Agregar encabezado

    # Agregar filas con los valores
    for annot_id, values in data_dict.items():
        row = [annot_id] + [', '.join(map(str, values[col])) if isinstance(values[col], list) else values[col] for col in headers]
        ws.append(row)


def annotate_split_cnvs_hereditary(split_dict, hereditary_file=HEREDITARY_CANCER_FILE):
    """Añade columna 'Hereditary cancer genes' a cada fila de CNVs 'split'."""
    with open(hereditary_file, mode='r', encoding='iso-8859-1') as f:
        hereditary_genes = {row[0].strip() for row in csv.reader(f) if row}

    for annot_id, rows in split_dict.items():
        new_rows = []
        for row in rows:
            gene = row['Gene_name'].strip()
            new_row = {}
            for key, value in row.items():
                new_row[key] = value
                if key == 'Gene_name':
                    new_row['Hereditary cancer genes'] = gene if gene in hereditary_genes else ''
            new_rows.append(new_row)
        split_dict[annot_id] = new_rows

    return split_dict


def create_excel_report(folder_path, proband_dict, relatives_dict, ped_dict, pheno_dict,
                        prob_split_dict, output_file, somatic_dict, proband):
    """
    Genera un archivo Excel con CNVs anotadas en varias hojas: 'Proband' (CNVs
    probando), 'Proband_split' (CNVs 'split' probando), 'Relatives' (CNVs
    familiares), '{IDrelative}_split' (CNVs 'split' familiares) y datos somáticos.
    """
    node, family_id = split_analysis_folder(folder_path)

    # Para CNAG, añadir la frecuencia interna a las CNVs de la ventana de 5 kbp
    if node == 'CNAG':
        ref_freq = load_freq_cnag_5kbp()
        proband_dict = add_freq_cnag_5kbp_to_cnvs(proband_dict, ref_freq)
        relatives_dict = add_freq_cnag_5kbp_to_cnvs(relatives_dict, ref_freq)
        logging.info(f"Frecuencia interna de la ventana de 5 kbp añadida a {len(ref_freq)} CNVs")

    # ---------- CREAR Y GUARDAR ARCHIVO EXCEL ----------
    # Crear un nuevo libro de Excel con el glosario (mantiene formato y estilos)
    wb = openpyxl.load_workbook(GLOSSARY_FILE)
    logging.info("Archivo Excel con el glosario creado.")

    # Añadir hoja con el resumen
    ws_summary = add_summary_sheet(wb, node, family_id, ped_dict, pheno_dict,
                                   proband_dict, prob_split_dict, relatives_dict, somatic_dict,
                                   variant_label='CNV')
    logging.info("Hoja de resumen añadida al Excel")
    
    # Crear nueva hoja para el probando
    ws_proband = wb.create_sheet('Proband')

    # ---------- 'Proband': ENCABEZADOS ----------
    # Obtener IDs de las muestras para las columnas de genotipo
    sample_headers = []
    for key in ped_dict.keys():
        smpl = key.split('-')[0]  # ID muestra
        smpl_id = get_full_sample_id(key, pheno_dict)  # ID completo

        if not smpl_id:
            logging.warning(f'No hay muestra germinal para {key}')
            continue
        
        # Añadir información de fenotipo
        status = translate_status(ped_dict[key].get('status', 'NA'))
        if status == 'affected':
            pheno = pheno_dict[smpl_id]['Subpatología']
            status = f'{status} ({pheno})'
        sex = translate_sex(ped_dict[key].get('sex', 'NA'))
        kinship = pheno_dict[smpl_id]['Relación familiar']

        if smpl == family_id:
            sample_headers.append((int(smpl), f'{smpl} - proband (Status: {status}, Sex: {sex})'))
        else:
            sample_headers.append((int(smpl), f'{smpl} (Status: {status}, Sex: {sex}, Kinship: {kinship})'))

    # Ordenar los encabezados de las muestras
    sample_headers = sorted(sample_headers, key=lambda x: x[0])
    sample_headers = [header[1] for header in sample_headers]

    # Eliminar duplicados manteniendo el orden
    sample_headers = list(dict.fromkeys(sample_headers))
    
    # Obtener IDs de las muestras tumorales
    somatic_headers = []
    for smpl in somatic_dict:
        somatic_headers.extend(somatic_dict[smpl].keys())

    # Encabezados del probando
    common_headers = ['CNV ID', 'Chrom', 'Start', 'End', 'Length', 'Type', 
                      'ClassifyCNV classification', 'ClassifyCNV Score', 'ClassifyCNV criteria',
                      'AnnotSV classification', 'AnnotSV score', 'AnnotSV criteria', 'Discrepancies',
                      'Gene count', 'Genes', 'Hereditary cancer genes', 'Cancer Cosmic', 
                      'Known/Predicted dosage-sensitive genes', 'OMIM', 'UniProt functions', 'KEGG pathways', 'Reactome pathways']
    proband_headers = somatic_headers.copy()
    proband_headers += ['Exact counts', 'Similar counts', 'Other counts']
    if node == 'CNAG':
        proband_headers.append('CNV_window')
        proband_headers.append(CNAG_FREQ_COLUMN)
    proband_headers += ['gnomAD', 'gnomAD_ID', 'gnomAD_freq', 'gnomAD_length', 'Match in SVs']
    ws_proband.append(common_headers + sample_headers + proband_headers)

    # ---------- 'Proband': INFORMACIÓN CNVs DEL PROBANDO ----------
    # Ordenar CNVs del probando
    sorted_cnvs = sorted(
        proband_dict.keys(),
        key=lambda k: (
            -float(proband_dict[k].get('AnnotSV_ranking_score', '0') or '0'),  # Ordenar por AnnotSV score
            -float(proband_dict[k].get('classify_score', '0') or '0'),         # Ordenar por Classify score
            -float(proband_dict[k].get('SV_length', '0') or '0')               # Ordenar por longitud
        )
    )

    # Agregar datos del probando al archivo Excel
    for cnv_id in sorted_cnvs:
        value = proband_dict[cnv_id]
        row = build_cnv_row(cnv_id=cnv_id, value=value, node=node, sample_headers=sample_headers,
                            cnv_dict=proband_dict, proband=proband, somatic_headers=somatic_headers)
        ws_proband.append(row)

    logging.info("Hoja del probando añadida al Excel")

    # ---------- 'Proband_split': CREAR HOJA Y AÑADIR ENCABEZADOS ----------
    # Obtener los encabezados
    excluded_fields = ['AnnotSV_ID', 'Samples_ID', 'Closest_left', 'Closest_right', 'Gene_count', 
                       'RE_gene', 'po_P_gain_phen', 'po_P_gain_hpo', 'po_P_gain_source', 
                       'po_P_gain_coord', 'po_P_gain_percent', 'po_P_loss_phen', 'po_P_loss_hpo', 
                       'po_P_loss_source', 'po_P_loss_coord', 'po_P_loss_percent', 
                       'po_B_gain_allG_source', 'po_B_gain_allG_coord', 'po_B_gain_someG_source', 
                       'po_B_gain_someG_coord', 'po_B_loss_allG_source', 'po_B_loss_allG_coord', 
                       'po_B_loss_someG_source', 'po_B_loss_someG_coord', 'GC_content_left', 
                       'GC_content_right', 'Repeat_coord_left', 'Repeat_type_left', 'Repeat_coord_right', 
                       'Repeat_type_right', 'Gap_left', 'Gap_right', 'SegDup_left', 'SegDup_right', 
                       'ENCODE_blacklist_left', 'ENCODE_blacklist_characteristics_left', 
                       'ENCODE_blacklist_right', 'ENCODE_blacklist_characteristics_right', 
                       'AnnotSV_ranking_score', 'AnnotSV_ranking_criteria', 'ACMG_class']
    
    prob_split_dict = annotate_split_cnvs_hereditary(prob_split_dict)
    sample_split_headers = list(next(iter(prob_split_dict.values()))[0].keys())
    split_headers = ['AnnotSV ID'] + [key for key in sample_split_headers if key not in excluded_fields]

    # Crear nueva hoja en el Excel y agregar encabezados
    ws_prob_split = wb.create_sheet(title='Proband_split')
    ws_prob_split.append(split_headers)
    
    # ---------- 'Proband_split': CNVs 'SPLIT' ----------
    # Ordenar CNVs por cromosoma y posición
    sorted_prob_split = {}
    for cnv_id in natsorted(prob_split_dict.keys(), key=lambda id: (prob_split_dict[id][0]['SV_chrom'], int(prob_split_dict[id][0]['SV_start']))):
        sorted_prob_split[cnv_id] = prob_split_dict[cnv_id]
    
    # Añadir los datos fila por fila
    for cnv_id, rows in sorted_prob_split.items():
        for row in rows:
            row_data = [cnv_id] + [value for key, value in row.items() if key not in excluded_fields]
            ws_prob_split.append(row_data)
    logging.info("Hoja con las CNVs 'split' del probando añadida al Excel")

    # ---------- 'Relatives': CREAR HOJA Y AÑADIR ENCABEZADOS ----------
    ws_relatives = None
    sample_headers_rel = sample_headers[1:]  # Excluir probando
    if relatives_dict:
        ws_relatives = wb.create_sheet(title='Relatives')
        rel_headers = common_headers + sample_headers_rel + ['Exact counts', 'Similar counts', 'Other counts']
        if node == 'CNAG':
            rel_headers.append('CNV_window')
            rel_headers.append(CNAG_FREQ_COLUMN)
        rel_headers += ['gnomAD', 'gnomAD_ID', 'gnomAD_freq', 'gnomAD_length']
        ws_relatives.append(rel_headers)

        # ---------- 'Relatives': INFORMACIÓN CNVs ----------
        # Ordenar CNVs
        rel_sorted_cnvs = sorted(
            relatives_dict.keys(),
            key=lambda k: (
                -float(relatives_dict[k].get('AnnotSV_ranking_score', '0') or '0'),  # Ordenar por AnnotSV score
                -float(relatives_dict[k].get('classify_score', '0') or '0'),         # Ordenar por ClassifyCNV score
                -float(relatives_dict[k].get('SV_length', '0') or '0')               # Ordenar por longitud
            )
        )

        # Agregar datos de cada familiar al Excel
        for cnv_id in rel_sorted_cnvs:
            value = relatives_dict[cnv_id]
            row = build_cnv_row(cnv_id=cnv_id, value=value, node=node, sample_headers=sample_headers_rel,
                                cnv_dict=relatives_dict, pheno_dict=pheno_dict)
            ws_relatives.append(row)
        logging.info("Hoja de los familiares añadida al Excel")
    
        # ---------- 'Relatives_split': CNVs 'split' ----------
        for rel in ped_dict:
            rel_smpl = rel.split('-')[0]
            if rel_smpl == family_id:
                continue

            pattern = os.path.join(folder_path, 'INTERMEDIATES', f'{rel_smpl}-{family_id}-4impact-*.CNVs.annotated_parsed_split.pkl')
            matches = glob.glob(pattern)
            if not matches:
                logging.warning(f"Falta el archivo de CNVs 'split' de {rel_smpl}")
                continue

            rel_split_dict = load_pickle(matches[0])
            rel_split_dict = annotate_split_cnvs_hereditary(rel_split_dict)

            # Crear nueva hoja en el Excel y agregar encabezados 
            ws_rel_split = wb.create_sheet(title=f'{rel_smpl}_split')
            ws_rel_split.append(split_headers)

            # Ordenar CNVs por cromosoma y posición
            sorted_rels_split = {}
            for cnv_id in natsorted(rel_split_dict.keys(), key=lambda id: (rel_split_dict[id][0]['SV_chrom'], int(rel_split_dict[id][0]['SV_start']))):
                sorted_rels_split[cnv_id] = rel_split_dict[cnv_id]
            
            # Añadir datos al Excel
            for cnv_id, rows in sorted_rels_split.items():
                for row in rows:
                    row_data = [cnv_id] + [value for key, value in row.items() if key not in excluded_fields]
                    ws_rel_split.append(row_data)

        logging.info("Hojas con las CNVs 'split' de los familiares añadidas al Excel")
    else:
        logging.info("No hay familiares, no se crean hojas de familiares en el Excel")


    # ---------- SOMATIC: CNVs ----------
    for smpl, sample_somatic_dict in somatic_dict.items():
        for somatic_id, somatic_data in sample_somatic_dict.items():
            if somatic_data is not None:
                sheet_name = ''
                if 'TO' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_Only'
                elif 'TN' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_vs_Normal'
                elif 'PON' in somatic_id:
                    sheet_name = f'{smpl}_Tumour_vs_PON'
    
                if sheet_name:
                    sorted_somatic = {}
                    for cnv_id in natsorted(
                        somatic_data.keys(),
                        key=lambda variant_id: (
                            somatic_data[variant_id]['SV_chrom'],
                            int(somatic_data[variant_id]['SV_start'])
                        )
                    ):
                        sorted_somatic[cnv_id] = somatic_data[cnv_id]

                    add_dict_to_excel_sheet(wb, sorted_somatic, sheet_name)
    if somatic_dict:
        logging.info("Hojas con las CNVs somáticas añadidas al Excel")
    
    # ---------- FORMATO CONDICIONAL ----------
    # Resaltar CNVs de 'Proband' y 'Relatives' según puntuación de ClassifyCNV y AnnotSV
    apply_conditional_format(ws_proband)
    if ws_relatives is not None:
        apply_conditional_format(ws_relatives)

    # ---------- AJUSTAR ANCHO DE COLUMNA ----------
    # Definir anchos personalizados para columnas específicas
    custom_widths = {
        'ClassifyCNV criteria': 18,
        'AnnotSV classification': 20,
        'AnnotSV criteria': 18,
        'Genes': 20,
        'Hereditary cancer genes': 22,
        'Cancer Cosmic': 15,
        'Known/Predicted dosage-sensitive genes': 35,
        'OMIM': 15,
        'UniProt functions': 25, 'KEGG pathways': 25, 'Reactome pathways': 25,
        'Other counts': 25,
        'Match in SVs': 26
    }

    # Ajustar ancho de las columnas de 'Proband' y 'Relatives'
    adjust_columns_width(ws_proband, custom_widths)
    if ws_relatives is not None:
        adjust_columns_width(ws_relatives, custom_widths)

    # Ajustar las primeras 5 columnas de las hojas 'split' y somáticas
    for ws in wb.worksheets:
        if ws.title.endswith('_split'):
            adjust_columns_width(ws, max_columns=5)
        if any(suffix in ws.title for suffix in ['_Tumour_Only', '_Tumour_vs_Normal', '_Tumour_vs_PON']):
            adjust_columns_width(ws, max_columns=5, custom_widths={'SV_length': 15})

    # Ajustar ancho de las primeras columnas del resumen
    adjust_columns_width(ws_summary, max_columns=5)
    
    # ---------- GUARDAR EL ARCHIVO EXCEL ----------
    wb.save(output_file)
    logging.info(f"Archivo Excel completado y guardado en: {output_file}")


def run_generate_excel_report(folder_path, output_folder, pheno_dict, proband, ped_dict):
    """
    Ejecuta paso 3: genera un informe Excel con CNVs anotadas del probando y
    familiares. Integra datos somáticos, germinales, genes de cáncer y otras
    anotaciones con glosario.
    
    Args:
        folder_path (str): Ruta al directorio principal de la familia.
        output_folder (str): Ruta a la carpeta para guardar el informe.
        pheno_dict (dict): Diccionario de fenotipos de las muestras.
        proband (str): ID completo del probando.
        ped_dict (dict): Diccionario PED de las muestras.
    """
    _, family_id = split_analysis_folder(folder_path)

    # Cargar CNVs del probando y familiares
    proband_file = os.path.join(output_folder, f'{family_id}_proband_cnvs_annotated.pkl')
    proband_dict = load_pickle(proband_file)

    relatives_file = os.path.join(output_folder, f'{family_id}_relatives_cnvs_annotated.pkl')
    if os.path.exists(relatives_file):
        relatives_dict = load_pickle(relatives_file)
    else:
        relatives_dict = {}

    # Convertir a número entero 'Gene_count' y 'SV_length'
    proband_dict = convert_gene_count_to_int(proband_dict)
    relatives_dict = convert_gene_count_to_int(relatives_dict)

    proband_dict = convert_field_to_int(proband_dict, 'SV_length')
    relatives_dict = convert_field_to_int(relatives_dict, 'SV_length')

    # Cargar CNVs 'split' del probando (creado en script 'CNV_parse.py')
    pattern = os.path.join(folder_path, 'INTERMEDIATES', f'{family_id}-{family_id}-4impact-*.CNVs.annotated_parsed_split.pkl')
    matches = glob.glob(pattern)
    if not matches:
        raise SystemExit('[ ERROR ] - No se encontró el archivo "split" del probando')
    
    proband_split_dict = load_pickle(matches[0])

    # Archivo de CNVs somáticas (ausencia implica que no hay tumor, se validó en el paso 2.7)
    somatic_file = os.path.join(output_folder, f'{family_id}_family_tumours.pkl')
    if os.path.exists(somatic_file):
        somatic_dict = load_pickle(somatic_file)
    else:
        somatic_dict = {}
        logging.info("Familia sin muestras tumorales. No se incluirán hojas somáticas en el Excel.")
    
    # Crear y guardar el archivo Excel
    output_file = os.path.join(output_folder, f'{family_id}_prioritized_CNVs.xlsx')
    create_excel_report(folder_path, proband_dict, relatives_dict, ped_dict, pheno_dict,
                        proband_split_dict, output_file, somatic_dict, proband)


# ---------- FUNCIÓN GLOBAL ----------
def main(folder_path):
    """
    Función principal que orquesta la ejecucion completa de las pasos 2 y 3
    del algoritmo de CNVs.

    Args:
        folder_path (str): Ruta a la carpeta de análisis de la familia.
    """
    # Configurar carpeta para guardar los resultados
    output_folder = os.path.join(folder_path, 'OUTPUTS/')
    if os.path.exists(output_folder):
        logging.warning(f'La carpeta de salida ya existe: {output_folder}')
    else:
        os.makedirs(output_folder, exist_ok=True)
        logging.info(f"Creada carpeta de salida {output_folder}")

    # Cargar archivos fenotipos y PED de las muestras
    pheno_dict = load_phenotype_data(folder_path, SAMPLE_PHENO_FILE)
    ped_dict = parse_ped_file(folder_path)

    # ID del probando
    proband = get_proband_id(folder_path, pheno_dict)

    # Información de inicio del paso 2
    logging.info("INICIANDO EL PROCESO DE ANOTACIÓN Y PRIORIZACIÓN DE CNVs (PASO 2)")
    logging.info(f"Carpeta de análisis: {folder_path}")
    logging.info(f"Carpeta de salida: {output_folder}")

    # Paso 2.1: Comparación de priorización (AnnotSV/ClassifyCNV)
    run_compare_prioritization(folder_path, output_folder, pheno_dict, proband, ped_dict)
    # Paso 2.2: Anotación de CNVs del probando compartidas con familiares
    run_annotate_shared_cnvs(folder_path, output_folder, pheno_dict, ped_dict)
    # Paso 2.3: Combinación de CNVs de los familiares
    run_merge_relatives_cnvs(folder_path, output_folder, pheno_dict)
    # Paso 2.4: Anotación de genes relacionados con cáncer
    proband_dict, rels_dict = run_annotate_cancer_genes(folder_path, output_folder)
    # Paso 2.5: Anotación de pathways
    proband_dict, rels_dict = run_annotate_pathways(proband_dict, rels_dict)
    # Paso 2.6: Anotación con gnomAD
    proband_dict = run_annotate_gnomad(folder_path, output_folder, proband_dict, rels_dict)
    # Paso 2.7: Procesamiento de CNVs somáticas
    run_process_somatic(folder_path, output_folder, proband_dict, pheno_dict)
    # Paso 2.8: Anotación con SVs
    run_annotate_svs_overlap(folder_path, output_folder, pheno_dict)
    logging.info("ANOTACIÓN Y PRIORIZACIÓN DE CNVs COMPLETADA")
    
    # Paso 3: Generación del informe Excel
    logging.info("INICIO DE LA GENERACIÓN DEL INFORME FINAL (PASO 3)")
    run_generate_excel_report(folder_path, output_folder, pheno_dict, proband, ped_dict)
    logging.info("GENERACIÓN DEL INFORME FINAL COMPLETADA")
    
    logging.info("ANÁLISIS CNVs FINALIZADO")


if __name__ == '__main__':
    # folder_path = 'C:/Users/edurne.urrutia/Desktop/VUSCAN_PIPELINES/DATA/SAMPLES/CNAG/1282/CNV/'
    WORKDIR = sys.argv[1]
    folder_path = f'{WORKDIR}CNV/'
    
    if not os.path.isdir(folder_path):
        logging.error('La ruta introducida no es válida o no existe.')
    else:
        main(folder_path)
